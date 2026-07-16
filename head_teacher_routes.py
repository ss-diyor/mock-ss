import csv
import io
import secrets
import openpyxl
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from db import get_pool
from auth import get_current_head_teacher, FRONTEND_BASE_URL
from groups_db import DEFAULT_MAX_GROUPS_PER_CENTER, DEFAULT_MAX_STUDENTS_PER_CENTER
from scoring import get_band_score
from branding import BrandingUpdateIn, branding_payload, validate_branding

router = APIRouter(prefix="/api/head-teacher", tags=["head-teacher"])


class GroupCreateIn(BaseModel):
    name: str

class GenerateTeacherInviteIn(BaseModel):
    expires_in_hours: int = 168  # default 7 days


# ─── Markaz haqida umumiy ma'lumot ─────────────────────────────────────────

@router.get("/center")
async def get_center(current_user: dict = Depends(get_current_head_teacher)):
    db = await get_pool()
    async with db.acquire() as conn:
        center = await conn.fetchrow(
            """
            SELECT id, name, organization_type, slug, is_active, max_groups, max_students, created_at,
                   brand_name, brand_primary_color, brand_secondary_color, brand_logo_url,
                   brand_favicon_url, brand_contact_email, brand_contact_phone, show_powered_by,
                   directory_opt_in, directory_admin_override, directory_description,
                   directory_region, directory_address, directory_website_url,
                   directory_telegram_url, directory_instagram_url, directory_show_email,
                   directory_gallery_urls, directory_gallery_captions,
                   directory_show_phone, directory_show_address, directory_show_statistics,
                   directory_show_testimonials
            FROM centers WHERE id=$1
            """,
            current_user["center_id"]
        )
        if not center:
            raise HTTPException(status_code=404, detail="Markaz topilmadi")
        groups_count = await conn.fetchval("SELECT COUNT(*) FROM groups WHERE center_id=$1", center["id"])
        students_count = await conn.fetchval(
            "SELECT COUNT(*) FROM users WHERE center_id=$1 AND role='student'", center["id"]
        )

    return {
        "id": center["id"],
        "name": center["name"],
        "organization_type": center["organization_type"],
        "is_active": center["is_active"],
        "created_at": center["created_at"].isoformat(),
        "groups_count": groups_count,
        "students_count": students_count,
        "max_groups": center["max_groups"] or DEFAULT_MAX_GROUPS_PER_CENTER,
        "max_students": center["max_students"] or DEFAULT_MAX_STUDENTS_PER_CENTER,
        "branding": branding_payload(center),
    }


@router.put("/center/branding")
async def update_center_branding(
    data: BrandingUpdateIn,
    current_user: dict = Depends(get_current_head_teacher)
):
    values = validate_branding(data)
    db = await get_pool()
    async with db.acquire() as conn:
        if values["slug"]:
            duplicate = await conn.fetchval(
                "SELECT 1 FROM centers WHERE LOWER(slug)=LOWER($1) AND id<>$2",
                values["slug"], current_user["center_id"]
            )
            if duplicate:
                raise HTTPException(status_code=409, detail="Bu slug boshqa tashkilot tomonidan band qilingan")
        row = await conn.fetchrow(
            """
            UPDATE centers SET
                brand_name=$1, slug=$2, brand_primary_color=$3, brand_secondary_color=$4,
                brand_logo_url=$5, brand_favicon_url=$6, brand_contact_email=$7,
                brand_contact_phone=$8, show_powered_by=$9,
                directory_opt_in=$10, directory_description=$11, directory_region=$12,
                directory_address=$13, directory_website_url=$14, directory_telegram_url=$15,
                directory_instagram_url=$16, directory_gallery_urls=$17,
                directory_gallery_captions=$18, directory_show_email=$19,
                directory_show_phone=$20, directory_show_address=$21,
                directory_show_statistics=$22, directory_show_testimonials=$23
            WHERE id=$24
            RETURNING id, name, organization_type, slug, brand_name, brand_primary_color,
                      brand_secondary_color, brand_logo_url, brand_favicon_url,
                      brand_contact_email, brand_contact_phone, show_powered_by,
                      directory_opt_in, directory_admin_override, directory_description,
                      directory_region, directory_address, directory_website_url,
                      directory_telegram_url, directory_instagram_url, directory_gallery_urls,
                      directory_gallery_captions,
                      directory_show_email,
                      directory_show_phone, directory_show_address, directory_show_statistics,
                      directory_show_testimonials
            """,
            values["brand_name"], values["slug"], values["brand_primary_color"],
            values["brand_secondary_color"], values["brand_logo_url"], values["brand_favicon_url"],
            values["brand_contact_email"], values["brand_contact_phone"], values["show_powered_by"],
            values["directory_opt_in"], values["directory_description"], values["directory_region"],
            values["directory_address"], values["directory_website_url"], values["directory_telegram_url"],
            values["directory_instagram_url"], values["directory_gallery_urls"],
            values["directory_gallery_captions"],
            values["directory_show_email"],
            values["directory_show_phone"], values["directory_show_address"],
            values["directory_show_statistics"], values["directory_show_testimonials"],
            current_user["center_id"]
        )
    return {"message": "Brending saqlandi", "branding": branding_payload(row)}


# ─── Guruhlar ───────────────────────────────────────────────────────────────

@router.get("/groups")
async def list_groups(current_user: dict = Depends(get_current_head_teacher)):
    db = await get_pool()
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT g.id, g.name, g.invite_code, g.is_active, g.created_at,
                   g.teacher_id, g.teacher_invite_code, g.teacher_invite_expires_at,
                   t.full_name AS teacher_name, t.email AS teacher_email,
                   COUNT(u.id) FILTER (WHERE u.role = 'student') AS students_count
            FROM groups g
            LEFT JOIN users t ON t.id = g.teacher_id
            LEFT JOIN users u ON u.group_id = g.id
            WHERE g.center_id = $1
            GROUP BY g.id, t.full_name, t.email
            ORDER BY g.created_at DESC
            """,
            current_user["center_id"]
        )
    return [
        {
            "id": r["id"], "name": r["name"], "invite_code": r["invite_code"],
            "is_active": r["is_active"], "created_at": r["created_at"].isoformat(),
            "has_teacher": r["teacher_id"] is not None,
            "teacher_name": r["teacher_name"], "teacher_email": r["teacher_email"],
            "students_count": r["students_count"],
            "teacher_invite_code": r["teacher_invite_code"],
            "teacher_invite_expires_at": r["teacher_invite_expires_at"].isoformat() if r["teacher_invite_expires_at"] else None,
        }
        for r in rows
    ]


@router.post("/groups")
async def create_group(data: GroupCreateIn, current_user: dict = Depends(get_current_head_teacher)):
    name = data.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Guruh nomini kiriting")

    db = await get_pool()
    async with db.acquire() as conn:
        async with conn.transaction():
            center = await conn.fetchrow(
                "SELECT is_active, max_groups FROM centers WHERE id=$1 FOR UPDATE",
                current_user["center_id"]
            )
            if not center or not center["is_active"]:
                raise HTTPException(status_code=400, detail="Markaz faol emas")

            max_groups = center["max_groups"] or DEFAULT_MAX_GROUPS_PER_CENTER
            current_count = await conn.fetchval(
                "SELECT COUNT(*) FROM groups WHERE center_id=$1", current_user["center_id"]
            )
            if current_count >= max_groups:
                raise HTTPException(status_code=400, detail=f"Guruhlar limiti ({max_groups}) to'lgan")

            invite_code = secrets.token_urlsafe(6)
            row = await conn.fetchrow(
                """
                INSERT INTO groups (name, invite_code, center_id)
                VALUES ($1, $2, $3)
                RETURNING id, name, invite_code, is_active, created_at
                """,
                name, invite_code, current_user["center_id"]
            )

    return {
        "id": row["id"], "name": row["name"], "invite_code": row["invite_code"],
        "is_active": row["is_active"], "created_at": row["created_at"].isoformat(),
    }


async def _own_group_or_404(conn, group_id: int, center_id: int):
    group = await conn.fetchrow("SELECT * FROM groups WHERE id=$1", group_id)
    if not group or group["center_id"] != center_id:
        raise HTTPException(status_code=404, detail="Guruh topilmadi")
    return group


@router.post("/groups/{group_id}/generate-teacher-invite")
async def generate_teacher_invite(group_id: int, data: GenerateTeacherInviteIn, current_user: dict = Depends(get_current_head_teacher)):
    db = await get_pool()
    async with db.acquire() as conn:
        group = await _own_group_or_404(conn, group_id, current_user["center_id"])
        if group["teacher_id"]:
            raise HTTPException(
                status_code=400,
                detail="Bu guruhda allaqachon teacher bor — avval uni olib tashlang"
            )

        invite_code = secrets.token_urlsafe(24)  # yuqori entropiya — bu rol beruvchi token
        await conn.execute(
            "UPDATE groups SET teacher_invite_code=$1, teacher_invite_expires_at=NOW() + ($2 || ' hours')::interval WHERE id=$3",
            invite_code, str(data.expires_in_hours), group_id
        )

    return {
        "teacher_invite_code": invite_code,
        "invite_link": f"{FRONTEND_BASE_URL}/register?teacher_invite={invite_code}",
    }


@router.post("/groups/{group_id}/revoke-teacher-invite")
async def revoke_teacher_invite(group_id: int, current_user: dict = Depends(get_current_head_teacher)):
    db = await get_pool()
    async with db.acquire() as conn:
        await _own_group_or_404(conn, group_id, current_user["center_id"])
        await conn.execute("UPDATE groups SET teacher_invite_code = NULL, teacher_invite_expires_at = NULL WHERE id = $1", group_id)
    return {"message": "Teacher taklif kodi bekor qilindi"}


@router.post("/groups/{group_id}/remove-teacher")
async def remove_teacher(group_id: int, current_user: dict = Depends(get_current_head_teacher)):
    db = await get_pool()
    async with db.acquire() as conn:
        async with conn.transaction():
            group = await _own_group_or_404(conn, group_id, current_user["center_id"])
            if not group["teacher_id"]:
                raise HTTPException(status_code=400, detail="Bu guruhda teacher yo'q")

            await conn.execute(
                "UPDATE users SET role='student', center_id=NULL WHERE id=$1", group["teacher_id"]
            )
            await conn.execute(
                "UPDATE groups SET teacher_id=NULL, teacher_invite_code=NULL WHERE id=$1", group_id
            )

    return {"message": "Teacher guruhdan olib tashlandi, hisobi student sifatida saqlanib qoldi"}


@router.post("/groups/{group_id}/deactivate")
async def toggle_group(group_id: int, current_user: dict = Depends(get_current_head_teacher)):
    db = await get_pool()
    async with db.acquire() as conn:
        await _own_group_or_404(conn, group_id, current_user["center_id"])
        row = await conn.fetchrow(
            "UPDATE groups SET is_active = NOT is_active WHERE id=$1 RETURNING is_active", group_id
        )
    return {
        "message": "Guruh yopildi" if not row["is_active"] else "Guruh qayta ochildi",
        "is_active": row["is_active"]
    }


# ─── Bulk Excel/CSV import (roster) ─────────────────────────────────────────

ROSTER_IMPORT_HEADERS = ("email", "full_name")
ROSTER_IMPORT_MAX_BYTES = 5 * 1024 * 1024
ROSTER_IMPORT_MAX_ROWS = 2000


def _read_roster_rows(filename: str, raw: bytes) -> list[dict]:
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if suffix == "csv":
        try:
            content = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            content = raw.decode("cp1251", errors="ignore")
        reader = csv.DictReader(io.StringIO(content))
        headers = [str(value or "").strip().lower() for value in (reader.fieldnames or [])]
        if "email" not in headers:
            raise HTTPException(status_code=400, detail="CSV faylida 'email' ustuni topilmadi")
        return [{(key or "").strip().lower(): value for key, value in row.items()} for row in reader]
    if suffix == "xlsx":
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            sheet = workbook["Oquvchilar"] if "Oquvchilar" in workbook.sheetnames else workbook.active
            values = sheet.iter_rows(values_only=True)
            headers = [str(value or "").strip().lower() for value in next(values, [])]
            if "email" not in headers:
                raise HTTPException(status_code=400, detail="Excel faylida 'email' ustuni topilmadi")
            return [dict(zip(headers, row)) for row in values]
        except HTTPException:
            raise
        except Exception:
            raise HTTPException(status_code=400, detail="Excel faylini o'qib bo'lmadi")
    raise HTTPException(status_code=400, detail="Faqat .xlsx yoki .csv fayl qabul qilinadi")


@router.get("/groups/import-template")
async def group_roster_import_template(current_user: dict = Depends(get_current_head_teacher)):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Oquvchilar"
    sheet.append(list(ROSTER_IMPORT_HEADERS))
    sheet.append(["student@example.com", "O'quvchi Ismi"])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = "A1:B2"
    sheet.column_dimensions["A"].width = 32
    sheet.column_dimensions["B"].width = 28
    payload = io.BytesIO()
    workbook.save(payload)
    payload.seek(0)
    return StreamingResponse(
        payload,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="group-students-import.xlsx"'},
    )

@router.post("/groups/{group_id}/import-students")
async def import_students(
    group_id: int,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_head_teacher)
):
    raw = await file.read(ROSTER_IMPORT_MAX_BYTES + 1)
    if len(raw) > ROSTER_IMPORT_MAX_BYTES:
        raise HTTPException(status_code=413, detail="Fayl hajmi 5 MB dan oshmasligi kerak")
    rows = _read_roster_rows(file.filename or "", raw)
    rows = [row for row in rows if any(str(value or "").strip() for value in row.values())]
    if len(rows) > ROSTER_IMPORT_MAX_ROWS:
        raise HTTPException(status_code=400, detail=f"Bir importda ko'pi bilan {ROSTER_IMPORT_MAX_ROWS} o'quvchi")

    db = await get_pool()
    async with db.acquire() as conn:
        await _own_group_or_404(conn, group_id, current_user["center_id"])

        added, skipped, invalid = 0, 0, 0
        async with conn.transaction():
            for raw_row in rows:
                norm = {(k or "").strip().lower(): str(v or "").strip() for k, v in raw_row.items()}
                email = norm.get("email", "").lower()
                full_name = norm.get("full_name") or norm.get("full name") or None
                if not email or "@" not in email:
                    invalid += 1
                    continue
                result = await conn.execute(
                    """
                    INSERT INTO group_roster_emails (group_id, email, full_name)
                    VALUES ($1, $2, $3)
                    ON CONFLICT (group_id, email) DO NOTHING
                    """,
                    group_id, email, full_name
                )
                if result == "INSERT 0 1":
                    added += 1
                else:
                    skipped += 1

    return {"added": added, "skipped_duplicates": skipped, "invalid_rows": invalid}


@router.get("/groups/{group_id}/roster")
async def get_roster(group_id: int, current_user: dict = Depends(get_current_head_teacher)):
    db = await get_pool()
    async with db.acquire() as conn:
        await _own_group_or_404(conn, group_id, current_user["center_id"])
        rows = await conn.fetch(
            """
            SELECT email, full_name, used, created_at
            FROM group_roster_emails WHERE group_id=$1
            ORDER BY created_at DESC
            """,
            group_id
        )
    return [
        {"email": r["email"], "full_name": r["full_name"], "used": r["used"], "created_at": r["created_at"].isoformat()}
        for r in rows
    ]


# ─── Statistika ──────────────────────────────────────────────────────────────

@router.get("/stats")
async def head_teacher_stats(current_user: dict = Depends(get_current_head_teacher)):
    db = await get_pool()
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT er.section, er.score, er.total, er.writing_band
            FROM exam_results er
            JOIN users u ON u.email = er.email
            WHERE u.center_id = $1 AND u.role = 'student'
            """,
            current_user["center_id"]
        )

    bands_by_section = {}
    for r in rows:
        if r["section"] == "writing":
            band = r["writing_band"]
        else:
            band = get_band_score(r["score"], r["total"], r["section"]) if r["score"] is not None and r["total"] else None
        if band is not None:
            bands_by_section.setdefault(r["section"], []).append(float(band))

    averages = {section: round(sum(vals) / len(vals), 1) for section, vals in bands_by_section.items()}
    weakest_section = min(averages, key=averages.get) if averages else None

    return {
        "average_band_by_section": averages,
        "weakest_section": weakest_section,
        "total_attempts": len(rows),
    }


@router.get("/export/results")
async def export_center_results(format: str = "excel", group_id: Optional[int] = None, current_user: dict = Depends(get_current_head_teacher)):
    db = await get_pool()
    async with db.acquire() as conn:
        if group_id:
            await _own_group_or_404(conn, group_id, current_user["center_id"])
            group_filter = "AND u.group_id = $2"
            args = (current_user["center_id"], group_id)
        else:
            group_filter = ""
            args = (current_user["center_id"],)

        rows = await conn.fetch(
            f"""
            SELECT er.id, u.full_name, u.email, er.section, er.score, er.total, er.writing_band,
                   er.writing_task_achievement, er.writing_coherence_cohesion, er.writing_lexical_resource, er.writing_grammar_accuracy,
                   er.grader_name, er.writing_graded_at, er.writing_feedback, er.submitted_at,
                   g.name AS group_name
            FROM exam_results er
            JOIN users u ON u.email = er.email
            LEFT JOIN groups g ON g.id = u.group_id
            WHERE u.center_id = $1 {group_filter}
            ORDER BY er.submitted_at DESC
            """,
            *args
        )
    from result_export import build_results_export
    return build_results_export(rows, format, "markaz-natijalari")


@router.get("/trend")
async def head_teacher_trend(days: int = 30, current_user: dict = Depends(get_current_head_teacher)):
    db = await get_pool()
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT DATE(er.submitted_at) AS date, g.name AS group_name, g.id AS group_id,
                   COUNT(er.id) AS submissions,
                   AVG(
                       COALESCE(
                           er.writing_band,
                           CASE 
                               WHEN er.section = 'listening' THEN (SELECT band FROM (VALUES (0,1), (40,9)) AS b(score, band) LIMIT 1) -- sodda o'rniga, DB dan emas python da hisoblash yaxshiroq
                               ELSE NULL
                           END
                       )
                   ) AS avg_band_raw
            FROM exam_results er
            JOIN users u ON u.email = er.email
            JOIN groups g ON g.id = u.group_id
            WHERE u.center_id = $1 AND er.submitted_at >= NOW() - ($2 || ' days')::interval
            GROUP BY DATE(er.submitted_at), g.name, g.id
            ORDER BY date ASC
            """,
            current_user["center_id"], str(days)
        )
        
        # SQL da barcha bandlarni hisoblash qiyin bo'lgani uchun, soddaroq yo'l:
        # Barcha qatorlarni olib pythonda guruhlaymiz
        raw_results = await conn.fetch(
            """
            SELECT DATE(er.submitted_at) AS date, g.name AS group_name, g.id AS group_id,
                   er.section, er.score, er.total, er.writing_band
            FROM exam_results er
            JOIN users u ON u.email = er.email
            JOIN groups g ON g.id = u.group_id
            WHERE u.center_id = $1 AND er.submitted_at >= NOW() - ($2 || ' days')::interval
            """,
            current_user["center_id"], str(days)
        )

    # Pythonda trendni hisoblash
    from collections import defaultdict
    import datetime
    
    # Kuni -> Guruh ID -> { count, total_band, band_count }
    trends = defaultdict(lambda: defaultdict(lambda: {"count": 0, "total_band": 0, "band_count": 0, "group_name": ""}))
    group_names = {}
    
    for r in raw_results:
        d = r["date"].isoformat()
        gid = r["group_id"]
        group_names[gid] = r["group_name"]
        
        trends[d][gid]["count"] += 1
        trends[d][gid]["group_name"] = r["group_name"]
        
        if r["section"] == "writing":
            band = r["writing_band"]
        else:
            band = get_band_score(r["score"], r["total"], r["section"]) if r["score"] is not None and r["total"] else None
            
        if band is not None:
            trends[d][gid]["total_band"] += float(band)
            trends[d][gid]["band_count"] += 1

    # Formatlash
    dates = sorted(list(trends.keys()))
    datasets = {}
    
    for gid, gname in group_names.items():
        datasets[gid] = {
            "group_name": gname,
            "group_id": gid,
            "submissions": [],
            "avg_bands": []
        }
        
    for d in dates:
        for gid in group_names.keys():
            stats = trends[d][gid]
            datasets[gid]["submissions"].append(stats["count"])
            if stats["band_count"] > 0:
                datasets[gid]["avg_bands"].append(round(stats["total_band"] / stats["band_count"], 1))
            else:
                datasets[gid]["avg_bands"].append(0)

    # Overall center trend
    overall = {
        "group_name": "Umumiy Markaz",
        "group_id": "overall",
        "submissions": [],
        "avg_bands": []
    }
    for d in dates:
        day_subs = sum(trends[d][gid]["count"] for gid in group_names.keys())
        day_total_band = sum(trends[d][gid]["total_band"] for gid in group_names.keys())
        day_band_count = sum(trends[d][gid]["band_count"] for gid in group_names.keys())
        
        overall["submissions"].append(day_subs)
        overall["avg_bands"].append(round(day_total_band / day_band_count, 1) if day_band_count > 0 else 0)

    return {
        "labels": dates,
        "datasets": list(datasets.values()) + [overall]
    }


@router.get("/leaderboard")
async def head_teacher_leaderboard(
    group_id: Optional[int] = None,
    limit: int = 50,
    current_user: dict = Depends(get_current_head_teacher)
):
    """Markaz rahbari markaz yoki guruh bo'yicha reyting jadvalini ko'radi."""
    from scoring import get_band_score, calculate_overall_band
    db = await get_pool()
    async with db.acquire() as conn:
        if group_id:
            await _own_group_or_404(conn, group_id, current_user["center_id"])
            where = "WHERE u.center_id = $1 AND u.group_id = $2 AND u.role = 'student'"
            args = (current_user["center_id"], group_id)
        else:
            where = "WHERE u.center_id = $1 AND u.role = 'student'"
            args = (current_user["center_id"],)

        rows = await conn.fetch(
            f"""
            SELECT
                u.id, u.full_name, u.email, g.name AS group_name,
                MAX(CASE WHEN er.section = 'listening' THEN er.score END) AS l_score,
                MAX(CASE WHEN er.section = 'listening' THEN er.total END) AS l_total,
                MAX(CASE WHEN er.section = 'reading'   THEN er.score END) AS r_score,
                MAX(CASE WHEN er.section = 'reading'   THEN er.total END) AS r_total,
                MAX(CASE WHEN er.section = 'writing'   THEN er.writing_band END) AS w_band,
                MAX(CASE WHEN er.section = 'speaking'  THEN er.speaking_band END) AS s_band
            FROM users u
            LEFT JOIN exam_results er ON er.email = u.email
            LEFT JOIN groups g ON g.id = u.group_id
            {where}
            GROUP BY u.id, u.full_name, u.email, g.name
            LIMIT {limit}
            """,
            *args
        )

    result = []
    for row in rows:
        l_band = get_band_score(row["l_score"], row["l_total"], "listening") if row["l_score"] is not None else None
        r_band = get_band_score(row["r_score"], row["r_total"], "reading") if row["r_score"] is not None else None
        w_band = float(row["w_band"]) if row["w_band"] is not None else None
        s_band = float(row["s_band"]) if row["s_band"] is not None else None
        all_bands = [b for b in [l_band, r_band, w_band, s_band] if b is not None]
        overall = calculate_overall_band(all_bands) if all_bands else None
        result.append({
            "id": row["id"], "full_name": row["full_name"], "email": row["email"],
            "group_name": row["group_name"],
            "listening_band": l_band, "reading_band": r_band,
            "writing_band": w_band, "speaking_band": s_band,
            "overall_band": overall
        })

    result.sort(key=lambda x: x["overall_band"] or 0, reverse=True)
    for i, item in enumerate(result):
        item["rank"] = i + 1
    return {"leaderboard": result, "total": len(result)}
