from fastapi import Depends, FastAPI, Header, HTTPException, UploadFile, File, Form
from feature_routes import router as feature_router
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, Response, StreamingResponse
from pydantic import BaseModel
from typing import Optional
import os
import json
import httpx
import asyncio
import io
import base64
import csv
import hmac
import re
try:
    import openpyxl
except ImportError:
    openpyxl = None
from datetime import datetime
from fpdf import FPDF

from db import get_pool
from scoring import get_band_score
from auth import router as auth_router, ensure_users_table, get_current_user
from groups_db import ensure_center_group_tables, DEFAULT_MAX_GROUPS_PER_CENTER, DEFAULT_MAX_STUDENTS_PER_CENTER
from head_teacher_routes import router as head_teacher_router
from teacher_routes import router as teacher_router
from school_routes import router as school_router
from school_staff_routes import router as school_staff_router
from billing_routes import router as billing_router
from test_catalog_routes import router as test_catalog_router, _can_access_test
from test_builder_routes import router as test_builder_router
from branding import ORGANIZATION_TYPES, SLUG_RE, branding_payload
from notification_center import (
    router as notification_router,
    ensure_notification_tables,
    fetch_notifications,
    mark_notification_read,
    mark_all_notifications_read,
    create_notification,
    notify_admin,
)
from support_routes import router as support_router, ensure_support_tables

app = FastAPI(title="IELTS Mock SS")
app.include_router(feature_router)
app.include_router(auth_router)
app.include_router(head_teacher_router)
app.include_router(teacher_router)
app.include_router(school_router)
app.include_router(school_staff_router)
app.include_router(billing_router)
app.include_router(test_catalog_router)
app.include_router(test_builder_router)
app.include_router(notification_router)
app.include_router(support_router)


@app.middleware("http")
async def no_cache_api(request, call_next):
    response = await call_next(request)
    if request.url.path.startswith("/api/"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
        response.headers["Pragma"] = "no-cache"
    return response

ADMIN_SECRET = os.environ.get("ADMIN_SECRET")
MIN_ADMIN_SECRET_LENGTH = 32

# Email konfiguratsiyasi (Resend)
RESEND_API_KEY = os.environ.get("RESEND_API_KEY")
EMAIL_FROM = os.environ.get("EMAIL_FROM", "noreply@ielts.sultanov.space")


@app.on_event("startup")
async def startup():
    db = await get_pool()
    async with db.acquire() as conn:
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS exam_results (
                id SERIAL PRIMARY KEY,
                full_name TEXT NOT NULL,
                email TEXT NOT NULL,
                section TEXT NOT NULL,
                score INTEGER,
                total INTEGER,
                answers JSONB,
                writing_task1 TEXT,
                writing_task2 TEXT,
                duration_seconds INTEGER,
                submitted_at TIMESTAMP DEFAULT NOW()
            )
        """)
        await conn.execute("ALTER TABLE exam_results ADD COLUMN IF NOT EXISTS writing_band NUMERIC")
        await conn.execute("ALTER TABLE exam_results ADD COLUMN IF NOT EXISTS writing_feedback TEXT")
        await conn.execute("ALTER TABLE exam_results ADD COLUMN IF NOT EXISTS notified BOOLEAN DEFAULT FALSE")
        await conn.execute("ALTER TABLE exam_results ADD COLUMN IF NOT EXISTS writing_task_achievement NUMERIC")
        await conn.execute("ALTER TABLE exam_results ADD COLUMN IF NOT EXISTS writing_coherence_cohesion NUMERIC")
        await conn.execute("ALTER TABLE exam_results ADD COLUMN IF NOT EXISTS writing_lexical_resource NUMERIC")
        await conn.execute("ALTER TABLE exam_results ADD COLUMN IF NOT EXISTS writing_grammar_accuracy NUMERIC")
        await conn.execute("ALTER TABLE exam_results ADD COLUMN IF NOT EXISTS writing_graded_at TIMESTAMP")
        await conn.execute("ALTER TABLE exam_results ADD COLUMN IF NOT EXISTS grader_name TEXT")
        # Speaking bo'limi uchun ustunlar
        await conn.execute("ALTER TABLE exam_results ADD COLUMN IF NOT EXISTS speaking_telegram_file_id TEXT")
        await conn.execute("ALTER TABLE exam_results ADD COLUMN IF NOT EXISTS speaking_band NUMERIC")
        await conn.execute("ALTER TABLE exam_results ADD COLUMN IF NOT EXISTS speaking_feedback TEXT")
        await conn.execute("ALTER TABLE exam_results ADD COLUMN IF NOT EXISTS speaking_graded_at TIMESTAMP")
        await conn.execute("ALTER TABLE exam_results ADD COLUMN IF NOT EXISTS speaking_fluency_coherence NUMERIC")
        await conn.execute("ALTER TABLE exam_results ADD COLUMN IF NOT EXISTS speaking_lexical_resource NUMERIC")
        await conn.execute("ALTER TABLE exam_results ADD COLUMN IF NOT EXISTS speaking_grammar_accuracy NUMERIC")
        await conn.execute("ALTER TABLE exam_results ADD COLUMN IF NOT EXISTS speaking_pronunciation NUMERIC")
        await conn.execute("ALTER TABLE exam_results ADD COLUMN IF NOT EXISTS speaking_audio_data BYTEA")
        await conn.execute("ALTER TABLE exam_results ADD COLUMN IF NOT EXISTS speaking_audio_mime TEXT")
        await conn.execute("ALTER TABLE exam_results ADD COLUMN IF NOT EXISTS speaking_audio_filename TEXT")


        await conn.execute("""
            CREATE TABLE IF NOT EXISTS exam_sessions (
                id SERIAL PRIMARY KEY,
                full_name TEXT NOT NULL,
                email TEXT NOT NULL,
                started_at TIMESTAMP DEFAULT NOW(),
                sections_completed TEXT[] DEFAULT ARRAY[]::TEXT[]
            )
        """)

    await ensure_users_table()
    await ensure_center_group_tables()
    async with db.acquire() as conn:
        await ensure_notification_tables(conn)
        await ensure_support_tables(conn)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS feedback_templates (
                id SERIAL PRIMARY KEY,
                teacher_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                section TEXT NOT NULL CHECK(section IN ('writing','speaking')),
                title TEXT NOT NULL,
                content TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT NOW()
            )
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS resubmission_requests (
                id SERIAL PRIMARY KEY,
                result_id INTEGER NOT NULL REFERENCES exam_results(id) ON DELETE CASCADE,
                student_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                teacher_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
                section TEXT NOT NULL CHECK(section IN ('writing','speaking')),
                reason TEXT NOT NULL,
                due_at TIMESTAMP,
                status TEXT NOT NULL DEFAULT 'pending',
                replacement_result_id INTEGER REFERENCES exam_results(id) ON DELETE SET NULL,
                created_at TIMESTAMP DEFAULT NOW(),
                UNIQUE(result_id, status)
            )
        """)
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS speaking_timed_comments (
                id SERIAL PRIMARY KEY,
                result_id INTEGER NOT NULL REFERENCES exam_results(id) ON DELETE CASCADE,
                teacher_id INTEGER REFERENCES users(id) ON DELETE SET NULL,
                timestamp_seconds NUMERIC NOT NULL CHECK(timestamp_seconds >= 0),
                comment TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT NOW()
            )
        """)
        await conn.execute("ALTER TABLE exam_results ADD COLUMN IF NOT EXISTS test_id INTEGER REFERENCES tests(id) ON DELETE SET NULL")
        await conn.execute("ALTER TABLE exam_results ADD COLUMN IF NOT EXISTS test_slug TEXT")
        await conn.execute("ALTER TABLE exam_results ADD COLUMN IF NOT EXISTS test_mode TEXT")
        await conn.execute("ALTER TABLE exam_sessions ADD COLUMN IF NOT EXISTS test_id INTEGER REFERENCES tests(id) ON DELETE SET NULL")
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS organization_applications (
                id SERIAL PRIMARY KEY,
                organization_name TEXT NOT NULL,
                organization_type TEXT NOT NULL CHECK(organization_type IN ('learning_center','school')),
                contact_name TEXT NOT NULL,
                phone TEXT NOT NULL,
                email TEXT NOT NULL,
                student_count INTEGER,
                message TEXT,
                status TEXT NOT NULL DEFAULT 'new' CHECK(status IN ('new','contacted','approved','rejected')),
                admin_note TEXT,
                created_at TIMESTAMP DEFAULT NOW(),
                reviewed_at TIMESTAMP
            )
        """)
        await conn.execute("CREATE INDEX IF NOT EXISTS organization_applications_status_idx ON organization_applications(status, created_at DESC)")
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS testimonials (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL UNIQUE REFERENCES users(id) ON DELETE CASCADE,
                content TEXT NOT NULL,
                original_content TEXT NOT NULL,
                role_snapshot TEXT NOT NULL DEFAULT 'student',
                status TEXT NOT NULL DEFAULT 'pending'
                    CHECK(status IN ('pending','published','rejected','archived')),
                show_full_name BOOLEAN NOT NULL DEFAULT TRUE,
                show_organization BOOLEAN NOT NULL DEFAULT TRUE,
                show_avatar BOOLEAN NOT NULL DEFAULT TRUE,
                featured BOOLEAN NOT NULL DEFAULT FALSE,
                sort_order INTEGER NOT NULL DEFAULT 100,
                admin_note TEXT,
                created_at TIMESTAMP DEFAULT NOW(),
                updated_at TIMESTAMP DEFAULT NOW(),
                reviewed_at TIMESTAMP,
                published_at TIMESTAMP
            )
        """)
        await conn.execute("CREATE INDEX IF NOT EXISTS testimonials_public_idx ON testimonials(status, featured DESC, sort_order, published_at DESC)")
        await conn.execute("""
            CREATE TABLE IF NOT EXISTS faqs (
                id SERIAL PRIMARY KEY,
                category TEXT NOT NULL CHECK(category IN ('tests','results','subscription','organizations')),
                question TEXT NOT NULL UNIQUE,
                answer TEXT NOT NULL,
                is_published BOOLEAN NOT NULL DEFAULT TRUE,
                sort_order INTEGER NOT NULL DEFAULT 100,
                created_at TIMESTAMP DEFAULT NOW(),
                updated_at TIMESTAMP DEFAULT NOW()
            )
        """)
        await conn.execute("CREATE INDEX IF NOT EXISTS faqs_public_idx ON faqs(is_published, sort_order, id)")
        default_faqs = [
            ("tests", "Testlarni qanday boshlayman?", "Ro'yxatdan o'ting yoki profilingizga kiring, Testlar sahifasidan kerakli mock testni tanlang va Full mock yoki Practice rejimini boshlang.", 10),
            ("tests", "Full mock va Practice rejimlari o'rtasidagi farq nima?", "Full mock Listening, Reading, Writing va Speaking bo'limlarini ketma-ket topshiradi. Practice rejimida esa faqat kerakli bo'limni tanlaysiz.", 20),
            ("results", "Natijalarim qayerda saqlanadi?", "Barcha yakunlangan urinishlar profilingizdagi Natijalar bo'limida saqlanadi. Tashkilotga ulangan bo'lsangiz, vakolatli ustoz va rahbarlar ham natijalarni ko'ra oladi.", 30),
            ("results", "Writing va Speaking natijasi qachon chiqadi?", "Bu bo'limlar teacher tomonidan IELTS mezonlari asosida tekshiriladi. Baholash tugagach natija profilingizda ko'rinadi va sozlangan bo'lsa Telegram hamda email xabari yuboriladi.", 40),
            ("subscription", "Obuna kimlar uchun talab qilinadi?", "Obuna maktab yoki o'quv markazi uchun super-admin tomonidan alohida yoqiladi. Obuna yoqilmagan tashkilot platformadan belgilangan shartlarda davom etadi.", 50),
            ("subscription", "Obuna to'lovi qanday tasdiqlanadi?", "Ko'rsatilgan karta raqamiga to'lov qilinadi va chek tashkilot rahbari paneli orqali yuboriladi. Super-admin tekshirgach obuna faollashadi.", 60),
            ("organizations", "Maktab yoki o'quv markazini qanday ulash mumkin?", "Landing sahifadagi Tashkilot sifatida boshlash formasini yuboring. Super-admin arizani ko'rib chiqib, mas'ul shaxs bilan bog'lanadi va tashkilot panelini yaratadi.", 70),
            ("organizations", "White-label imkoniyati nima beradi?", "Tashkilot o'z nomi, logosi, ranglari, faviconi, aloqa ma'lumotlari va public slugidan foydalanishi mumkin. Asosiy platforma funksiyalari o'zgarmaydi.", 80),
        ]
        await conn.executemany(
            """
            INSERT INTO faqs(category, question, answer, is_published, sort_order)
            VALUES ($1,$2,$3,TRUE,$4) ON CONFLICT(question) DO NOTHING
            """,
            default_faqs
        )


# ─── Models ───────────────────────────────────────────────────────────────────

class StartSession(BaseModel):
    full_name: str
    email: str
    test_id: Optional[int] = None


class OrganizationApplicationIn(BaseModel):
    organization_name: str
    organization_type: str
    contact_name: str
    phone: str
    email: str
    student_count: Optional[int] = None
    message: Optional[str] = None
    website: Optional[str] = None


class OrganizationApplicationReviewIn(BaseModel):
    status: str
    note: Optional[str] = None


class TestimonialSubmitIn(BaseModel):
    content: str
    show_full_name: bool = True
    show_organization: bool = True
    show_avatar: bool = True


class TestimonialReviewIn(BaseModel):
    status: str
    content: Optional[str] = None
    featured: bool = False
    sort_order: int = 100
    admin_note: Optional[str] = None


class FaqUpsertIn(BaseModel):
    category: str
    question: str
    answer: str
    is_published: bool = True
    sort_order: int = 100


FAQ_CATEGORIES = {"tests", "results", "subscription", "organizations"}


def validate_faq(data: FaqUpsertIn) -> tuple[str, str, str, bool, int]:
    category = (data.category or "").strip().lower()
    question = re.sub(r"\s+", " ", data.question or "").strip()
    answer = re.sub(r"\s+", " ", data.answer or "").strip()
    if category not in FAQ_CATEGORIES:
        raise HTTPException(status_code=400, detail="FAQ kategoriyasi noto'g'ri")
    if not 10 <= len(question) <= 300:
        raise HTTPException(status_code=400, detail="Savol 10 dan 300 tagacha belgidan iborat bo'lishi kerak")
    if not 20 <= len(answer) <= 3000:
        raise HTTPException(status_code=400, detail="Javob 20 dan 3000 tagacha belgidan iborat bo'lishi kerak")
    if not 0 <= data.sort_order <= 10000:
        raise HTTPException(status_code=400, detail="Tartib raqami 0 dan 10000 gacha bo'lishi kerak")
    return category, question, answer, data.is_published, data.sort_order


class SubmitResult(BaseModel):
    session_id: int
    full_name: str
    email: str
    section: str
    score: Optional[int] = None
    total: Optional[int] = None
    answers: Optional[dict] = None
    writing_task1: Optional[str] = None
    writing_task2: Optional[str] = None
    duration_seconds: Optional[int] = None
    test_id: Optional[int] = None
    test_slug: Optional[str] = None
    test_mode: Optional[str] = None


class GradeWriting(BaseModel):
    result_id: int
    band: float
    feedback: Optional[str] = None
    send_email: bool = True
    task_achievement: Optional[float] = None
    coherence_cohesion: Optional[float] = None
    lexical_resource: Optional[float] = None
    grammar_accuracy: Optional[float] = None


@app.get("/api/public/stats")
async def public_platform_stats():
    """Landing sahifa uchun faqat haqiqiy, shaxsiy ma'lumotsiz ko'rsatkichlar."""
    db = await get_pool()
    async with db.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT
                (SELECT COUNT(*) FROM users
                 WHERE role='student' AND deleted_at IS NULL) AS students,
                (SELECT COUNT(*) FROM exam_sessions
                 WHERE COALESCE(cardinality(sections_completed), 0) > 0) AS completed_tests,
                (SELECT COUNT(*) FROM centers
                 WHERE organization_type='school'
                   AND is_active=TRUE AND deleted_at IS NULL) AS schools,
                (SELECT COUNT(*) FROM centers
                 WHERE organization_type='learning_center'
                   AND is_active=TRUE AND deleted_at IS NULL) AS learning_centers
            """
        )
    return dict(row)


# ─── Email yuborish (Resend API) ───────────────────────────────────────────────

async def send_email(to_email: str, to_name: str, subject: str, html_body: str):
    if not RESEND_API_KEY:
        raise RuntimeError("RESEND_API_KEY sozlanmagan")

    async with httpx.AsyncClient(timeout=15) as client:
        response = await client.post(
            "https://api.resend.com/emails",
            headers={
                "Authorization": f"Bearer {RESEND_API_KEY}",
                "Content-Type": "application/json"
            },
            json={
                "from": f"IELTS Mock SS <{EMAIL_FROM}>",
                "to": [to_email],
                "subject": subject,
                "html": html_body
            }
        )
        if response.status_code >= 400:
            raise RuntimeError(f"Resend xatosi: {response.text}")

async def send_email_with_attachment(to_email: str, to_name: str, subject: str, html_body: str, attachment_bytes: bytes, attachment_filename: str):
    if not RESEND_API_KEY:
        raise RuntimeError("RESEND_API_KEY sozlanmagan")

    b64_content = base64.b64encode(attachment_bytes).decode('utf-8')

    async with httpx.AsyncClient(timeout=15) as client:
        response = await client.post(
            "https://api.resend.com/emails",
            headers={
                "Authorization": f"Bearer {RESEND_API_KEY}",
                "Content-Type": "application/json"
            },
            json={
                "from": f"IELTS Mock SS <{EMAIL_FROM}>",
                "to": [to_email],
                "subject": subject,
                "html": html_body,
                "attachments": [
                    {
                        "filename": attachment_filename,
                        "content": b64_content
                    }
                ]
            }
        )
        if response.status_code >= 400:
            raise RuntimeError(f"Resend xatosi: {response.text}")


def build_result_email(name: str, section: str, score, total, band, feedback=None) -> str:
    section_names = {"listening": "Listening", "reading": "Reading", "writing": "Writing"}
    section_name = section_names.get(section, section)

    if section == "writing" or score is None or total is None:
        body = f"""
        <p><b>Holat:</b> {'Baholandi' if band is not None else 'Qabul qilindi, baholash kutilmoqda'}</p>
        <p><b>Band Score:</b> {band if band is not None else 'Hali baholanmagan'}</p>
        {f'<p><b>Izoh:</b> {feedback}</p>' if feedback else ''}
        """
    else:
        body = f"""
        <p><b>Natija:</b> {score} / {total} to'g'ri</p>
        <p><b>Band Score:</b> {band}</p>
        """

    return f"""
    <div style="font-family:Arial,sans-serif; max-width:520px; margin:0 auto; padding:24px; border:1px solid #c9d8ff; border-radius:12px;">
      <h2 style="color:#1a56e8;">IELTS Mock SS — {section_name} natijasi</h2>
      <p>Assalomu alaykum, {name}!</p>
      {body}
      <p style="margin-top:20px; color:#4a5978; font-size:13px;">
        Batafsil: <a href="https://ielts.sultanov.space" style="color:#1a56e8;">ielts.sultanov.space</a>
      </p>
      <p style="color:#4a5978; font-size:12px; margin-top:16px;"> © 2026-2027 Bo'stonliq tuman ixtisoslashtirilgan maktabi</p>
    </div>
    """


# ─── Endpoints ─────────────────────────────────────────────────────────────────

@app.post("/api/start")
async def start_session(data: StartSession, current_user: dict = Depends(get_current_user)):
    if not current_user.get("email_verified"):
        raise HTTPException(status_code=403, detail="Emailni tasdiqlash talab qilinadi")

    db = await get_pool()
    async with db.acquire() as conn:
        if data.test_id:
            if not await _can_access_test(conn, data.test_id, current_user):
                raise HTTPException(status_code=403, detail="Bu test sizga biriktirilmagan yoki muddati kelmagan")
            limit = await conn.fetchval(
                """
                SELECT COALESCE((
                  SELECT a.attempt_limit FROM test_assignments a
                  WHERE a.test_id=t.id AND a.center_id=$2
                    AND (a.available_from IS NULL OR a.available_from<=NOW())
                    AND (a.available_until IS NULL OR a.available_until>=NOW())
                    AND (a.group_id=$3 OR EXISTS (
                      SELECT 1 FROM school_class_students cs WHERE cs.class_id=a.class_id AND cs.student_id=$4 AND cs.left_at IS NULL
                    ) OR (a.group_id IS NULL AND a.class_id IS NULL))
                  ORDER BY CASE WHEN a.group_id IS NOT NULL OR a.class_id IS NOT NULL THEN 0 ELSE 1 END LIMIT 1
                ),t.attempt_limit,1) FROM tests t WHERE t.id=$1
                """,
                data.test_id,current_user.get("center_id"),current_user.get("group_id"),current_user["id"]
            )
            attempts = await conn.fetchval(
                "SELECT COUNT(*) FROM exam_sessions WHERE email=$1 AND test_id=$2",
                current_user["email"].strip().lower(), data.test_id
            )
            if attempts >= limit:
                raise HTTPException(status_code=409, detail=f"Urinish limitingiz tugagan ({limit})")
        row = await conn.fetchrow(
            """
            INSERT INTO exam_sessions (full_name, email, test_id)
            VALUES ($1, $2, $3)
            RETURNING id, started_at
            """,
            current_user["full_name"].strip(),
            current_user["email"].strip().lower(),
            data.test_id
        )
    return {
        "session_id": row["id"],
        "message": "Imtihon boshlandi",
        "started_at": row["started_at"].isoformat()
    }


@app.post("/api/submit")
async def submit_result(data: SubmitResult, current_user: dict = Depends(get_current_user)):
    if not current_user.get("email_verified"):
        raise HTTPException(status_code=403, detail="Emailni tasdiqlash talab qilinadi")
    if data.section not in {"listening", "reading", "writing", "speaking"}:
        raise HTTPException(status_code=400, detail="Bo'lim noto'g'ri")
    if data.test_id and data.section in {"listening", "reading"} and data.answers is not None:
        db_score = await get_pool()
        async with db_score.acquire() as score_conn:
            questions = await score_conn.fetch(
                """
                SELECT q.id,q.question_type,q.correct_answer,q.points
                FROM test_builder_questions q JOIN test_builder_sections s ON s.id=q.section_id
                WHERE s.test_id=$1 AND s.section=$2 ORDER BY q.sort_order,q.id
                """, data.test_id,data.section
            )
        if questions:
            def normal(value): return str(value if value is not None else "").strip().lower()
            earned,total_points=0.0,0.0
            for question in questions:
                expected=question["correct_answer"]
                if isinstance(expected,str):
                    try: expected=json.loads(expected)
                    except json.JSONDecodeError: pass
                submitted=data.answers.get(str(question["id"]),data.answers.get(question["id"]))
                points=float(question["points"]);total_points+=points
                if isinstance(expected,list):
                    left=sorted(normal(v) for v in (submitted if isinstance(submitted,list) else [submitted]))
                    right=sorted(normal(v) for v in expected)
                    if left==right: earned+=points
                elif normal(submitted)==normal(expected): earned+=points
            data.score=round(earned)
            data.total=round(total_points)
    if data.score is not None and data.total is not None:
        if data.total <= 0 or data.score < 0 or data.score > data.total:
            raise HTTPException(status_code=400, detail="Natija qiymatlari noto'g'ri")
    if data.section in {"listening", "reading"} and (data.score is None or data.total is None):
        raise HTTPException(status_code=400, detail="Listening/Reading uchun score va total majburiy")
    if data.answers and len(json.dumps(data.answers)) > 1_000_000:
        raise HTTPException(status_code=413, detail="Javoblar hajmi juda katta")

    writing_task1, writing_task2 = data.writing_task1, data.writing_task2
    if data.section == "writing" and data.answers:
        # Turli HTML test konstruktorlari matnni turli kalitlarda yuborishi mumkin.
        writing_task1 = writing_task1 or next((data.answers.get(key) for key in (
            "writing_task1", "task1", "task_1", "part1", "part_1", "answer1"
        ) if data.answers.get(key)), None)
        writing_task2 = writing_task2 or next((data.answers.get(key) for key in (
            "writing_task2", "task2", "task_2", "part2", "part_2", "answer2", "essay"
        ) if data.answers.get(key)), None)
    writing_task1 = str(writing_task1).strip() if writing_task1 is not None else None
    writing_task2 = str(writing_task2).strip() if writing_task2 is not None else None
    if writing_task1 and len(writing_task1) > 100_000:
        raise HTTPException(status_code=413, detail="Writing Task 1 matni juda katta")
    if writing_task2 and len(writing_task2) > 100_000:
        raise HTTPException(status_code=413, detail="Writing Task 2 matni juda katta")

    user_email = current_user["email"].strip().lower()
    user_name = current_user["full_name"].strip()

    db = await get_pool()
    async with db.acquire() as conn:
        if data.test_id and not await conn.fetchval(
            "SELECT 1 FROM tests WHERE id=$1 AND status='published'", data.test_id
        ):
            raise HTTPException(status_code=404, detail="Test topilmadi yoki publish qilinmagan")
        session_row = await conn.fetchrow(
            "SELECT id,test_id FROM exam_sessions WHERE id = $1 AND email = $2",
            data.session_id, user_email
        )
        if not session_row:
            raise HTTPException(status_code=403, detail="Bu sessiyaga ruxsat yo'q")
        if data.test_id and session_row["test_id"] and data.test_id != session_row["test_id"]:
            raise HTTPException(status_code=403, detail="Test sessiyasi mos kelmadi")

        result_row = await conn.fetchrow(
            """
            INSERT INTO exam_results 
                (full_name, email, section, score, total, answers, 
                 writing_task1, writing_task2, duration_seconds, test_id, test_slug, test_mode)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12)
            RETURNING id
            """,
            user_name,
            user_email,
            data.section,
            data.score,
            data.total,
            json.dumps(data.answers) if data.answers else None,
            writing_task1,
            writing_task2,
            data.duration_seconds,
            data.test_id,
            (data.test_slug or "")[:120] or None,
            data.test_mode if data.test_mode in {"full", "practice"} else None
        )

        await conn.execute(
            """
            UPDATE resubmission_requests
            SET status='submitted',replacement_result_id=$1
            WHERE student_id=$2 AND section=$3 AND status='pending'
            """,
            result_row["id"], current_user["id"], data.section
        )

        await conn.execute(
            """
            UPDATE exam_sessions
            SET sections_completed = array_append(sections_completed, $1)
            WHERE id = $2
              AND NOT ($1 = ANY(sections_completed))
            """,
            data.section,
            data.session_id
        )

        user_row = await conn.fetchrow("SELECT telegram_chat_id FROM users WHERE email = $1", user_email)
        chat_id = user_row["telegram_chat_id"] if user_row else None
        await create_notification(
            conn,
            recipient_user_id=current_user["id"],
            kind="task" if data.section == "writing" else "success",
            title=f"{data.section.title()} natijasi saqlandi",
            message=("Ishingiz teacher tekshiruviga yuborildi." if data.section == "writing" else f"Natijangiz: {data.score}/{data.total}."),
            action_url="/profile",
            metadata={"event": "result_saved", "result_id": result_row["id"], "section": data.section},
        )
        await notify_admin(
            conn,
            "Yangi test natijasi",
            f"{user_name} {data.section.title()} bo'limini topshirdi.",
            kind="info",
            action_url="/admin",
            metadata={"event": "result_saved", "result_id": result_row["id"], "section": data.section},
        )

    percentage = round((data.score / data.total) * 100) if data.score is not None and data.total else None
    band = get_band_score(data.score, data.total, data.section) if data.score is not None and data.total else None

    try:
        from telegram import notify_admin_new_result
        await notify_admin_new_result(user_name, user_email, data.section, data.score, data.total, band)
    except Exception:
        pass

    if current_user.get("group_id"):
        try:
            from notifications import notify_teacher_new_result
            db3 = await get_pool()
            async with db3.acquire() as conn3:
                await notify_teacher_new_result(conn3, current_user["group_id"], user_name, data.section, band)
        except Exception:
            pass

    if current_user.get("center_id"):
        try:
            from notifications import notify_head_teacher_new_result
            db4 = await get_pool()
            async with db4.acquire() as conn4:
                await notify_head_teacher_new_result(conn4, current_user["center_id"], user_name, data.section, band)
        except Exception:
            pass

    if data.section != "writing":
        try:
            db = await get_pool()
            async with db.acquire() as conn:
                rows = await conn.fetch(
                    """
                    SELECT DISTINCT ON (section) full_name, section, score, total, writing_band
                    FROM exam_results
                    WHERE email = $1
                    ORDER BY section, submitted_at DESC
                    """,
                    user_email
                )
            
            sections = []
            for r in rows:
                if r["section"] == "writing":
                    r_band = float(r["writing_band"]) if r["writing_band"] is not None else None
                elif r["score"] is not None and r["total"] is not None:
                    r_band = get_band_score(r["score"], r["total"], r["section"])
                else:
                    r_band = None
                sections.append({
                    "section": r["section"],
                    "score": r["score"],
                    "total": r["total"],
                    "band": r_band
                })
            
            pdf_bytes = build_result_pdf(user_name, user_email, sections)
            html = build_result_email(user_name, data.section, data.score, data.total, band)
            
            await send_email_with_attachment(
                user_email, user_name, f"IELTS Mock SS — {data.section.capitalize()} natijangiz",
                html, pdf_bytes, f"ielts_natija_{user_email.split('@')[0]}.pdf"
            )
            
            async with db.acquire() as conn:
                await conn.execute("UPDATE exam_results SET notified = TRUE WHERE id = $1", result_row["id"])
                
            if chat_id:
                from telegram import notify_user_result_ready
                await notify_user_result_ready(chat_id, user_name, data.section, band, data.score, data.total)
                
        except Exception as e:
            print("Email/PDF jo'natishda xatolik:", e)
    else:
        try:
            await send_email(
                user_email, user_name, "IELTS Mock SS — Writing javobingiz qabul qilindi",
                build_result_email(user_name, "writing", None, None, None)
            )
            if chat_id:
                from telegram import notify_user_result_ready
                await notify_user_result_ready(chat_id, user_name, "writing", None, None, None)
        except Exception as e:
            print("Writing xabarnomasida xatolik:", e)

    return {
        "result_id": result_row["id"],
        "message": "Natija saqlandi",
        "section": data.section,
        "score": data.score,
        "total": data.total,
        "percentage": percentage,
        "band": band
    }


@app.post("/api/submit-speaking")
async def submit_speaking(
    session_id: int = Form(...),
    test_id: Optional[int] = Form(None),
    audio: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """O'quvchi speaking audiosini yuboradi — Telegram guruhiga saqlanadi."""
    if not current_user.get("email_verified"):
        raise HTTPException(status_code=403, detail="Emailni tasdiqlash talab qilinadi")

    # Fayl hajmi va audio turini tekshirish: max 25MB (Telegram limiti)
    MAX_SIZE = 25 * 1024 * 1024
    audio_bytes = await audio.read()
    if len(audio_bytes) > MAX_SIZE:
        raise HTTPException(status_code=413, detail="Audio fayli 25MB dan katta bo'lmasin")
    if not audio_bytes:
        raise HTTPException(status_code=400, detail="Audio fayli bo'sh")
    allowed_audio_types = {"audio/webm", "audio/ogg", "audio/mpeg", "audio/mp4", "audio/wav", "audio/x-wav"}
    audio_mime = (audio.content_type or "").lower()
    if audio_mime not in allowed_audio_types:
        raise HTTPException(status_code=400, detail="Audio formati qo'llab-quvvatlanmaydi")
    looks_like_audio = (
        audio_bytes.startswith(b"OggS") or audio_bytes.startswith(b"RIFF") or
        audio_bytes.startswith(b"ID3") or audio_bytes[:2] in {b"\xff\xfb", b"\xff\xf3", b"\xff\xf2"} or
        audio_bytes.startswith(b"\x1aE\xdf\xa3") or b"ftyp" in audio_bytes[:32]
    )
    if not looks_like_audio:
        raise HTTPException(status_code=400, detail="Audio fayl tarkibi noto'g'ri")

    user_email = current_user["email"].strip().lower()
    user_name = current_user["full_name"].strip()

    db = await get_pool()
    async with db.acquire() as conn:
        session_row = await conn.fetchrow(
            "SELECT id,test_id FROM exam_sessions WHERE id = $1 AND email = $2",
            session_id, user_email
        )
        if not session_row:
            raise HTTPException(status_code=403, detail="Bu sessiyaga ruxsat yo'q")
        if test_id and session_row["test_id"] and test_id != session_row["test_id"]:
            raise HTTPException(status_code=403, detail="Test sessiyasi mos kelmadi")

        # Avval DB ga yozib olamiz (file_id keyinroq yangilanadi)
        result_row = await conn.fetchrow(
            """
            INSERT INTO exam_results
                (full_name, email, section, duration_seconds, speaking_audio_data,
                 speaking_audio_mime, speaking_audio_filename,test_id)
            VALUES ($1, $2, 'speaking', NULL, $3, $4, $5,$6)
            RETURNING id
            """,
            user_name, user_email, audio_bytes, audio_mime, (audio.filename or "speaking-audio")[:180],test_id
        )
        result_id = result_row["id"]

        await conn.execute(
            "UPDATE resubmission_requests SET status='submitted',replacement_result_id=$1 WHERE student_id=$2 AND section='speaking' AND status='pending'",
            result_id,current_user["id"]
        )

        await conn.execute(
            """
            UPDATE exam_sessions
            SET sections_completed = array_append(sections_completed, 'speaking')
            WHERE id = $1
              AND NOT ('speaking' = ANY(sections_completed))
            """,
            session_id
        )
        await create_notification(
            conn,
            recipient_user_id=current_user["id"],
            kind="task",
            title="Speaking audiongiz saqlandi",
            message="Speaking javobingiz teacher tekshiruviga yuborildi.",
            action_url="/profile",
            metadata={"event": "speaking_submitted", "result_id": result_id, "section": "speaking"},
        )
        await notify_admin(
            conn,
            "Yangi Speaking javobi",
            f"{user_name} Speaking audiosini yubordi.",
            kind="task",
            action_url="/admin",
            metadata={"event": "speaking_submitted", "result_id": result_id},
        )

    # Telegram guruhiga audio yuborish
    from telegram import send_voice_to_telegram, notify_admin_new_speaking
    caption = (
        f"🎤 <b>Speaking Audio</b>\n\n"
        f"👤 {user_name}\n"
        f"📧 {user_email}\n"
        f"🆔 result_id: {result_id}"
    )
    filename = f"speaking_{result_id}_{user_email.split('@')[0]}.ogg"
    try:
        file_id = await send_voice_to_telegram(audio_bytes, filename, caption)
    except Exception as exc:
        print(f"[Speaking] Telegram yuborishda xatolik: {exc}")
        file_id = None

    if file_id:
        db = await get_pool()
        async with db.acquire() as conn:
            await conn.execute(
                "UPDATE exam_results SET speaking_telegram_file_id = $1 WHERE id = $2",
                file_id, result_id
            )
    else:
        # file_id olmasa ham natija saqlanadi, log qilamiz
        print(f"[Speaking] Telegram yuborishda xatolik. result_id={result_id}")

    # Admin ga xabar
    try:
        await notify_admin_new_speaking(user_name, user_email, file_id)
    except Exception:
        pass

    # Teacher ga xabar
    if current_user.get("group_id"):
        try:
            from notifications import notify_teacher_new_result
            db2 = await get_pool()
            async with db2.acquire() as conn2:
                await notify_teacher_new_result(conn2, current_user["group_id"], user_name, "speaking", None)
        except Exception:
            pass

    return {
        "result_id": result_id,
        "message": "Speaking audioniz muvaffaqiyatli yuborildi",
        "telegram_saved": file_id is not None
    }

@app.get("/api/results")
async def get_my_results(current_user: dict = Depends(get_current_user)):
    return await get_results(current_user["email"], current_user)


@app.get("/api/results/pdf")
async def get_my_results_pdf(current_user: dict = Depends(get_current_user)):
    return await get_results_pdf(current_user["email"], current_user)


@app.get("/api/results/{email}")
async def get_results(email: str, current_user: dict = Depends(get_current_user)):
    if email.strip().lower() != current_user["email"].strip().lower():
        raise HTTPException(status_code=403, detail="Faqat o'zingizning natijalaringizni ko'rishingiz mumkin")

    db = await get_pool()
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT section, score, total, submitted_at
            FROM exam_results
            WHERE email = $1
            ORDER BY submitted_at DESC
            LIMIT 20
            """,
            email.lower()
        )
    results = [
        {
            "section": r["section"],
            "score": r["score"],
            "total": r["total"],
            "band": get_band_score(r["score"], r["total"], r["section"])
                    if r["score"] is not None and r["total"] is not None else None,
            "submitted_at": r["submitted_at"].isoformat()
        }
        for r in rows
    ]
    return {"email": email, "results": results}


# ─── PDF Sertifikat ─────────────────────────────────────────────────────────────

def build_result_pdf(full_name: str, email: str, sections: list) -> bytes:
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_auto_page_break(auto=False)

    # Header
    pdf.set_fill_color(26, 86, 232)
    pdf.rect(0, 0, 210, 30, style="F")
    pdf.set_xy(0, 9)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(210, 8, "IELTS Mock SS Exam - Natija", align="C")
    pdf.set_xy(0, 18)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(210, 6, "ielts.sultanov.space", align="C")

    # Demo disclaimer banner
    pdf.set_xy(10, 36)
    pdf.set_fill_color(255, 251, 235)
    pdf.set_draw_color(217, 119, 6)
    pdf.rect(10, 36, 190, 12, style="DF")
    pdf.set_xy(12, 39)
    pdf.set_text_color(217, 119, 6)
    pdf.set_font("Helvetica", "", 9)
    pdf.multi_cell(186, 5,
        "Diqqat: bu DEMO test rejimi natijasi. Savollar namunaviy, haqiqiy imtihon "
        "formatidagi savollar hali qo'shilmagan. Rasmiy sertifikat emas.")

    # Candidate info
    y = 56
    pdf.set_text_color(11, 23, 51)
    pdf.set_xy(10, y)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(95, 7, "Ism-familiya:")
    pdf.set_font("Helvetica", "", 12)
    pdf.cell(95, 7, full_name)
    pdf.set_xy(10, y + 7)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(95, 7, "Email:")
    pdf.set_font("Helvetica", "", 12)
    pdf.cell(95, 7, email)
    pdf.set_xy(10, y + 14)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(95, 7, "Sana:")
    pdf.set_font("Helvetica", "", 12)
    pdf.cell(95, 7, datetime.now().strftime("%d.%m.%Y"))

    # Overall band (average of graded sections)
    graded_bands = [s["band"] for s in sections if s.get("band") is not None]
    overall = round(sum(graded_bands) / len(graded_bands), 1) if graded_bands else None

    y = 84
    pdf.set_fill_color(238, 243, 255)
    pdf.rect(10, y, 190, 26, style="F")
    pdf.set_xy(10, y + 4)
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(74, 89, 120)
    pdf.cell(190, 6, "UMUMIY BAND SCORE", align="C")
    pdf.set_xy(10, y + 10)
    pdf.set_font("Helvetica", "B", 26)
    pdf.set_text_color(26, 86, 232)
    pdf.cell(190, 12, str(overall) if overall is not None else "Kutilmoqda", align="C")

    # Section breakdown table
    y = 118
    pdf.set_text_color(11, 23, 51)
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_xy(10, y)
    pdf.cell(190, 8, "Bo'limlar bo'yicha natija")

    y += 10
    section_labels = {"listening": "Listening", "reading": "Reading", "writing": "Writing"}
    pdf.set_fill_color(26, 86, 232)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_xy(10, y)
    pdf.cell(63, 9, "Bo'lim", border=1, fill=True, align="C")
    pdf.cell(63, 9, "Natija", border=1, fill=True, align="C")
    pdf.cell(64, 9, "Band", border=1, fill=True, align="C")

    pdf.set_text_color(11, 23, 51)
    pdf.set_font("Helvetica", "", 10)
    for s in sections:
        y += 9
        pdf.set_xy(10, y)
        label = section_labels.get(s["section"], s["section"])
        if s["section"] == "writing":
            detail = "Tekshirilmoqda" if s.get("band") is None else "Admin tomonidan baholandi"
        else:
            detail = f'{s["score"]}/{s["total"]}' if s.get("score") is not None else "-"
        band_txt = str(s["band"]) if s.get("band") is not None else "-"
        pdf.cell(63, 9, label, border=1, align="C")
        pdf.cell(63, 9, detail, border=1, align="C")
        pdf.cell(64, 9, band_txt, border=1, align="C")

    # Footer
    pdf.set_xy(10, 275)
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(74, 89, 120)
    pdf.cell(190, 5, "Bu hujjat avtomatik generatsiya qilingan - ielts.sultanov.space", align="C")

    return bytes(pdf.output())


@app.get("/api/results/{email}/pdf")
async def get_results_pdf(email: str, current_user: dict = Depends(get_current_user)):
    if email.strip().lower() != current_user["email"].strip().lower():
        raise HTTPException(status_code=403, detail="Faqat o'zingizning natijalaringizni yuklab olishingiz mumkin")

    db = await get_pool()
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT DISTINCT ON (section) full_name, section, score, total, writing_band
            FROM exam_results
            WHERE email = $1
            ORDER BY section, submitted_at DESC
            """,
            email.lower()
        )

    if not rows:
        raise HTTPException(status_code=404, detail="Natija topilmadi")

    full_name = rows[0]["full_name"]
    sections = []
    for r in rows:
        if r["section"] == "writing":
            band = float(r["writing_band"]) if r["writing_band"] is not None else None
        elif r["score"] is not None and r["total"] is not None:
            band = get_band_score(r["score"], r["total"], r["section"])
        else:
            band = None
        sections.append({
            "section": r["section"],
            "score": r["score"],
            "total": r["total"],
            "band": band
        })

    order = {"listening": 0, "reading": 1, "writing": 2}
    sections.sort(key=lambda s: order.get(s["section"], 99))

    pdf_bytes = build_result_pdf(full_name, email.lower(), sections)
    filename = f"ielts_natija_{email.lower().split('@')[0]}.pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ─── ORGANIZATION APPLICATIONS ────────────────────────────────────────────────

@app.post("/api/organization-applications")
async def create_organization_application(data: OrganizationApplicationIn):
    if data.website:
        return {"ok": True, "message": "Arizangiz qabul qilindi"}

    organization_name = data.organization_name.strip()
    contact_name = data.contact_name.strip()
    phone = data.phone.strip()
    email = data.email.strip().lower()
    message = (data.message or "").strip() or None

    if data.organization_type not in ORGANIZATION_TYPES:
        raise HTTPException(status_code=400, detail="Tashkilot turi noto'g'ri")
    if not 2 <= len(organization_name) <= 160:
        raise HTTPException(status_code=400, detail="Tashkilot nomini to'g'ri kiriting")
    if not 2 <= len(contact_name) <= 120:
        raise HTTPException(status_code=400, detail="Mas'ul shaxs ismini to'g'ri kiriting")
    if not re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", email) or len(email) > 254:
        raise HTTPException(status_code=400, detail="Email manzil noto'g'ri")
    if not 7 <= len(phone) <= 30:
        raise HTTPException(status_code=400, detail="Telefon raqamini to'g'ri kiriting")
    if data.student_count is not None and not 1 <= data.student_count <= 1_000_000:
        raise HTTPException(status_code=400, detail="O'quvchilar soni noto'g'ri")
    if message and len(message) > 1500:
        raise HTTPException(status_code=400, detail="Izoh 1500 belgidan oshmasligi kerak")

    db = await get_pool()
    async with db.acquire() as conn:
        existing = await conn.fetchval(
            """
            SELECT id FROM organization_applications
            WHERE LOWER(email)=$1 AND LOWER(organization_name)=LOWER($2)
              AND status IN ('new','contacted') AND created_at > NOW() - INTERVAL '24 hours'
            ORDER BY created_at DESC LIMIT 1
            """,
            email, organization_name
        )
        if existing:
            return {"ok": True, "application_id": existing, "message": "Arizangiz avval qabul qilingan"}
        application_id = await conn.fetchval(
            """
            INSERT INTO organization_applications
                (organization_name, organization_type, contact_name, phone, email, student_count, message)
            VALUES ($1,$2,$3,$4,$5,$6,$7) RETURNING id
            """,
            organization_name, data.organization_type, contact_name, phone, email,
            data.student_count, message
        )
        await notify_admin(
            conn,
            "Yangi tashkilot arizasi",
            f"{organization_name} nomidan {contact_name} ariza yubordi.",
            kind="task",
            action_url="/admin",
            metadata={"event": "organization_application", "application_id": application_id},
        )
    return {"ok": True, "application_id": application_id, "message": "Arizangiz qabul qilindi"}


def testimonial_role_label(role: str, organization_type: Optional[str] = None) -> str:
    if role == "teacher":
        return "Ustoz"
    if role == "head_teacher":
        return "Maktab direktori" if organization_type == "school" else "O'quv markazi rahbari"
    if role == "school_staff":
        return "Maktab xodimi"
    return "O'quvchi"


@app.get("/api/testimonials/public")
async def public_testimonials():
    db = await get_pool()
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT t.id, t.content, t.show_full_name, t.show_organization, t.show_avatar,
                   t.featured, t.published_at, u.full_name, u.username, u.role,
                   u.avatar_mime, c.name AS organization_name, c.organization_type
            FROM testimonials t
            JOIN users u ON u.id=t.user_id
            LEFT JOIN centers c ON c.id=u.center_id
            WHERE t.status='published'
              AND COALESCE(u.is_suspended, FALSE)=FALSE
              AND u.deleted_at IS NULL
            ORDER BY t.featured DESC, t.sort_order ASC, t.published_at DESC NULLS LAST
            LIMIT 12
            """
        )
    return [
        {
            "id": row["id"],
            "content": row["content"],
            "full_name": (row["full_name"] or row["username"]) if row["show_full_name"] else "Tasdiqlangan foydalanuvchi",
            "role": testimonial_role_label(row["role"], row["organization_type"]),
            "organization_name": row["organization_name"] if row["show_organization"] else None,
            "avatar_url": f"/api/auth/avatar/{row['username']}" if row["show_avatar"] and row["avatar_mime"] else None,
            "featured": row["featured"],
            "published_at": row["published_at"],
        }
        for row in rows
    ]


@app.get("/api/faqs/public")
async def public_faqs():
    db = await get_pool()
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT id, category, question, answer
            FROM faqs WHERE is_published=TRUE
            ORDER BY sort_order ASC, id ASC
            LIMIT 100
            """
        )
    return [dict(row) for row in rows]


@app.get("/api/testimonials/me")
async def my_testimonial(current_user: dict = Depends(get_current_user)):
    db = await get_pool()
    async with db.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT id, content, original_content, status, show_full_name,
                   show_organization, show_avatar, admin_note, created_at,
                   updated_at, reviewed_at, published_at
            FROM testimonials WHERE user_id=$1
            """,
            current_user["id"]
        )
    return dict(row) if row else None


@app.post("/api/testimonials/me")
async def submit_testimonial(data: TestimonialSubmitIn, current_user: dict = Depends(get_current_user)):
    if not current_user.get("email_verified"):
        raise HTTPException(status_code=403, detail="Fikr yuborish uchun emailingizni tasdiqlang")
    content = re.sub(r"\s+", " ", data.content or "").strip()
    if len(content) < 40:
        raise HTTPException(status_code=400, detail="Fikr kamida 40 belgidan iborat bo'lishi kerak")
    if len(content) > 800:
        raise HTTPException(status_code=400, detail="Fikr 800 belgidan oshmasligi kerak")
    db = await get_pool()
    async with db.acquire() as conn:
        row = await conn.fetchrow(
            """
            INSERT INTO testimonials
                (user_id, content, original_content, role_snapshot, show_full_name,
                 show_organization, show_avatar)
            VALUES ($1,$2,$2,$3,$4,$5,$6)
            ON CONFLICT (user_id) DO UPDATE SET
                content=EXCLUDED.content,
                original_content=EXCLUDED.original_content,
                role_snapshot=EXCLUDED.role_snapshot,
                status='pending',
                show_full_name=EXCLUDED.show_full_name,
                show_organization=EXCLUDED.show_organization,
                show_avatar=EXCLUDED.show_avatar,
                featured=FALSE,
                admin_note=NULL,
                updated_at=NOW(),
                reviewed_at=NULL,
                published_at=NULL
            RETURNING id, status, updated_at
            """,
            current_user["id"], content, current_user.get("role") or "student",
            data.show_full_name, data.show_organization, data.show_avatar
        )
        await notify_admin(
            conn,
            "Yangi foydalanuvchi fikri",
            f"{current_user['full_name']} fikrini moderatsiyaga yubordi.",
            kind="task",
            action_url="/admin",
            metadata={"event": "testimonial_pending", "testimonial_id": row["id"]},
        )
    return {**dict(row), "message": "Fikringiz moderatsiyaga yuborildi"}


@app.delete("/api/testimonials/me")
async def delete_my_testimonial(current_user: dict = Depends(get_current_user)):
    db = await get_pool()
    async with db.acquire() as conn:
        deleted = await conn.fetchval(
            "DELETE FROM testimonials WHERE user_id=$1 RETURNING id", current_user["id"]
        )
    if not deleted:
        raise HTTPException(status_code=404, detail="Fikr topilmadi")
    return {"ok": True}


# ─── ADMIN ENDPOINTS ────────────────────────────────────────────────────────────

def check_admin(secret: str):
    if not ADMIN_SECRET or ADMIN_SECRET == "admin123" or len(ADMIN_SECRET) < MIN_ADMIN_SECRET_LENGTH:
        raise HTTPException(
            status_code=503,
            detail=f"ADMIN_SECRET sozlanmagan yoki {MIN_ADMIN_SECRET_LENGTH} belgidan qisqa"
        )
    if not hmac.compare_digest(secret, ADMIN_SECRET):
        raise HTTPException(status_code=403, detail="Ruxsat yo'q — noto'g'ri parol")


def require_admin(x_admin_secret: str = Header("", alias="X-Admin-Secret")):
    check_admin(x_admin_secret)


@app.get("/api/admin/notifications")
async def admin_notifications(_: None = Depends(require_admin)):
    db = await get_pool()
    async with db.acquire() as conn:
        return await fetch_notifications(conn, role="admin", limit=40)


@app.post("/api/admin/notifications/{notification_id}/read")
async def admin_read_notification(notification_id: int, _: None = Depends(require_admin)):
    db = await get_pool()
    async with db.acquire() as conn:
        updated = await mark_notification_read(conn, notification_id, role="admin")
    if not updated:
        raise HTTPException(status_code=404, detail="Bildirishnoma topilmadi")
    return {"ok": True}


@app.post("/api/admin/notifications/read-all")
async def admin_read_all_notifications(_: None = Depends(require_admin)):
    db = await get_pool()
    async with db.acquire() as conn:
        updated = await mark_all_notifications_read(conn, role="admin")
    return {"ok": True, "updated": updated}


@app.get("/api/admin/faqs")
async def admin_faqs(_: None = Depends(require_admin)):
    db = await get_pool()
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT id, category, question, answer, is_published, sort_order,
                   created_at, updated_at
            FROM faqs ORDER BY sort_order ASC, id ASC
            """
        )
    return [dict(row) for row in rows]


@app.post("/api/admin/faqs")
async def admin_create_faq(data: FaqUpsertIn, _: None = Depends(require_admin)):
    category, question, answer, is_published, sort_order = validate_faq(data)
    db = await get_pool()
    async with db.acquire() as conn:
        if await conn.fetchval("SELECT 1 FROM faqs WHERE LOWER(question)=LOWER($1)", question):
            raise HTTPException(status_code=409, detail="Bu savol allaqachon mavjud")
        row = await conn.fetchrow(
            """
            INSERT INTO faqs(category, question, answer, is_published, sort_order)
            VALUES ($1,$2,$3,$4,$5)
            RETURNING id, category, question, answer, is_published, sort_order,
                      created_at, updated_at
            """,
            category, question, answer, is_published, sort_order
        )
    return dict(row)


@app.put("/api/admin/faqs/{faq_id}")
async def admin_update_faq(faq_id: int, data: FaqUpsertIn, _: None = Depends(require_admin)):
    category, question, answer, is_published, sort_order = validate_faq(data)
    db = await get_pool()
    async with db.acquire() as conn:
        if await conn.fetchval(
            "SELECT 1 FROM faqs WHERE LOWER(question)=LOWER($1) AND id<>$2", question, faq_id
        ):
            raise HTTPException(status_code=409, detail="Bu savol allaqachon mavjud")
        row = await conn.fetchrow(
            """
            UPDATE faqs SET category=$2, question=$3, answer=$4,
                is_published=$5, sort_order=$6, updated_at=NOW()
            WHERE id=$1
            RETURNING id, category, question, answer, is_published, sort_order,
                      created_at, updated_at
            """,
            faq_id, category, question, answer, is_published, sort_order
        )
    if not row:
        raise HTTPException(status_code=404, detail="FAQ topilmadi")
    return dict(row)


@app.delete("/api/admin/faqs/{faq_id}")
async def admin_delete_faq(faq_id: int, _: None = Depends(require_admin)):
    db = await get_pool()
    async with db.acquire() as conn:
        deleted = await conn.fetchval("DELETE FROM faqs WHERE id=$1 RETURNING id", faq_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="FAQ topilmadi")
    return {"ok": True}


@app.get("/api/admin/testimonials")
async def admin_testimonials(status: Optional[str] = None, _: None = Depends(require_admin)):
    if status and status not in {"pending", "published", "rejected", "archived"}:
        raise HTTPException(status_code=400, detail="Fikr holati noto'g'ri")
    db = await get_pool()
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT t.id, t.content, t.original_content, t.status, t.role_snapshot,
                   t.show_full_name, t.show_organization, t.show_avatar, t.featured,
                   t.sort_order, t.admin_note, t.created_at, t.updated_at, t.reviewed_at,
                   t.published_at, u.id AS user_id, u.full_name, u.username, u.email,
                   u.role, c.name AS organization_name, c.organization_type
            FROM testimonials t
            JOIN users u ON u.id=t.user_id
            LEFT JOIN centers c ON c.id=u.center_id
            WHERE ($1::text IS NULL OR t.status=$1)
            ORDER BY (t.status='pending') DESC, t.featured DESC, t.updated_at DESC
            LIMIT 500
            """,
            status
        )
    return [
        {
            **dict(row),
            "role_label": testimonial_role_label(row["role"], row["organization_type"]),
        }
        for row in rows
    ]


@app.post("/api/admin/testimonials/{testimonial_id}/review")
async def admin_review_testimonial(
    testimonial_id: int,
    data: TestimonialReviewIn,
    _: None = Depends(require_admin)
):
    if data.status not in {"pending", "published", "rejected", "archived"}:
        raise HTTPException(status_code=400, detail="Fikr holati noto'g'ri")
    content = re.sub(r"\s+", " ", data.content or "").strip() if data.content is not None else None
    if content is not None and not 40 <= len(content) <= 800:
        raise HTTPException(status_code=400, detail="Fikr 40 dan 800 tagacha belgidan iborat bo'lishi kerak")
    if not 0 <= data.sort_order <= 10000:
        raise HTTPException(status_code=400, detail="Tartib raqami 0 dan 10000 gacha bo'lishi kerak")
    note = (data.admin_note or "").strip() or None
    if note and len(note) > 1500:
        raise HTTPException(status_code=400, detail="Admin izohi juda uzun")
    db = await get_pool()
    async with db.acquire() as conn:
        row = await conn.fetchrow(
            """
            UPDATE testimonials SET
                content=COALESCE($2, content),
                status=$3,
                featured=CASE WHEN $3='published' THEN $4 ELSE FALSE END,
                sort_order=$5,
                admin_note=$6,
                updated_at=NOW(),
                reviewed_at=CASE WHEN $3='pending' THEN NULL ELSE NOW() END,
                published_at=CASE WHEN $3='published' THEN COALESCE(published_at, NOW()) ELSE NULL END
            WHERE id=$1
            RETURNING id, user_id, content, original_content, status, featured, sort_order,
                      admin_note, updated_at, reviewed_at, published_at
            """,
            testimonial_id, content, data.status, data.featured, data.sort_order, note
        )
        if row and data.status in {"published", "rejected", "archived"}:
            status_text = {
                "published": "Fikringiz landing page'da chop etildi.",
                "rejected": "Fikringiz moderatsiyadan o'tmadi.",
                "archived": "Fikringiz arxivlandi.",
            }[data.status]
            await create_notification(
                conn,
                recipient_user_id=row["user_id"],
                kind="success" if data.status == "published" else "warning",
                title="Fikringiz holati yangilandi",
                message=status_text + (f" Admin izohi: {note}" if note else ""),
                action_url="/profile",
                metadata={"event": "testimonial_review", "testimonial_id": testimonial_id, "status": data.status},
            )
    if not row:
        raise HTTPException(status_code=404, detail="Fikr topilmadi")
    return dict(row)


@app.delete("/api/admin/testimonials/{testimonial_id}")
async def admin_delete_testimonial(testimonial_id: int, _: None = Depends(require_admin)):
    db = await get_pool()
    async with db.acquire() as conn:
        deleted = await conn.fetchval(
            "DELETE FROM testimonials WHERE id=$1 RETURNING id", testimonial_id
        )
    if not deleted:
        raise HTTPException(status_code=404, detail="Fikr topilmadi")
    return {"ok": True}


@app.get("/api/admin/organization-applications")
async def admin_organization_applications(_: None = Depends(require_admin)):
    db = await get_pool()
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT id, organization_name, organization_type, contact_name, phone, email,
                   student_count, message, status, admin_note, created_at, reviewed_at
            FROM organization_applications ORDER BY (status='new') DESC, created_at DESC
            LIMIT 500
            """
        )
    return [dict(row) for row in rows]


@app.post("/api/admin/organization-applications/{application_id}/review")
async def admin_review_organization_application(
    application_id: int,
    data: OrganizationApplicationReviewIn,
    _: None = Depends(require_admin)
):
    if data.status not in {"new", "contacted", "approved", "rejected"}:
        raise HTTPException(status_code=400, detail="Ariza holati noto'g'ri")
    note = (data.note or "").strip() or None
    if note and len(note) > 1500:
        raise HTTPException(status_code=400, detail="Izoh juda uzun")
    db = await get_pool()
    async with db.acquire() as conn:
        row = await conn.fetchrow(
            """
            UPDATE organization_applications
            SET status=$2, admin_note=$3, reviewed_at=CASE WHEN $2='new' THEN NULL ELSE NOW() END
            WHERE id=$1 RETURNING id, status
            """,
            application_id, data.status, note
        )
    if not row:
        raise HTTPException(status_code=404, detail="Ariza topilmadi")
    return dict(row)


@app.delete("/api/admin/organization-applications/{application_id}")
async def admin_delete_organization_application(application_id: int, _: None = Depends(require_admin)):
    db = await get_pool()
    async with db.acquire() as conn:
        deleted = await conn.fetchval(
            "DELETE FROM organization_applications WHERE id=$1 RETURNING id", application_id
        )
    if not deleted:
        raise HTTPException(status_code=404, detail="Ariza topilmadi")
    return {"ok": True}


@app.get("/api/admin/results")
async def admin_results(_: None = Depends(require_admin)):
    db = await get_pool()
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT er.id, er.full_name, er.email, er.section, er.score, er.total,
                   er.writing_band, er.writing_feedback, er.notified, er.submitted_at,
                   COALESCE(t.title, er.test_slug, 'IELTS Mock SS') AS test_title
            FROM exam_results er LEFT JOIN tests t ON t.id=er.test_id
            ORDER BY er.submitted_at DESC
            LIMIT 300
            """
        )
    return {
        "total": len(rows),
        "results": [
            {
                "id": r["id"],
                "full_name": r["full_name"],
                "email": r["email"],
                "section": r["section"],
                "test_title": r["test_title"],
                "score": r["score"],
                "total": r["total"],
                "band": (
                    float(r["writing_band"]) if r["section"] == "writing" and r["writing_band"] is not None
                    else get_band_score(r["score"], r["total"], r["section"])
                    if r["section"] != "writing" and r["score"] is not None and r["total"] is not None
                    else None
                ),
                "writing_feedback": r["writing_feedback"],
                "notified": r["notified"],
                "submitted_at": r["submitted_at"].isoformat()
            }
            for r in rows
        ]
    }


@app.get("/api/admin/users")
async def admin_users(
    search: Optional[str] = None,
    center_id: Optional[int] = None,
    affiliation: str = "all",
    role: Optional[str] = None,
    status: str = "all",
    page: int = 1,
    page_size: int = 50,
    _: None = Depends(require_admin),
):
    if affiliation not in {"all", "affiliated", "independent"}:
        raise HTTPException(status_code=400, detail="Tashkilot filtri noto'g'ri")
    if center_id is not None and center_id <= 0:
        raise HTTPException(status_code=400, detail="Tashkilot ID noto'g'ri")
    allowed_roles = {"student", "teacher", "head_teacher", "school_staff", "director", "admin"}
    if role and role not in allowed_roles:
        raise HTTPException(status_code=400, detail="Foydalanuvchi roli noto'g'ri")
    if status not in {"all", "active", "suspended", "unverified", "deleted"}:
        raise HTTPException(status_code=400, detail="Foydalanuvchi holati noto'g'ri")
    page = max(1, page)
    page_size = min(200, max(10, page_size))
    db = await get_pool()
    search_term = f"%{search.strip()}%" if search and search.strip() else None
    async with db.acquire() as conn:
        total = await conn.fetchval(
            """
            SELECT COUNT(*)
            FROM users u
            WHERE ($1::text IS NULL OR u.username ILIKE $1 OR u.email ILIKE $1 OR u.full_name ILIKE $1)
              AND ($2::int IS NULL OR u.center_id=$2)
              AND (
                $3::text='all'
                OR ($3='affiliated' AND u.center_id IS NOT NULL)
                OR ($3='independent' AND u.center_id IS NULL)
              )
              AND ($4::text IS NULL OR u.role=$4)
              AND (
                $5::text='all'
                OR ($5='active' AND u.deleted_at IS NULL AND COALESCE(u.is_suspended,FALSE)=FALSE)
                OR ($5='suspended' AND u.deleted_at IS NULL AND COALESCE(u.is_suspended,FALSE)=TRUE)
                OR ($5='unverified' AND u.deleted_at IS NULL AND u.email_verified IS NOT TRUE)
                OR ($5='deleted' AND u.deleted_at IS NOT NULL)
              )
            """,
            search_term, center_id, affiliation, role, status
        )
        total_pages = max(1, (total + page_size - 1) // page_size)
        page = min(page, total_pages)
        offset = (page - 1) * page_size
        rows = await conn.fetch(
            """
            SELECT u.id, u.username, u.email, u.full_name, u.created_at,
                   u.email_verified, u.is_suspended, u.deleted_at, u.role,
                   u.center_id, c.name AS organization_name,
                   c.organization_type,
                   (u.avatar_mime IS NOT NULL) AS has_avatar,
                   COALESCE(er.attempts, 0) AS attempts
            FROM users u
            LEFT JOIN centers c ON c.id=u.center_id
            LEFT JOIN (
                SELECT LOWER(email) AS email_key, COUNT(*) AS attempts
                FROM exam_results GROUP BY LOWER(email)
            ) er ON er.email_key=LOWER(u.email)
            WHERE ($1::text IS NULL OR u.username ILIKE $1 OR u.email ILIKE $1 OR u.full_name ILIKE $1)
              AND ($2::int IS NULL OR u.center_id=$2)
              AND (
                $3::text='all'
                OR ($3='affiliated' AND u.center_id IS NOT NULL)
                OR ($3='independent' AND u.center_id IS NULL)
              )
              AND ($4::text IS NULL OR u.role=$4)
              AND (
                $5::text='all'
                OR ($5='active' AND u.deleted_at IS NULL AND COALESCE(u.is_suspended,FALSE)=FALSE)
                OR ($5='suspended' AND u.deleted_at IS NULL AND COALESCE(u.is_suspended,FALSE)=TRUE)
                OR ($5='unverified' AND u.deleted_at IS NULL AND u.email_verified IS NOT TRUE)
                OR ($5='deleted' AND u.deleted_at IS NOT NULL)
              )
            ORDER BY u.created_at DESC
            LIMIT $6 OFFSET $7
            """,
            search_term, center_id, affiliation, role, status, page_size, offset
        )
    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages,
        "users": [
            {
                "id": r["id"],
                "username": r["username"],
                "email": r["email"],
                "full_name": r["full_name"],
                "created_at": r["created_at"].isoformat(),
                "email_verified": r["email_verified"],
                "is_suspended": r["is_suspended"],
                "deleted_at": r["deleted_at"].isoformat() if r["deleted_at"] else None,
                "role": r["role"],
                "center_id": r["center_id"],
                "organization_name": r["organization_name"],
                "organization_type": r["organization_type"],
                "attempts": r["attempts"],
                "avatar_url": f"/api/auth/avatar/{r['username']}" if r.get("has_avatar") else None
            }
            for r in rows
        ]
    }


class CenterCreateIn(BaseModel):
    name: str
    organization_type: str = "learning_center"
    max_groups: Optional[int] = None
    max_students: Optional[int] = None


class AdminDirectoryOverrideIn(BaseModel):
    mode: str = "inherit"
    reason: Optional[str] = None
    slug: Optional[str] = None
    featured: bool = False
    sort_order: int = 100


class AssignHeadTeacherIn(BaseModel):
    user_id: int


class AdminPaymentReviewIn(BaseModel):
    action: str
    note: Optional[str] = None


@app.post("/api/admin/centers")
async def admin_create_center(data: CenterCreateIn, _: None = Depends(require_admin)):
    name = data.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Tashkilot nomini kiriting")
    if data.organization_type not in ORGANIZATION_TYPES:
        raise HTTPException(status_code=400, detail="Tashkilot turi noto'g'ri")

    db = await get_pool()
    async with db.acquire() as conn:
        row = await conn.fetchrow(
            """
            INSERT INTO centers (name, organization_type, max_groups, max_students, brand_name)
            VALUES ($1, $2, $3, $4, $1)
            RETURNING id, name, organization_type, max_groups, max_students, is_active, created_at
            """,
            name, data.organization_type, data.max_groups, data.max_students
        )
    return dict(row) | {"created_at": row["created_at"].isoformat()}


@app.get("/api/admin/centers")
async def admin_list_centers(_: None = Depends(require_admin)):
    db = await get_pool()
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT c.id, c.name, c.organization_type, c.slug, c.brand_name, c.subscription_required, c.test_upload_enabled,
                   c.brand_primary_color, c.brand_logo_url,
                   c.directory_opt_in, c.directory_admin_override, c.directory_admin_reason,
                   c.directory_region, c.directory_featured, c.directory_sort_order,
                   c.is_active, c.deleted_at, c.max_groups, c.max_students, c.created_at,
                   c.owner_id, owner.full_name AS owner_name, owner.email AS owner_email,
                   sub.status AS subscription_status, sub.current_period_end,
                   COUNT(DISTINCT g.id) AS groups_count,
                   COUNT(DISTINCT u.id) FILTER (WHERE u.role = 'student') AS students_count
            FROM centers c
            LEFT JOIN users owner ON owner.id = c.owner_id
            LEFT JOIN organization_subscriptions sub ON sub.center_id = c.id
            LEFT JOIN groups g ON g.center_id = c.id
            LEFT JOIN users u ON u.center_id = c.id
            GROUP BY c.id, owner.full_name, owner.email, sub.status, sub.current_period_end
            ORDER BY c.created_at DESC
            """
        )
    return [
        {
            "id": r["id"], "name": r["name"], "organization_type": r["organization_type"],
            "slug": r["slug"], "brand_name": r["brand_name"],
            "subscription_required": r["subscription_required"],
            "test_upload_enabled": r["test_upload_enabled"],
            "subscription_status": r["subscription_status"],
            "subscription_period_end": r["current_period_end"].isoformat() if r["current_period_end"] else None,
            "primary_color": r["brand_primary_color"], "logo_url": r["brand_logo_url"],
            "directory_opt_in": r["directory_opt_in"],
            "directory_admin_override": r["directory_admin_override"],
            "directory_admin_reason": r["directory_admin_reason"],
            "directory_region": r["directory_region"],
            "directory_featured": r["directory_featured"],
            "directory_sort_order": r["directory_sort_order"],
            "is_active": r["is_active"],
            "deleted_at": r["deleted_at"].isoformat() if r["deleted_at"] else None,
            "max_groups": r["max_groups"] or DEFAULT_MAX_GROUPS_PER_CENTER,
            "max_students": r["max_students"] or DEFAULT_MAX_STUDENTS_PER_CENTER,
            "created_at": r["created_at"].isoformat(),
            "owner_id": r["owner_id"], "owner_name": r["owner_name"], "owner_email": r["owner_email"],
            "groups_count": r["groups_count"], "students_count": r["students_count"],
        }
        for r in rows
    ]


@app.get("/api/admin/centers/{center_id}/detail")
async def admin_center_detail(center_id: int, _: None = Depends(require_admin)):
    """Public katalog ma'lumotlarini admin uchun visibility cheklovisiz ko'rsatadi."""
    db = await get_pool()
    async with db.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT c.id, c.name, c.organization_type, c.slug, c.brand_name,
                   c.brand_primary_color, c.brand_logo_url, c.brand_contact_email,
                   c.brand_contact_phone, c.directory_description, c.directory_region,
                   c.directory_address, c.directory_website_url, c.directory_telegram_url,
                   c.directory_instagram_url, c.directory_show_email, c.directory_show_phone,
                   c.directory_show_address, c.directory_show_statistics,
                   c.directory_show_testimonials, c.directory_featured,
                   c.directory_opt_in, c.directory_admin_override,
                   c.subscription_required, c.test_upload_enabled,
                   c.is_active, c.deleted_at, c.max_groups, c.max_students, c.created_at,
                   c.owner_id, owner.full_name AS owner_name, owner.email AS owner_email,
                   sub.status AS subscription_status, sub.current_period_end,
                   COUNT(DISTINCT g.id) FILTER (WHERE g.deleted_at IS NULL) AS groups_count,
                   COUNT(DISTINCT u.id) FILTER (
                       WHERE u.role='student' AND u.deleted_at IS NULL
                   ) AS students_count
            FROM centers c
            LEFT JOIN users owner ON owner.id=c.owner_id
            LEFT JOIN organization_subscriptions sub ON sub.center_id=c.id
            LEFT JOIN groups g ON g.center_id=c.id
            LEFT JOIN users u ON u.center_id=c.id
            WHERE c.id=$1
            GROUP BY c.id, owner.full_name, owner.email,
                     sub.status, sub.current_period_end
            """,
            center_id
        )
        if not row:
            raise HTTPException(status_code=404, detail="Tashkilot topilmadi")

        feedback_rows = await conn.fetch(
            """
            SELECT t.content, t.show_full_name, u.full_name, u.username, u.role
            FROM testimonials t
            JOIN users u ON u.id=t.user_id
            WHERE u.center_id=$1 AND t.status='published' AND u.deleted_at IS NULL
            ORDER BY t.featured DESC, t.sort_order, t.published_at DESC NULLS LAST
            LIMIT 6
            """,
            center_id
        )

    directory_public = (
        row["is_active"] is True
        and row["deleted_at"] is None
        and bool(row["slug"])
        and (
            row["directory_admin_override"] == "force_public"
            or (
                row["directory_admin_override"] == "inherit"
                and row["directory_opt_in"] is True
            )
        )
    )
    return {
        "id": row["id"],
        "internal_name": row["name"],
        "name": row["brand_name"] or row["name"],
        "organization_type": row["organization_type"],
        "slug": row["slug"],
        "description": row["directory_description"],
        "region": row["directory_region"],
        "address": row["directory_address"],
        "logo_url": row["brand_logo_url"],
        "primary_color": row["brand_primary_color"] or "#1a56e8",
        "contact_email": row["brand_contact_email"],
        "contact_phone": row["brand_contact_phone"],
        "website_url": row["directory_website_url"],
        "telegram_url": row["directory_telegram_url"],
        "instagram_url": row["directory_instagram_url"],
        "show_email": row["directory_show_email"] is True,
        "show_phone": row["directory_show_phone"] is True,
        "show_address": row["directory_show_address"] is True,
        "show_statistics": row["directory_show_statistics"] is True,
        "show_testimonials": row["directory_show_testimonials"] is True,
        "featured": row["directory_featured"] is True,
        "directory_public": directory_public,
        "directory_mode": row["directory_admin_override"],
        "is_active": row["is_active"],
        "deleted_at": row["deleted_at"].isoformat() if row["deleted_at"] else None,
        "groups_count": row["groups_count"],
        "students_count": row["students_count"],
        "max_groups": row["max_groups"] or DEFAULT_MAX_GROUPS_PER_CENTER,
        "max_students": row["max_students"] or DEFAULT_MAX_STUDENTS_PER_CENTER,
        "owner_name": row["owner_name"],
        "owner_email": row["owner_email"],
        "subscription_required": row["subscription_required"],
        "subscription_status": row["subscription_status"],
        "subscription_period_end": (
            row["current_period_end"].isoformat() if row["current_period_end"] else None
        ),
        "test_upload_enabled": row["test_upload_enabled"],
        "created_at": row["created_at"].isoformat(),
        "public_url": f"/org/{row['slug']}" if directory_public else None,
        "testimonials": [
            {
                "content": item["content"],
                "full_name": (
                    (item["full_name"] or item["username"])
                    if item["show_full_name"] else "Tasdiqlangan foydalanuvchi"
                ),
                "role": testimonial_role_label(item["role"], row["organization_type"]),
            }
            for item in feedback_rows
        ],
    }


@app.post("/api/admin/centers/{center_id}/directory")
async def admin_update_center_directory(
    center_id: int,
    data: AdminDirectoryOverrideIn,
    _: None = Depends(require_admin)
):
    if data.mode not in {"inherit", "force_public", "force_hidden"}:
        raise HTTPException(status_code=400, detail="Katalog override holati noto'g'ri")
    reason = (data.reason or "").strip() or None
    if data.mode != "inherit" and not reason:
        raise HTTPException(status_code=400, detail="Admin override sababini yozing")
    if reason and len(reason) > 500:
        raise HTTPException(status_code=400, detail="Override sababi juda uzun")
    slug = (data.slug or "").strip().lower() or None
    if slug and not SLUG_RE.fullmatch(slug):
        raise HTTPException(status_code=400, detail="Slug faqat kichik harf, raqam va tirelardan iborat bo'lishi kerak")
    if not 0 <= data.sort_order <= 10000:
        raise HTTPException(status_code=400, detail="Tartib raqami 0 dan 10000 gacha bo'lishi kerak")
    db = await get_pool()
    async with db.acquire() as conn:
        current_slug = await conn.fetchval("SELECT slug FROM centers WHERE id=$1", center_id)
        if current_slug is None and not await conn.fetchval("SELECT EXISTS(SELECT 1 FROM centers WHERE id=$1)", center_id):
            raise HTTPException(status_code=404, detail="Tashkilot topilmadi")
        if data.mode == "force_public" and not (slug or current_slug):
            raise HTTPException(status_code=400, detail="Public qilish uchun slug kiriting")
        if slug:
            duplicate = await conn.fetchval(
                "SELECT 1 FROM centers WHERE LOWER(slug)=LOWER($1) AND id<>$2", slug, center_id
            )
            if duplicate:
                raise HTTPException(status_code=409, detail="Bu public slug band qilingan")
        row = await conn.fetchrow(
            """
            UPDATE centers SET
                directory_admin_override=$2, directory_admin_reason=$3,
                slug=COALESCE($4, slug), directory_featured=$5,
                directory_sort_order=$6, directory_override_updated_at=NOW()
            WHERE id=$1
            RETURNING id, slug, directory_opt_in, directory_admin_override,
                      directory_admin_reason, directory_featured, directory_sort_order
            """,
            center_id, data.mode, reason, slug, data.featured, data.sort_order
        )
    return dict(row)


@app.post("/api/admin/centers/{center_id}/toggle-test-upload")
async def admin_toggle_test_upload(center_id: int, _: None = Depends(require_admin)):
    db = await get_pool()
    async with db.acquire() as conn:
        row = await conn.fetchrow("UPDATE centers SET test_upload_enabled=NOT test_upload_enabled WHERE id=$1 RETURNING test_upload_enabled", center_id)
    if not row:
        raise HTTPException(status_code=404, detail="Tashkilot topilmadi")
    return {"test_upload_enabled": row["test_upload_enabled"]}


@app.post("/api/admin/centers/{center_id}/toggle-subscription")
async def admin_toggle_subscription(center_id: int, _: None = Depends(require_admin)):
    db = await get_pool()
    async with db.acquire() as conn:
        async with conn.transaction():
            center = await conn.fetchrow(
                """
                UPDATE centers SET subscription_required=NOT subscription_required
                WHERE id=$1 RETURNING id, subscription_required
                """,
                center_id
            )
            if not center:
                raise HTTPException(status_code=404, detail="Tashkilot topilmadi")
            if center["subscription_required"]:
                await conn.execute(
                    """
                    INSERT INTO organization_subscriptions(center_id, status, trial_ends_at, updated_at)
                    VALUES ($1, 'trial', NOW() + INTERVAL '14 days', NOW())
                    ON CONFLICT(center_id) DO NOTHING
                    """,
                    center_id
                )
    return {
        "subscription_required": center["subscription_required"],
        "message": "Obuna talabi yoqildi" if center["subscription_required"] else "Obuna talabi o'chirildi",
    }


async def _own_speaking_result(conn, result_id: int, user_id: int):
    return await conn.fetchrow(
        """
        SELECT er.id,er.speaking_audio_data,er.speaking_audio_mime,er.speaking_audio_filename
        FROM exam_results er JOIN users u ON u.email=er.email
        WHERE er.id=$1 AND u.id=$2 AND er.section='speaking'
        """, result_id, user_id
    )


@app.get("/api/student/speaking/{result_id}/audio")
async def student_speaking_audio(result_id: int, current_user: dict = Depends(get_current_user)):
    db = await get_pool()
    async with db.acquire() as conn:
        row = await _own_speaking_result(conn, result_id, current_user["id"])
    if not row or not row["speaking_audio_data"]:
        raise HTTPException(status_code=404, detail="Audio topilmadi")
    return Response(content=bytes(row["speaking_audio_data"]), media_type=row["speaking_audio_mime"] or "audio/ogg",
                    headers={"Content-Disposition": f'inline; filename="speaking-{result_id}"'})


@app.get("/api/student/speaking/{result_id}/comments")
async def student_speaking_comments(result_id: int, current_user: dict = Depends(get_current_user)):
    db = await get_pool()
    async with db.acquire() as conn:
        if not await _own_speaking_result(conn, result_id, current_user["id"]):
            raise HTTPException(status_code=404, detail="Speaking natijasi topilmadi")
        rows = await conn.fetch(
            "SELECT id,timestamp_seconds,comment,created_at FROM speaking_timed_comments WHERE result_id=$1 ORDER BY timestamp_seconds,id",
            result_id
        )
    return [dict(r) | {"timestamp_seconds": float(r["timestamp_seconds"]), "created_at": r["created_at"].isoformat()} for r in rows]


@app.get("/api/admin/subscription-payments")
async def admin_subscription_payments(status: Optional[str] = None, _: None = Depends(require_admin)):
    db = await get_pool()
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT pay.id, pay.order_code, pay.billing_cycle, pay.amount, pay.payer_name,
                   pay.transaction_reference, pay.status, pay.review_note, pay.created_at,
                   pay.receipt_mime, c.id AS center_id, c.name AS center_name, p.name AS plan_name
            FROM subscription_payments pay
            JOIN centers c ON c.id=pay.center_id
            JOIN subscription_plans p ON p.id=pay.plan_id
            WHERE ($1::text IS NULL OR pay.status=$1)
            ORDER BY CASE WHEN pay.status='pending' THEN 0 ELSE 1 END, pay.created_at DESC
            LIMIT 500
            """,
            status
        )

    return [dict(row) | {"created_at": row["created_at"].isoformat()} for row in rows]


@app.get("/api/admin/subscription-payments/{payment_id}/receipt")
async def admin_subscription_receipt(payment_id: int, _: None = Depends(require_admin)):
    db = await get_pool()
    async with db.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT receipt_data, receipt_mime FROM subscription_payments WHERE id=$1", payment_id
        )
    if not row:
        raise HTTPException(status_code=404, detail="To'lov topilmadi")
    return Response(content=bytes(row["receipt_data"]), media_type=row["receipt_mime"])


@app.post("/api/admin/subscription-payments/{payment_id}/review")
async def admin_review_subscription_payment(
    payment_id: int,
    data: AdminPaymentReviewIn,
    _: None = Depends(require_admin)
):
    if data.action not in {"approve", "reject"}:
        raise HTTPException(status_code=400, detail="Amal noto'g'ri")
    note = data.note.strip()[:500] if data.note else None
    db = await get_pool()
    async with db.acquire() as conn:
        async with conn.transaction():
            payment = await conn.fetchrow(
                "SELECT * FROM subscription_payments WHERE id=$1 FOR UPDATE", payment_id
            )
            if not payment:
                raise HTTPException(status_code=404, detail="To'lov topilmadi")
            if payment["status"] != "pending":
                raise HTTPException(status_code=409, detail="To'lov oldin ko'rib chiqilgan")
            new_status = "approved" if data.action == "approve" else "rejected"
            await conn.execute(
                """
                UPDATE subscription_payments
                SET status=$1, review_note=$2, reviewed_at=NOW() WHERE id=$3
                """,
                new_status, note, payment_id
            )
            if data.action == "approve":
                await conn.execute(
                    """
                    INSERT INTO organization_subscriptions(
                        center_id, plan_id, status, current_period_start, current_period_end, grace_ends_at, updated_at
                    ) VALUES (
                        $1, $2, 'active', NOW(),
                        NOW() + CASE WHEN $3='yearly' THEN INTERVAL '1 year' ELSE INTERVAL '1 month' END,
                        NOW() + CASE WHEN $3='yearly' THEN INTERVAL '1 year 7 days' ELSE INTERVAL '1 month 7 days' END,
                        NOW()
                    )
                    ON CONFLICT(center_id) DO UPDATE SET
                        plan_id=EXCLUDED.plan_id,
                        status='active',
                        current_period_start=GREATEST(COALESCE(organization_subscriptions.current_period_end, NOW()), NOW()),
                        current_period_end=GREATEST(COALESCE(organization_subscriptions.current_period_end, NOW()), NOW())
                            + CASE WHEN $3='yearly' THEN INTERVAL '1 year' ELSE INTERVAL '1 month' END,
                        grace_ends_at=GREATEST(COALESCE(organization_subscriptions.current_period_end, NOW()), NOW())
                            + CASE WHEN $3='yearly' THEN INTERVAL '1 year 7 days' ELSE INTERVAL '1 month 7 days' END,
                        updated_at=NOW()
                    """,
                    payment["center_id"], payment["plan_id"], payment["billing_cycle"]
                )
            owner_id = await conn.fetchval("SELECT owner_id FROM centers WHERE id=$1", payment["center_id"])
            if owner_id:
                await create_notification(
                    conn,
                    recipient_user_id=owner_id,
                    kind="success" if data.action == "approve" else "warning",
                    title="Obuna to'lovi ko'rib chiqildi",
                    message=("To'lov tasdiqlandi va obuna faollashtirildi." if data.action == "approve" else "To'lov rad etildi.")
                            + (f" Izoh: {note}" if note else ""),
                    action_url="/head-teacher",
                    metadata={"event": "payment_review", "payment_id": payment_id, "status": new_status},
                )
    return {"message": "To'lov tasdiqlandi" if data.action == "approve" else "To'lov rad etildi"}


def public_organization_payload(row: dict) -> dict:
    show_stats = row["directory_show_statistics"] is True
    return {
        "id": row["id"],
        "slug": row["slug"],
        "name": row["brand_name"] or row["name"],
        "organization_type": row["organization_type"],
        "description": row["directory_description"],
        "region": row["directory_region"],
        "address": row["directory_address"] if row["directory_show_address"] else None,
        "logo_url": row["brand_logo_url"],
        "primary_color": row["brand_primary_color"] or "#1a56e8",
        "contact_email": row["brand_contact_email"] if row["directory_show_email"] else None,
        "contact_phone": row["brand_contact_phone"] if row["directory_show_phone"] else None,
        "website_url": row["directory_website_url"],
        "telegram_url": row["directory_telegram_url"],
        "instagram_url": row["directory_instagram_url"],
        "students_count": row["students_count"] if show_stats else None,
        "groups_count": row["groups_count"] if show_stats else None,
        "featured": row["directory_featured"] is True,
        "show_testimonials": row["directory_show_testimonials"] is True,
    }


@app.get("/api/organizations/public")
async def public_organizations(organization_type: Optional[str] = None, q: Optional[str] = None):
    if organization_type and organization_type not in ORGANIZATION_TYPES:
        raise HTTPException(status_code=400, detail="Tashkilot turi noto'g'ri")
    search = (q or "").strip()[:100] or None
    db = await get_pool()
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT c.id, c.name, c.organization_type, c.slug, c.brand_name,
                   c.brand_primary_color, c.brand_logo_url, c.brand_contact_email,
                   c.brand_contact_phone, c.directory_description, c.directory_region,
                   c.directory_address, c.directory_website_url, c.directory_telegram_url,
                   c.directory_instagram_url, c.directory_show_email, c.directory_show_phone,
                   c.directory_show_address, c.directory_show_statistics,
                   c.directory_show_testimonials, c.directory_featured,
                   COUNT(DISTINCT g.id) FILTER (WHERE g.deleted_at IS NULL) AS groups_count,
                   COUNT(DISTINCT u.id) FILTER (WHERE u.role='student' AND u.deleted_at IS NULL) AS students_count
            FROM centers c
            LEFT JOIN groups g ON g.center_id=c.id
            LEFT JOIN users u ON u.center_id=c.id
            WHERE c.is_active=TRUE AND c.deleted_at IS NULL AND c.slug IS NOT NULL
              AND c.directory_admin_override<>'force_hidden'
              AND (c.directory_admin_override='force_public' OR
                   (c.directory_admin_override='inherit' AND c.directory_opt_in=TRUE))
              AND ($1::text IS NULL OR c.organization_type=$1)
              AND ($2::text IS NULL OR c.name ILIKE '%' || $2 || '%' OR
                   COALESCE(c.brand_name,'') ILIKE '%' || $2 || '%' OR
                   COALESCE(c.directory_region,'') ILIKE '%' || $2 || '%')
            GROUP BY c.id
            ORDER BY c.directory_featured DESC, c.directory_sort_order ASC, c.name
            LIMIT 200
            """,
            organization_type, search
        )
    return [public_organization_payload(dict(row)) for row in rows]


@app.get("/api/organizations/public/{slug}")
async def public_organization_detail(slug: str):
    db = await get_pool()
    async with db.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT c.id, c.name, c.organization_type, c.slug, c.brand_name,
                   c.brand_primary_color, c.brand_logo_url, c.brand_contact_email,
                   c.brand_contact_phone, c.directory_description, c.directory_region,
                   c.directory_address, c.directory_website_url, c.directory_telegram_url,
                   c.directory_instagram_url, c.directory_show_email, c.directory_show_phone,
                   c.directory_show_address, c.directory_show_statistics,
                   c.directory_show_testimonials, c.directory_featured,
                   COUNT(DISTINCT g.id) FILTER (WHERE g.deleted_at IS NULL) AS groups_count,
                   COUNT(DISTINCT u.id) FILTER (WHERE u.role='student' AND u.deleted_at IS NULL) AS students_count
            FROM centers c
            LEFT JOIN groups g ON g.center_id=c.id
            LEFT JOIN users u ON u.center_id=c.id
            WHERE LOWER(c.slug)=LOWER($1) AND c.is_active=TRUE AND c.deleted_at IS NULL
              AND c.directory_admin_override<>'force_hidden'
              AND (c.directory_admin_override='force_public' OR
                   (c.directory_admin_override='inherit' AND c.directory_opt_in=TRUE))
            GROUP BY c.id
            """,
            slug.strip()
        )
        if not row:
            raise HTTPException(status_code=404, detail="Tashkilot topilmadi")
        payload = public_organization_payload(dict(row))
        testimonials = []
        if row["directory_show_testimonials"]:
            feedback_rows = await conn.fetch(
                """
                SELECT t.content, t.show_full_name, t.show_avatar,
                       u.full_name, u.username, u.role, u.avatar_mime
                FROM testimonials t JOIN users u ON u.id=t.user_id
                WHERE u.center_id=$1 AND t.status='published' AND u.deleted_at IS NULL
                ORDER BY t.featured DESC, t.sort_order, t.published_at DESC NULLS LAST LIMIT 6
                """,
                row["id"]
            )
            testimonials = [{
                "content": item["content"],
                "full_name": (item["full_name"] or item["username"]) if item["show_full_name"] else "Tasdiqlangan foydalanuvchi",
                "role": testimonial_role_label(item["role"], row["organization_type"]),
                "avatar_url": f"/api/auth/avatar/{item['username']}" if item["show_avatar"] and item["avatar_mime"] else None,
            } for item in feedback_rows]
        payload["testimonials"] = testimonials
    return payload


@app.get("/api/branding/{slug}")
async def public_branding(slug: str):
    db = await get_pool()
    async with db.acquire() as conn:
        row = await conn.fetchrow(
            """
            SELECT id, name, organization_type, slug, brand_name, brand_primary_color,
                   brand_secondary_color, brand_logo_url, brand_favicon_url,
                   brand_contact_email, brand_contact_phone, show_powered_by,
                   directory_opt_in, directory_admin_override, directory_description,
                   directory_region, directory_address, directory_website_url,
                   directory_telegram_url, directory_instagram_url, directory_show_email,
                   directory_show_phone, directory_show_address, directory_show_statistics,
                   directory_show_testimonials
            FROM centers WHERE LOWER(slug)=LOWER($1) AND is_active=TRUE
            """,
            slug.strip()
        )
    if not row:
        raise HTTPException(status_code=404, detail="Tashkilot topilmadi")
    return branding_payload(row)


@app.post("/api/admin/centers/{center_id}/assign-head-teacher")
async def admin_assign_head_teacher(center_id: int, data: AssignHeadTeacherIn, _: None = Depends(require_admin)):
    db = await get_pool()
    async with db.acquire() as conn:
        async with conn.transaction():
            center = await conn.fetchrow("SELECT id FROM centers WHERE id=$1 FOR UPDATE", center_id)
            if not center:
                raise HTTPException(status_code=404, detail="Markaz topilmadi")

            user = await conn.fetchrow("SELECT id, role FROM users WHERE id=$1", data.user_id)
            if not user:
                raise HTTPException(status_code=404, detail="Foydalanuvchi topilmadi")
            if user["role"] != "student":
                raise HTTPException(
                    status_code=400,
                    detail="Faqat oddiy (student) hisob head-teacher qilib tayinlanishi mumkin"
                )

            await conn.execute(
                "UPDATE users SET role='head_teacher', center_id=$1 WHERE id=$2", center_id, data.user_id
            )
            await conn.execute("UPDATE centers SET owner_id=$1 WHERE id=$2", data.user_id, center_id)

    return {"message": "Head-teacher tayinlandi"}


@app.post("/api/admin/centers/{center_id}/deactivate")
async def admin_toggle_center(center_id: int, _: None = Depends(require_admin)):
    db = await get_pool()
    async with db.acquire() as conn:
        row = await conn.fetchrow(
            "UPDATE centers SET is_active = NOT is_active WHERE id=$1 RETURNING id, is_active",
            center_id
        )
        if not row:
            raise HTTPException(status_code=404, detail="Markaz topilmadi")
    return {
        "message": "Markaz yopildi" if not row["is_active"] else "Markaz qayta ochildi",
        "is_active": row["is_active"]
    }


ADMIN_ENTITY_CONFIG = {
    "group": ("groups", "name", "is_active"),
    "class": ("school_classes", "name", "is_active"),
    "staff": ("school_staff", "employee_code", "is_active"),
    "subject": ("school_subjects", "name", "is_active"),
    "position": ("school_positions", "name", None),
    "user": ("users", "full_name", "is_suspended"),
    "test": ("tests", "title", None),
}


@app.get("/api/admin/centers/{center_id}/inventory")
async def admin_center_inventory(center_id: int, _: None = Depends(require_admin)):
    """Tashkilot tarkibini bitta xavfsiz boshqaruv oynasida ko'rsatadi."""
    db = await get_pool()
    async with db.acquire() as conn:
        center = await conn.fetchrow("SELECT id, name FROM centers WHERE id=$1", center_id)
        if not center:
            raise HTTPException(status_code=404, detail="Tashkilot topilmadi")
        groups = await conn.fetch("SELECT id,name,is_active,deleted_at FROM groups WHERE center_id=$1 ORDER BY name", center_id)
        classes = await conn.fetch("SELECT id,name,academic_year,is_active,deleted_at FROM school_classes WHERE center_id=$1 ORDER BY name", center_id)
        subjects = await conn.fetch("SELECT id,name,is_active,deleted_at FROM school_subjects WHERE center_id=$1 ORDER BY name", center_id)
        positions = await conn.fetch("SELECT id,name,deleted_at FROM school_positions WHERE center_id=$1 ORDER BY name", center_id)
        staff = await conn.fetch("""
            SELECT ss.id, COALESCE(u.full_name,ss.employee_code,'Xodim') AS name,
                   ss.is_active,ss.deleted_at
            FROM school_staff ss LEFT JOIN users u ON u.id=ss.user_id
            WHERE ss.center_id=$1 ORDER BY name
        """, center_id)
        users = await conn.fetch("SELECT id,full_name AS name,email,role,is_suspended,deleted_at FROM users WHERE center_id=$1 ORDER BY full_name", center_id)
        tests = await conn.fetch("SELECT id,title AS name,status,deleted_at FROM tests WHERE center_id=$1 ORDER BY title", center_id)

    def serialize(rows):
        return [dict(r) | {"deleted_at": r["deleted_at"].isoformat() if r["deleted_at"] else None} for r in rows]
    return {"center": dict(center), "groups": serialize(groups), "classes": serialize(classes),
            "subjects": serialize(subjects), "positions": serialize(positions),
            "staff": serialize(staff), "users": serialize(users), "tests": serialize(tests)}


@app.post("/api/admin/centers/{center_id}/trash")
async def admin_trash_center(center_id: int, _: None = Depends(require_admin)):
    db = await get_pool()
    async with db.acquire() as conn:
        row = await conn.fetchrow("UPDATE centers SET deleted_at=NOW(),is_active=FALSE WHERE id=$1 RETURNING name", center_id)
    if not row:
        raise HTTPException(status_code=404, detail="Tashkilot topilmadi")
    return {"message": f"{row['name']} savatga o'tkazildi"}


@app.post("/api/admin/centers/{center_id}/restore")
async def admin_restore_center(center_id: int, _: None = Depends(require_admin)):
    db = await get_pool()
    async with db.acquire() as conn:
        row = await conn.fetchrow("UPDATE centers SET deleted_at=NULL,is_active=TRUE WHERE id=$1 RETURNING name", center_id)
    if not row:
        raise HTTPException(status_code=404, detail="Tashkilot topilmadi")
    return {"message": f"{row['name']} tiklandi"}


@app.delete("/api/admin/centers/{center_id}")
async def admin_purge_center(center_id: int, confirm_name: str, _: None = Depends(require_admin)):
    db = await get_pool()
    async with db.acquire() as conn:
        async with conn.transaction():
            center = await conn.fetchrow("SELECT name,deleted_at FROM centers WHERE id=$1 FOR UPDATE", center_id)
            if not center:
                raise HTTPException(status_code=404, detail="Tashkilot topilmadi")
            if not center["deleted_at"]:
                raise HTTPException(status_code=409, detail="Avval tashkilotni savatga o'tkazing")
            if confirm_name.strip() != center["name"]:
                raise HTTPException(status_code=400, detail="Tashkilot nomi mos kelmadi")
            await conn.execute("UPDATE users SET role='student',center_id=NULL,group_id=NULL WHERE center_id=$1", center_id)
            await conn.execute("DELETE FROM centers WHERE id=$1", center_id)
    return {"message": "Tashkilot va unga bog'liq ma'lumotlar butunlay o'chirildi"}


@app.post("/api/admin/entities/{entity_type}/{entity_id}/trash")
async def admin_trash_entity(entity_type: str, entity_id: int, _: None = Depends(require_admin)):
    config = ADMIN_ENTITY_CONFIG.get(entity_type)
    if not config:
        raise HTTPException(status_code=400, detail="Obyekt turi noto'g'ri")
    table, label_col, active_col = config
    if entity_type == "user":
        sql = f"UPDATE {table} SET deleted_at=NOW(),is_suspended=TRUE WHERE id=$1 RETURNING {label_col} AS label"
    elif entity_type == "test":
        sql = f"UPDATE {table} SET deleted_at=NOW(),status='archived' WHERE id=$1 RETURNING {label_col} AS label"
    elif active_col:
        sql = f"UPDATE {table} SET deleted_at=NOW(),{active_col}=FALSE WHERE id=$1 RETURNING {label_col} AS label"
    else:
        sql = f"UPDATE {table} SET deleted_at=NOW() WHERE id=$1 RETURNING {label_col} AS label"
    db = await get_pool()
    async with db.acquire() as conn:
        row = await conn.fetchrow(sql, entity_id)
    if not row:
        raise HTTPException(status_code=404, detail="Obyekt topilmadi")
    return {"message": f"{row['label'] or 'Obyekt'} savatga o'tkazildi"}


@app.post("/api/admin/entities/{entity_type}/{entity_id}/restore")
async def admin_restore_entity(entity_type: str, entity_id: int, _: None = Depends(require_admin)):
    config = ADMIN_ENTITY_CONFIG.get(entity_type)
    if not config:
        raise HTTPException(status_code=400, detail="Obyekt turi noto'g'ri")
    table, label_col, active_col = config
    if entity_type == "user":
        sql = f"UPDATE {table} SET deleted_at=NULL,is_suspended=FALSE WHERE id=$1 RETURNING {label_col} AS label"
    elif entity_type == "test":
        sql = f"UPDATE {table} SET deleted_at=NULL,status='draft' WHERE id=$1 RETURNING {label_col} AS label"
    elif active_col:
        sql = f"UPDATE {table} SET deleted_at=NULL,{active_col}=TRUE WHERE id=$1 RETURNING {label_col} AS label"
    else:
        sql = f"UPDATE {table} SET deleted_at=NULL WHERE id=$1 RETURNING {label_col} AS label"
    db = await get_pool()
    async with db.acquire() as conn:
        row = await conn.fetchrow(sql, entity_id)
    if not row:
        raise HTTPException(status_code=404, detail="Obyekt topilmadi")
    return {"message": f"{row['label'] or 'Obyekt'} tiklandi"}


@app.delete("/api/admin/entities/{entity_type}/{entity_id}")
async def admin_purge_entity(entity_type: str, entity_id: int, _: None = Depends(require_admin)):
    config = ADMIN_ENTITY_CONFIG.get(entity_type)
    if not config:
        raise HTTPException(status_code=400, detail="Obyekt turi noto'g'ri")
    table, _, _ = config
    db = await get_pool()
    async with db.acquire() as conn:
        async with conn.transaction():
            row = await conn.fetchrow(f"SELECT deleted_at FROM {table} WHERE id=$1 FOR UPDATE", entity_id)
            if not row:
                raise HTTPException(status_code=404, detail="Obyekt topilmadi")
            if not row["deleted_at"]:
                raise HTTPException(status_code=409, detail="Avval obyektni savatga o'tkazing")
            if entity_type == "user":
                await conn.execute("UPDATE users SET referred_by=NULL WHERE referred_by=$1", entity_id)
            await conn.execute(f"DELETE FROM {table} WHERE id=$1", entity_id)
    return {"message": "Obyekt butunlay o'chirildi"}


@app.post("/api/admin/users/{user_id}/suspend")
async def admin_suspend_user(user_id: int, _: None = Depends(require_admin)):
    db = await get_pool()
    async with db.acquire() as conn:
        row = await conn.fetchrow(
            "UPDATE users SET is_suspended = NOT is_suspended WHERE id = $1 RETURNING id, is_suspended",
            user_id
        )
        if not row:
            raise HTTPException(status_code=404, detail="Foydalanuvchi topilmadi")

    return {
        "message": "Foydalanuvchi bloklandi" if row["is_suspended"] else "Foydalanuvchi blokdan chiqarildi",
        "is_suspended": row["is_suspended"]
    }


@app.delete("/api/admin/users/{user_id}")
async def admin_delete_user(user_id: int, _: None = Depends(require_admin)):
    db = await get_pool()
    async with db.acquire() as conn:
        row = await conn.fetchrow(
            "UPDATE users SET deleted_at=NOW(),is_suspended=TRUE WHERE id=$1 RETURNING id", user_id
        )
        if not row:
            raise HTTPException(status_code=404, detail="Foydalanuvchi topilmadi")

    return {"message": "Foydalanuvchi savatga o'tkazildi"}


@app.post("/api/admin/grade-writing")
async def grade_writing(data: GradeWriting, _: None = Depends(require_admin)):
    db = await get_pool()
    async with db.acquire() as conn:
        row = await conn.fetchrow(
            """
            UPDATE exam_results
            SET writing_band = $1, writing_feedback = $2,
                writing_task_achievement = $4, writing_coherence_cohesion = $5,
                writing_lexical_resource = $6, writing_grammar_accuracy = $7,
                grader_name = $8, writing_graded_at = NOW()
            WHERE id = $3
            RETURNING full_name, email, section
            """,
            data.band, data.feedback, data.result_id,
            data.task_achievement, data.coherence_cohesion, data.lexical_resource, data.grammar_accuracy, "Admin"
        )
        if not row:
            raise HTTPException(status_code=404, detail="Natija topilmadi")

    if data.send_email:
        try:
            html = build_result_email(row["full_name"], "writing", None, None, data.band, data.feedback)
            
            async with db.acquire() as conn:
                rows = await conn.fetch(
                    """
                    SELECT DISTINCT ON (section) full_name, section, score, total, writing_band
                    FROM exam_results
                    WHERE email = $1
                    ORDER BY section, submitted_at DESC
                    """,
                    row["email"].lower()
                )
            sections = []
            for r in rows:
                if r["section"] == "writing":
                    r_band = float(r["writing_band"]) if r["writing_band"] is not None else None
                elif r["score"] is not None and r["total"] is not None:
                    r_band = get_band_score(r["score"], r["total"], r["section"])
                else:
                    r_band = None
                sections.append({
                    "section": r["section"],
                    "score": r["score"],
                    "total": r["total"],
                    "band": r_band
                })
                
            pdf_bytes = build_result_pdf(row["full_name"], row["email"].lower(), sections)
            await send_email_with_attachment(
                row["email"], row["full_name"], "IELTS Mock SS — Writing natijangiz tayyor",
                html, pdf_bytes, f"ielts_natija_{row['email'].lower().split('@')[0]}.pdf"
            )
            
            async with db.acquire() as conn:
                await conn.execute("UPDATE exam_results SET notified = TRUE WHERE id = $1", data.result_id)
                user_row = await conn.fetchrow("SELECT telegram_chat_id FROM users WHERE email = $1", row["email"].lower())
                
            if user_row and user_row["telegram_chat_id"]:
                from telegram import notify_user_writing_graded
                await notify_user_writing_graded(user_row["telegram_chat_id"], row["full_name"], data.band, data.feedback)
                
        except Exception as e:
            return {"message": "Baho saqlandi, lekin email yuborilmadi", "error": str(e)}

    return {"message": "Baho saqlandi va email yuborildi"}


@app.post("/api/admin/notify/{result_id}")
async def notify_result(result_id: int, _: None = Depends(require_admin)):
    """Listening/Reading natijasini emailga yuborish"""
    db = await get_pool()
    async with db.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT full_name, email, section, score, total FROM exam_results WHERE id = $1",
            result_id
        )
        if not row:
            raise HTTPException(status_code=404, detail="Natija topilmadi")

        rows = await conn.fetch(
            """
            SELECT DISTINCT ON (section) full_name, section, score, total, writing_band
            FROM exam_results
            WHERE email = $1
            ORDER BY section, submitted_at DESC
            """,
            row["email"].lower()
        )
        sections = []
        for r in rows:
            if r["section"] == "writing":
                r_band = float(r["writing_band"]) if r["writing_band"] is not None else None
            elif r["score"] is not None and r["total"] is not None:
                r_band = get_band_score(r["score"], r["total"], r["section"])
            else:
                r_band = None
            sections.append({
                "section": r["section"],
                "score": r["score"],
                "total": r["total"],
                "band": r_band
            })

    band = get_band_score(row["score"], row["total"], row["section"]) if row["score"] is not None and row["total"] is not None else None
    try:
        pdf_bytes = build_result_pdf(row["full_name"], row["email"].lower(), sections)
        html = build_result_email(row["full_name"], row["section"], row["score"], row["total"], band)
        
        await send_email_with_attachment(
            row["email"], row["full_name"], f"IELTS Mock SS — {row['section'].capitalize()} natijangiz",
            html, pdf_bytes, f"ielts_natija_{row['email'].lower().split('@')[0]}.pdf"
        )
        
        async with db.acquire() as conn:
            await conn.execute("UPDATE exam_results SET notified = TRUE WHERE id = $1", result_id)
            user_row = await conn.fetchrow("SELECT telegram_chat_id FROM users WHERE email = $1", row["email"].lower())
            
        if user_row and user_row["telegram_chat_id"]:
            from telegram import notify_user_result_ready
            await notify_user_result_ready(user_row["telegram_chat_id"], row["full_name"], row["section"], band, row["score"], row["total"])
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Email yuborilmadi: {str(e)}")

    return {"message": "Email yuborildi"}


@app.get("/api/admin/stats")
async def admin_stats(days: int = 30, _: None = Depends(require_admin)):
    db = await get_pool()
    async with db.acquire() as conn:
        total_users = await conn.fetchval("SELECT COUNT(*) FROM users")
        verified_users = await conn.fetchval("SELECT COUNT(*) FROM users WHERE email_verified = TRUE")
        total_results = await conn.fetchval("SELECT COUNT(*) FROM exam_results")
        active_today = await conn.fetchval("SELECT COUNT(DISTINCT email) FROM exam_results WHERE DATE(submitted_at) = CURRENT_DATE")
        
        daily_reg_rows = await conn.fetch(
            """
            SELECT DATE(created_at) as date, COUNT(*) as count 
            FROM users 
            WHERE created_at >= NOW() - INTERVAL '30 days'
            GROUP BY DATE(created_at) 
            ORDER BY date
            """
        )
        
        daily_act_rows = await conn.fetch(
            """
            SELECT DATE(submitted_at) as date, COUNT(DISTINCT email) as count 
            FROM exam_results 
            WHERE submitted_at >= NOW() - INTERVAL '30 days'
            GROUP BY DATE(submitted_at) 
            ORDER BY date
            """
        )

        sec_rows = await conn.fetch("SELECT section, COUNT(*) as count FROM exam_results GROUP BY section")
        all_res = await conn.fetch("SELECT section, score, total, writing_band FROM exam_results")
    
    overview = {
        "total_users": total_users,
        "verified_users": verified_users,
        "total_results": total_results,
        "active_today": active_today
    }
    
    daily_registrations = [{"date": str(r["date"]), "count": r["count"]} for r in daily_reg_rows]
    daily_active_users = [{"date": str(r["date"]), "count": r["count"]} for r in daily_act_rows]
    section_attempts = {r["section"]: r["count"] for r in sec_rows}
    
    band_distribution = {"listening": {}, "reading": {}, "writing": {}}
    for r in all_res:
        s = r["section"]
        if s == "writing":
            b = float(r["writing_band"]) if r["writing_band"] is not None else None
        elif r["score"] is not None and r["total"] is not None:
            b = get_band_score(r["score"], r["total"], s)
        else:
            b = None
            
        if b is not None:
            b_str = str(b)
            band_distribution[s][b_str] = band_distribution[s].get(b_str, 0) + 1

    return {
        "overview": overview,
        "daily_registrations": daily_registrations,
        "daily_active_users": daily_active_users,
        "section_attempts": section_attempts,
        "band_distribution": band_distribution
    }


@app.get("/api/admin/export/results")
async def export_results(format: str = "csv", _: None = Depends(require_admin)):
    db = await get_pool()
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT id, full_name, email, section, score, total,
                   writing_band, submitted_at
            FROM exam_results
            ORDER BY submitted_at DESC
            """
        )
        
    data = []
    for r in rows:
        b = float(r["writing_band"]) if r["section"] == "writing" and r["writing_band"] is not None else get_band_score(r["score"], r["total"], r["section"]) if r["section"] != "writing" and r["score"] is not None and r["total"] is not None else ""
        data.append({
            "ID": r["id"],
            "Ism": r["full_name"],
            "Email": r["email"],
            "Bo'lim": r["section"],
            "Natija": f"{r['score']}/{r['total']}" if r["score"] is not None else "",
            "Band": b,
            "Sana": r["submitted_at"].isoformat()
        })
        
    if format == "csv":
        output = io.StringIO()
        if data:
            writer = csv.DictWriter(output, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        output.seek(0)
        return StreamingResponse(
            iter([b'\xef\xbb\xbf' + output.getvalue().encode('utf-8')]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=results.csv"}
        )
    elif format == "excel":
        if not openpyxl:
            raise HTTPException(status_code=500, detail="openpyxl o'rnatilmagan")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"
        if data:
            ws.append(list(data[0].keys()))
            for d in data:
                ws.append(list(d.values()))
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=results.xlsx"}
        )

@app.get("/api/admin/export/users")
async def export_users(format: str = "csv", _: None = Depends(require_admin)):
    db = await get_pool()
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT u.id, u.username, u.email, u.full_name, u.created_at,
                   u.email_verified, u.is_suspended, u.telegram_chat_id, u.referral_count,
                   COUNT(er.id) AS attempts
            FROM users u
            LEFT JOIN exam_results er ON er.email = u.email
            GROUP BY u.id
            ORDER BY u.created_at DESC
            """
        )
        
    data = []
    for r in rows:
        data.append({
            "ID": r["id"],
            "Username": r["username"],
            "Email": r["email"],
            "To'liq Ism": r["full_name"],
            "Sana": r["created_at"].isoformat(),
            "Tasdiqlangan": "Ha" if r["email_verified"] else "Yo'q",
            "Bloklangan": "Ha" if r["is_suspended"] else "Yo'q",
            "Urinishlar": r["attempts"],
            "Telegram": r["telegram_chat_id"] or "",
            "Taklif qildi": r["referral_count"] or 0
        })
        
    if format == "csv":
        output = io.StringIO()
        if data:
            writer = csv.DictWriter(output, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        output.seek(0)
        return StreamingResponse(
            iter([b'\xef\xbb\xbf' + output.getvalue().encode('utf-8')]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=users.csv"}
        )
    elif format == "excel":
        if not openpyxl:
            raise HTTPException(status_code=500, detail="openpyxl o'rnatilmagan")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"
        if data:
            ws.append(list(data[0].keys()))
            for d in data:
                ws.append(list(d.values()))
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=users.xlsx"}
        )


# ─── Leaderboard (Reyting) ─────────────────────────────────────────────────────

def _compute_overall(row) -> float | None:
    """Talabaning barcha bo'limlar bo'yicha overall band ni hisoblash."""
    from scoring import calculate_overall_band
    bands = []
    for section in ["listening", "reading", "writing", "speaking"]:
        if section in ("writing", "speaking"):
            key = f"{section}_band"
            val = row.get(key)
            if val is not None:
                bands.append(float(val))
        else:
            b = get_band_score(row.get("score"), row.get("total"), section)
            # Bu yerda row section-specific emas, shuning uchun alohida yo'l bilan
    return calculate_overall_band(bands) if bands else None


async def _build_leaderboard(conn, where_clause: str, args: tuple, limit: int = 50) -> list:
    """Leaderboard ma'lumotlarini DB dan olib, band hisoblash."""
    from scoring import get_band_score, calculate_overall_band

    rows = await conn.fetch(
        f"""
        SELECT
            u.id, u.full_name, u.email,
            g.name AS group_name,
            c.name AS center_name,
            BOOL_OR(er.section = 'listening') AS has_listening,
            MAX(CASE WHEN er.section = 'listening' THEN er.score END) AS l_score,
            MAX(CASE WHEN er.section = 'listening' THEN er.total END) AS l_total,
            MAX(CASE WHEN er.section = 'reading' THEN er.score END) AS r_score,
            MAX(CASE WHEN er.section = 'reading' THEN er.total END) AS r_total,
            MAX(CASE WHEN er.section = 'writing' THEN er.writing_band END) AS w_band,
            MAX(CASE WHEN er.section = 'speaking' THEN er.speaking_band END) AS s_band
        FROM users u
        LEFT JOIN exam_results er ON er.email = u.email
        LEFT JOIN groups g ON g.id = u.group_id
        LEFT JOIN centers c ON c.id = u.center_id
        {where_clause}
        GROUP BY u.id, u.full_name, u.email, g.name, c.name
        HAVING COUNT(er.id) > 0
        LIMIT {limit}
        """,
        *args
    )

    result = []
    for row in rows:
        l_band = get_band_score(row["l_score"], row["l_total"], "listening") if row["l_score"] is not None else None
        r_band = get_band_score(row["r_score"], row["r_total"], "reading") if row["r_score"] is not None else None
        w_band = float(row["w_band"]) if row["w_band"] is not None else None
        s_band = float(row["s_band"]) if row["s_band"] is not None else None

        all_bands = [b for b in [l_band, r_band, w_band, s_band] if b is not None]
        overall = calculate_overall_band(all_bands) if all_bands else None

        result.append({
            "id": row["id"],
            "full_name": row["full_name"],
            "email": row["email"],
            "group_name": row["group_name"],
            "center_name": row["center_name"],
            "listening_band": l_band,
            "reading_band": r_band,
            "writing_band": w_band,
            "speaking_band": s_band,
            "overall_band": overall
        })

    # Overall band bo'yicha tartib
    result.sort(key=lambda x: x["overall_band"] or 0, reverse=True)
    for i, item in enumerate(result):
        item["rank"] = i + 1
    return result


@app.get("/api/student/leaderboard")
async def student_leaderboard(current_user: dict = Depends(get_current_user)):
    """O'quvchi o'z guruhidagi reyting jadvalini ko'radi."""
    db = await get_pool()
    async with db.acquire() as conn:
        group_id = current_user.get("group_id")
        if not group_id:
            return {"leaderboard": [], "my_rank": None}

        data = await _build_leaderboard(conn, "WHERE u.group_id = $1 AND u.role = 'student'", (group_id,))

    my_rank = next((item["rank"] for item in data if item["email"] == current_user["email"]), None)
    return {"leaderboard": data[:20], "my_rank": my_rank, "total": len(data)}


@app.get("/api/student/resubmissions")
async def student_resubmissions(current_user: dict = Depends(get_current_user)):
    db = await get_pool()
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT rr.id,rr.section,rr.reason,rr.due_at,rr.status,rr.created_at,
                   COALESCE(t.title,er.test_slug,'IELTS Mock SS') AS test_title
            FROM resubmission_requests rr
            JOIN exam_results er ON er.id=rr.result_id
            LEFT JOIN tests t ON t.id=er.test_id
            WHERE rr.student_id=$1 ORDER BY rr.created_at DESC
            """,
            current_user["id"]
        )
    return [dict(r) | {
        "due_at": r["due_at"].isoformat() if r["due_at"] else None,
        "created_at": r["created_at"].isoformat(),
    } for r in rows]


@app.get("/api/admin/grade-speaking")
async def admin_grade_speaking_form(_: None = Depends(require_admin)):
    """Admin uchun baholanmagan speaking ro'yxati."""
    db = await get_pool()
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT er.id, er.full_name, er.email, er.speaking_telegram_file_id, er.submitted_at,
                   u.id AS student_id, g.name AS group_name
            FROM exam_results er
            LEFT JOIN users u ON u.email = er.email
            LEFT JOIN groups g ON g.id = u.group_id
            WHERE er.section = 'speaking' AND er.speaking_band IS NULL
            ORDER BY er.submitted_at ASC
            """
        )
    return [
        {
            "id": r["id"],
            "student_id": r["student_id"],
            "full_name": r["full_name"],
            "email": r["email"],
            "group_name": r["group_name"],
            "telegram_file_id": r["speaking_telegram_file_id"],
            "submitted_at": r["submitted_at"].isoformat()
        }
        for r in rows
    ]


class GradeSpeakingAdmin(BaseModel):
    result_id: int
    band: float
    feedback: Optional[str] = None
    send_notification: bool = True


@app.post("/api/admin/grade-speaking")
async def admin_grade_speaking(data: GradeSpeakingAdmin, _: None = Depends(require_admin)):
    """Admin speaking ni baholaydi."""
    db = await get_pool()
    async with db.acquire() as conn:
        row = await conn.fetchrow(
            """
            UPDATE exam_results
            SET speaking_band = $1, speaking_feedback = $2,
                grader_name = 'Admin', speaking_graded_at = NOW()
            WHERE id = $3 AND section = 'speaking'
            RETURNING full_name, email
            """,
            data.band, data.feedback, data.result_id
        )
        if not row:
            raise HTTPException(status_code=404, detail="Natija topilmadi")

    if data.send_notification:
        try:
            async with db.acquire() as conn2:
                user_row = await conn2.fetchrow(
                    "SELECT telegram_chat_id FROM users WHERE email = $1", row["email"].lower()
                )
            if user_row and user_row["telegram_chat_id"]:
                from telegram import notify_user_speaking_graded
                await notify_user_speaking_graded(
                    user_row["telegram_chat_id"], row["full_name"], data.band, data.feedback
                )
        except Exception:
            pass

    return {"message": "Speaking baholandi"}


# ─── Static Files ──────────────────────────────────────────────────────────────

app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/")
async def root():
    return FileResponse("static/index.html")

@app.head("/")
async def root_healthcheck():
    return Response(status_code=200)

@app.get("/admin")
async def admin_page():
    return FileResponse("static/admin.html")

@app.get("/admin/organizations/{center_id}")
async def admin_organization_detail_page(center_id: int):
    return FileResponse("static/organization-admin.html")

@app.get("/profile")
async def profile_page():
    return FileResponse("static/profile.html")

@app.get("/head-teacher")
async def head_teacher_page():
    return FileResponse("static/head-teacher.html")

@app.get("/teacher")
async def teacher_page():
    return FileResponse("static/teacher.html")

@app.get("/school-staff")
async def school_staff_page():
    return FileResponse("static/school-staff.html")

@app.get("/tests")
async def tests_page():
    return FileResponse("static/tests.html")

@app.get("/organizations")
async def organizations_page():
    return FileResponse("static/organizations.html")

@app.get("/org/{slug}")
async def organization_public_page(slug: str):
    return FileResponse("static/organizations.html")

@app.get("/test-builder")
async def test_builder_page():
    return FileResponse("static/test-builder.html")

@app.get("/mock-mode")
async def mock_mode_page():
    return FileResponse("static/mock-mode.html")

@app.get("/mock-result")
async def mock_result_page():
    return FileResponse("static/mock-result.html")

@app.get("/tests/{test_id}/run")
async def test_runner_page(test_id: int):
    return FileResponse("static/test-runner.html")

@app.get("/tests/{test_id}/mode")
async def test_mode_page(test_id: int):
    return FileResponse("static/test-mode.html")

@app.get("/listening-demo")
async def listening_demo_page():
    return FileResponse("static/Listening-demo.html")

@app.get("/reading-demo")
async def reading_demo_page():
    return FileResponse("static/Reading-demo.html")

@app.get("/writing-demo")
async def writing_demo_page():
    return FileResponse("static/writing-demo.html")

@app.get("/speaking-demo")
async def speaking_demo_page():
    return FileResponse("static/speaking-demo.html")

@app.get("/register")
async def register_page():
    # teacher_invite va group_invite linklar shu URL ga keladi
    return FileResponse("static/profile.html")

@app.get("/{path:path}")
async def catch_all(path: str):
    return FileResponse("static/index.html")
