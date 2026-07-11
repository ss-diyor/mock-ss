import json
import csv
import io
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
from email_validator import EmailNotValidError, validate_email
import openpyxl

from auth import USERNAME_RE, get_current_head_teacher, hash_password
from db import get_pool


router = APIRouter(prefix="/api/school", tags=["school"])

ALLOWED_PERMISSIONS = {
    "manage_staff", "manage_classes", "manage_students", "view_results",
    "manage_exams", "view_reports", "manage_branding",
}
DEFAULT_POSITIONS = (
    ("Direktor o'rinbosari", ["manage_staff", "manage_classes", "manage_students", "view_results", "view_reports"]),
    ("Sinf rahbari", ["manage_students", "view_results"]),
    ("Fan o'qituvchisi", ["view_results", "manage_exams"]),
    ("Administrator", ["manage_classes", "manage_students"]),
)


class PositionCreateIn(BaseModel):
    name: str
    permissions: list[str] = Field(default_factory=list)


class StaffCreateIn(BaseModel):
    login: str
    position_id: int
    employee_code: Optional[str] = None


class ClassCreateIn(BaseModel):
    name: str
    academic_year: str
    grade_level: Optional[int] = None
    homeroom_staff_id: Optional[int] = None


class ClassStudentIn(BaseModel):
    login: str


class SubjectCreateIn(BaseModel):
    name: str
    code: Optional[str] = None


class TeacherAssignmentIn(BaseModel):
    class_id: int
    subject_id: int
    staff_id: int


IMPORT_HEADERS = ("full_name", "email", "username", "password")
MAX_IMPORT_ROWS = 1000
MAX_IMPORT_BYTES = 5 * 1024 * 1024


def _read_student_rows(filename: str, content: bytes) -> list[dict]:
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if suffix == "csv":
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            raise HTTPException(status_code=400, detail="CSV UTF-8 formatida bo'lishi kerak")
        reader = csv.DictReader(io.StringIO(text))
        headers = [str(h or "").strip().lower() for h in (reader.fieldnames or [])]
        if headers != list(IMPORT_HEADERS):
            raise HTTPException(status_code=400, detail=f"CSV ustunlari aynan: {', '.join(IMPORT_HEADERS)}")
        return [dict(row) for row in reader]
    if suffix == "xlsx":
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = workbook["Oquvchilar"] if "Oquvchilar" in workbook.sheetnames else workbook.active
            values = sheet.iter_rows(values_only=True)
            headers = [str(value or "").strip().lower() for value in next(values, [])]
            if headers != list(IMPORT_HEADERS):
                raise HTTPException(status_code=400, detail=f"Excel ustunlari aynan: {', '.join(IMPORT_HEADERS)}")
            return [dict(zip(IMPORT_HEADERS, row)) for row in values]
        except HTTPException:
            raise
        except Exception:
            raise HTTPException(status_code=400, detail="Excel faylini o'qib bo'lmadi")
    raise HTTPException(status_code=400, detail="Faqat .xlsx yoki .csv fayl qabul qilinadi")


async def _school_or_404(conn, current_user: dict, for_update: bool = False):
    suffix = " FOR UPDATE" if for_update else ""
    row = await conn.fetchrow(
        f"SELECT id, name, organization_type, is_active FROM centers WHERE id=$1{suffix}",
        current_user["center_id"]
    )
    if not row or row["organization_type"] != "school":
        raise HTTPException(status_code=404, detail="Maktab topilmadi")
    if not row["is_active"]:
        raise HTTPException(status_code=403, detail="Maktab faol emas")
    return row


async def _ensure_default_positions(conn, center_id: int):
    for name, permissions in DEFAULT_POSITIONS:
        await conn.execute(
            """
            INSERT INTO school_positions(center_id, name, permissions, is_system)
            VALUES ($1, $2, $3::jsonb, TRUE)
            ON CONFLICT(center_id, name) DO NOTHING
            """,
            center_id, name, json.dumps(permissions)
        )


@router.get("/positions")
async def list_positions(current_user: dict = Depends(get_current_head_teacher)):
    db = await get_pool()
    async with db.acquire() as conn:
        school = await _school_or_404(conn, current_user)
        await _ensure_default_positions(conn, school["id"])
        rows = await conn.fetch(
            """
            SELECT p.id, p.name, p.permissions, p.is_system, p.created_at,
                   COUNT(s.id) FILTER (WHERE s.is_active=TRUE) AS staff_count
            FROM school_positions p
            LEFT JOIN school_staff s ON s.position_id=p.id
            WHERE p.center_id=$1
            GROUP BY p.id ORDER BY p.is_system DESC, p.name
            """,
            school["id"]
        )
    return [
        dict(r) | {
            "permissions": json.loads(r["permissions"]) if isinstance(r["permissions"], str) else r["permissions"],
            "created_at": r["created_at"].isoformat(),
        }
        for r in rows
    ]


@router.post("/positions")
async def create_position(data: PositionCreateIn, current_user: dict = Depends(get_current_head_teacher)):
    name = data.name.strip()
    if not name or len(name) > 80:
        raise HTTPException(status_code=400, detail="Lavozim nomi 1-80 belgi bo'lishi kerak")
    permissions = list(dict.fromkeys(data.permissions))
    if any(permission not in ALLOWED_PERMISSIONS for permission in permissions):
        raise HTTPException(status_code=400, detail="Noma'lum ruxsat yuborildi")
    db = await get_pool()
    async with db.acquire() as conn:
        school = await _school_or_404(conn, current_user)
        exists = await conn.fetchval(
            "SELECT 1 FROM school_positions WHERE center_id=$1 AND LOWER(name)=LOWER($2)", school["id"], name
        )
        if exists:
            raise HTTPException(status_code=409, detail="Bu lavozim mavjud")
        row = await conn.fetchrow(
            """
            INSERT INTO school_positions(center_id, name, permissions)
            VALUES ($1, $2, $3::jsonb)
            RETURNING id, name, permissions, is_system, created_at
            """,
            school["id"], name, json.dumps(permissions)
        )
    return dict(row) | {
        "permissions": json.loads(row["permissions"]) if isinstance(row["permissions"], str) else row["permissions"],
        "created_at": row["created_at"].isoformat(), "staff_count": 0,
    }


@router.get("/staff")
async def list_staff(current_user: dict = Depends(get_current_head_teacher)):
    db = await get_pool()
    async with db.acquire() as conn:
        school = await _school_or_404(conn, current_user)
        rows = await conn.fetch(
            """
            SELECT s.id, s.user_id, s.employee_code, s.is_active, s.created_at,
                   u.full_name, u.username, u.email, p.id AS position_id, p.name AS position_name
            FROM school_staff s
            JOIN users u ON u.id=s.user_id
            LEFT JOIN school_positions p ON p.id=s.position_id
            WHERE s.center_id=$1 ORDER BY s.is_active DESC, u.full_name
            """,
            school["id"]
        )
    return [dict(r) | {"created_at": r["created_at"].isoformat()} for r in rows]


@router.post("/staff")
async def add_staff(data: StaffCreateIn, current_user: dict = Depends(get_current_head_teacher)):
    login = data.login.strip().lower().lstrip("@")
    if not login:
        raise HTTPException(status_code=400, detail="Username yoki email kiriting")
    db = await get_pool()
    async with db.acquire() as conn:
        school = await _school_or_404(conn, current_user)
        position = await conn.fetchrow(
            "SELECT id FROM school_positions WHERE id=$1 AND center_id=$2", data.position_id, school["id"]
        )
        if not position:
            raise HTTPException(status_code=404, detail="Lavozim topilmadi")
        user = await conn.fetchrow("SELECT id, center_id FROM users WHERE username=$1 OR email=$1", login)
        if not user:
            raise HTTPException(status_code=404, detail="Foydalanuvchi topilmadi; avval ro'yxatdan o'tishi kerak")
        if user["center_id"] and user["center_id"] != school["id"]:
            raise HTTPException(status_code=409, detail="Foydalanuvchi boshqa tashkilotga biriktirilgan")
        duplicate = await conn.fetchval(
            "SELECT 1 FROM school_staff WHERE center_id=$1 AND user_id=$2", school["id"], user["id"]
        )
        if duplicate:
            raise HTTPException(status_code=409, detail="Bu xodim allaqachon qo'shilgan")
        row = await conn.fetchrow(
            """
            INSERT INTO school_staff(center_id, user_id, position_id, employee_code)
            VALUES ($1, $2, $3, $4) RETURNING id
            """,
            school["id"], user["id"], data.position_id,
            data.employee_code.strip()[:40] if data.employee_code else None
        )
        await conn.execute(
            """
            UPDATE users SET center_id=$1,
                role=CASE WHEN role='head_teacher' THEN role ELSE 'school_staff' END
            WHERE id=$2
            """,
            school["id"], user["id"]
        )
    return {"message": "Xodim qo'shildi", "id": row["id"]}


@router.post("/staff/{staff_id}/toggle")
async def toggle_staff(staff_id: int, current_user: dict = Depends(get_current_head_teacher)):
    db = await get_pool()
    async with db.acquire() as conn:
        school = await _school_or_404(conn, current_user)
        row = await conn.fetchrow(
            """
            UPDATE school_staff SET is_active=NOT is_active
            WHERE id=$1 AND center_id=$2 RETURNING is_active
            """,
            staff_id, school["id"]
        )
        if not row:
            raise HTTPException(status_code=404, detail="Xodim topilmadi")
    return {"is_active": row["is_active"]}


@router.get("/classes")
async def list_classes(current_user: dict = Depends(get_current_head_teacher)):
    db = await get_pool()
    async with db.acquire() as conn:
        school = await _school_or_404(conn, current_user)
        rows = await conn.fetch(
            """
            SELECT c.id, c.name, c.academic_year, c.grade_level, c.is_active, c.created_at,
                   c.homeroom_staff_id, u.full_name AS homeroom_teacher,
                   COUNT(cs.id) FILTER (WHERE cs.left_at IS NULL) AS students_count
            FROM school_classes c
            LEFT JOIN school_staff s ON s.id=c.homeroom_staff_id
            LEFT JOIN users u ON u.id=s.user_id
            LEFT JOIN school_class_students cs ON cs.class_id=c.id
            WHERE c.center_id=$1
            GROUP BY c.id, u.full_name ORDER BY c.academic_year DESC, c.grade_level, c.name
            """,
            school["id"]
        )
    return [dict(r) | {"created_at": r["created_at"].isoformat()} for r in rows]


@router.post("/classes")
async def create_class(data: ClassCreateIn, current_user: dict = Depends(get_current_head_teacher)):
    name, academic_year = data.name.strip(), data.academic_year.strip()
    if not name or not academic_year:
        raise HTTPException(status_code=400, detail="Sinf nomi va o'quv yilini kiriting")
    if data.grade_level is not None and not 1 <= data.grade_level <= 12:
        raise HTTPException(status_code=400, detail="Sinf bosqichi 1-12 oralig'ida bo'lishi kerak")
    db = await get_pool()
    async with db.acquire() as conn:
        school = await _school_or_404(conn, current_user)
        if data.homeroom_staff_id:
            own_staff = await conn.fetchval(
                "SELECT 1 FROM school_staff WHERE id=$1 AND center_id=$2 AND is_active=TRUE",
                data.homeroom_staff_id, school["id"]
            )
            if not own_staff:
                raise HTTPException(status_code=404, detail="Sinf rahbari topilmadi")
        duplicate = await conn.fetchval(
            "SELECT 1 FROM school_classes WHERE center_id=$1 AND LOWER(name)=LOWER($2) AND academic_year=$3",
            school["id"], name, academic_year
        )
        if duplicate:
            raise HTTPException(status_code=409, detail="Bu sinf shu o'quv yilida mavjud")
        row = await conn.fetchrow(
            """
            INSERT INTO school_classes(center_id, name, academic_year, grade_level, homeroom_staff_id)
            VALUES ($1, $2, $3, $4, $5) RETURNING id, name, academic_year
            """,
            school["id"], name, academic_year, data.grade_level, data.homeroom_staff_id
        )
    return dict(row)


@router.get("/classes/{class_id}/students")
async def list_class_students(class_id: int, current_user: dict = Depends(get_current_head_teacher)):
    db = await get_pool()
    async with db.acquire() as conn:
        school = await _school_or_404(conn, current_user)
        own_class = await conn.fetchval("SELECT 1 FROM school_classes WHERE id=$1 AND center_id=$2", class_id, school["id"])
        if not own_class:
            raise HTTPException(status_code=404, detail="Sinf topilmadi")
        rows = await conn.fetch(
            """
            SELECT u.id, u.full_name, u.username, u.email, cs.joined_at
            FROM school_class_students cs JOIN users u ON u.id=cs.student_id
            WHERE cs.class_id=$1 AND cs.left_at IS NULL ORDER BY u.full_name
            """,
            class_id
        )
    return [dict(r) | {"joined_at": r["joined_at"].isoformat()} for r in rows]


@router.get("/students/import-template")
async def student_import_template(current_user: dict = Depends(get_current_head_teacher)):
    db = await get_pool()
    async with db.acquire() as conn:
        await _school_or_404(conn, current_user)
    return FileResponse(
        "static/templates/school-students-import.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="school-students-import.xlsx",
    )


@router.post("/classes/{class_id}/import-students")
async def import_class_students(
    class_id: int,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_head_teacher)
):
    content = await file.read(MAX_IMPORT_BYTES + 1)
    if len(content) > MAX_IMPORT_BYTES:
        raise HTTPException(status_code=413, detail="Fayl hajmi 5 MB dan oshmasligi kerak")
    raw_rows = _read_student_rows(file.filename or "", content)
    rows = [row for row in raw_rows if any(str(value or "").strip() for value in row.values())]
    if not rows:
        raise HTTPException(status_code=400, detail="Import faylida o'quvchilar yo'q")
    if len(rows) > MAX_IMPORT_ROWS:
        raise HTTPException(status_code=400, detail=f"Bir importda ko'pi bilan {MAX_IMPORT_ROWS} o'quvchi")

    db = await get_pool()
    report = []
    seen_emails, seen_usernames = set(), set()
    async with db.acquire() as conn:
        school = await _school_or_404(conn, current_user)
        await _own_class_or_404(conn, class_id, school["id"])
        for index, raw in enumerate(rows, start=2):
            full_name = str(raw.get("full_name") or "").strip()
            email = str(raw.get("email") or "").strip().lower()
            username = str(raw.get("username") or "").strip().lower().lstrip("@")
            password = str(raw.get("password") or "").strip()
            errors = []
            if len(full_name) < 3:
                errors.append("To'liq ism kamida 3 belgi")
            try:
                email = validate_email(email, check_deliverability=False).normalized.lower()
            except EmailNotValidError:
                errors.append("Email noto'g'ri")
            if not USERNAME_RE.fullmatch(username):
                errors.append("Username 3-20 belgi: kichik harf, raqam yoki _")
            if len(password) < 6:
                errors.append("Parol kamida 6 belgi")
            if email in seen_emails:
                errors.append("Email fayl ichida takrorlangan")
            if username in seen_usernames:
                errors.append("Username fayl ichida takrorlangan")
            seen_emails.add(email)
            seen_usernames.add(username)

            if not errors:
                existing_email = await conn.fetchval("SELECT 1 FROM users WHERE email=$1", email)
                existing_username = await conn.fetchval("SELECT 1 FROM users WHERE username=$1", username)
                if existing_email:
                    errors.append("Email tizimda mavjud")
                if existing_username:
                    errors.append("Username tizimda mavjud")
            if errors:
                report.append({"row": index, "status": "error", "email": email, "username": username, "message": "; ".join(errors)})
                continue

            try:
                async with conn.transaction():
                    user_id = await conn.fetchval(
                        """
                        INSERT INTO users(username, email, full_name, password_hash, email_verified, role, center_id)
                        VALUES ($1, $2, $3, $4, TRUE, 'student', $5) RETURNING id
                        """,
                        username, email, full_name, hash_password(password), school["id"]
                    )
                    await conn.execute(
                        "INSERT INTO school_class_students(class_id, student_id) VALUES ($1, $2)",
                        class_id, user_id
                    )
                report.append({"row": index, "status": "imported", "email": email, "username": username, "message": "Import qilindi"})
            except Exception:
                report.append({"row": index, "status": "error", "email": email, "username": username, "message": "Bazaga saqlashda konflikt"})

    imported = sum(1 for item in report if item["status"] == "imported")
    return {"total": len(report), "imported": imported, "failed": len(report) - imported, "rows": report}


@router.get("/subjects")
async def list_subjects(current_user: dict = Depends(get_current_head_teacher)):
    db = await get_pool()
    async with db.acquire() as conn:
        school = await _school_or_404(conn, current_user)
        rows = await conn.fetch(
            "SELECT id, name, code, is_active, created_at FROM school_subjects WHERE center_id=$1 ORDER BY is_active DESC, name",
            school["id"]
        )
    return [dict(row) | {"created_at": row["created_at"].isoformat()} for row in rows]


@router.post("/subjects")
async def create_subject(data: SubjectCreateIn, current_user: dict = Depends(get_current_head_teacher)):
    name = data.name.strip()
    code = data.code.strip().upper()[:20] if data.code else None
    if not name or len(name) > 80:
        raise HTTPException(status_code=400, detail="Fan nomi 1-80 belgi bo'lishi kerak")
    db = await get_pool()
    async with db.acquire() as conn:
        school = await _school_or_404(conn, current_user)
        duplicate = await conn.fetchval(
            "SELECT 1 FROM school_subjects WHERE center_id=$1 AND LOWER(name)=LOWER($2)", school["id"], name
        )
        if duplicate:
            raise HTTPException(status_code=409, detail="Bu fan mavjud")
        row = await conn.fetchrow(
            "INSERT INTO school_subjects(center_id, name, code) VALUES ($1, $2, $3) RETURNING id, name, code, is_active",
            school["id"], name, code
        )
    return dict(row)


@router.get("/teacher-assignments")
async def list_teacher_assignments(current_user: dict = Depends(get_current_head_teacher)):
    db = await get_pool()
    async with db.acquire() as conn:
        school = await _school_or_404(conn, current_user)
        rows = await conn.fetch(
            """
            SELECT a.id, a.class_id, c.name AS class_name, a.subject_id, sub.name AS subject_name,
                   a.staff_id, u.full_name AS teacher_name, u.email AS teacher_email, a.created_at
            FROM school_teacher_assignments a
            JOIN school_classes c ON c.id=a.class_id
            JOIN school_subjects sub ON sub.id=a.subject_id
            JOIN school_staff s ON s.id=a.staff_id
            JOIN users u ON u.id=s.user_id
            WHERE a.center_id=$1 ORDER BY c.name, sub.name, u.full_name
            """,
            school["id"]
        )
    return [dict(row) | {"created_at": row["created_at"].isoformat()} for row in rows]


@router.post("/teacher-assignments")
async def create_teacher_assignment(data: TeacherAssignmentIn, current_user: dict = Depends(get_current_head_teacher)):
    db = await get_pool()
    async with db.acquire() as conn:
        school = await _school_or_404(conn, current_user)
        own_class = await conn.fetchval("SELECT 1 FROM school_classes WHERE id=$1 AND center_id=$2 AND is_active=TRUE", data.class_id, school["id"])
        own_subject = await conn.fetchval("SELECT 1 FROM school_subjects WHERE id=$1 AND center_id=$2 AND is_active=TRUE", data.subject_id, school["id"])
        own_staff = await conn.fetchval("SELECT 1 FROM school_staff WHERE id=$1 AND center_id=$2 AND is_active=TRUE", data.staff_id, school["id"])
        if not own_class or not own_subject or not own_staff:
            raise HTTPException(status_code=404, detail="Sinf, fan yoki o'qituvchi topilmadi")
        duplicate = await conn.fetchval(
            "SELECT 1 FROM school_teacher_assignments WHERE class_id=$1 AND subject_id=$2 AND staff_id=$3",
            data.class_id, data.subject_id, data.staff_id
        )
        if duplicate:
            raise HTTPException(status_code=409, detail="Bu biriktirish mavjud")
        assignment_id = await conn.fetchval(
            """
            INSERT INTO school_teacher_assignments(center_id, class_id, subject_id, staff_id)
            VALUES ($1, $2, $3, $4) RETURNING id
            """,
            school["id"], data.class_id, data.subject_id, data.staff_id
        )
    return {"message": "O'qituvchi biriktirildi", "id": assignment_id}


@router.delete("/teacher-assignments/{assignment_id}")
async def delete_teacher_assignment(assignment_id: int, current_user: dict = Depends(get_current_head_teacher)):
    db = await get_pool()
    async with db.acquire() as conn:
        school = await _school_or_404(conn, current_user)
        row = await conn.fetchrow(
            "DELETE FROM school_teacher_assignments WHERE id=$1 AND center_id=$2 RETURNING id",
            assignment_id, school["id"]
        )
    if not row:
        raise HTTPException(status_code=404, detail="Biriktirish topilmadi")
    return {"message": "Biriktirish olib tashlandi"}


@router.post("/classes/{class_id}/students")
async def add_class_student(class_id: int, data: ClassStudentIn, current_user: dict = Depends(get_current_head_teacher)):
    login = data.login.strip().lower().lstrip("@")
    db = await get_pool()
    async with db.acquire() as conn:
        async with conn.transaction():
            school = await _school_or_404(conn, current_user, for_update=True)
            own_class = await conn.fetchval("SELECT 1 FROM school_classes WHERE id=$1 AND center_id=$2", class_id, school["id"])
            if not own_class:
                raise HTTPException(status_code=404, detail="Sinf topilmadi")
            user = await conn.fetchrow("SELECT id, role, center_id FROM users WHERE username=$1 OR email=$1", login)
            if not user:
                raise HTTPException(status_code=404, detail="O'quvchi topilmadi; avval ro'yxatdan o'tishi kerak")
            if user["role"] != "student":
                raise HTTPException(status_code=400, detail="Faqat student hisobini sinfga qo'shish mumkin")
            if user["center_id"] and user["center_id"] != school["id"]:
                raise HTTPException(status_code=409, detail="O'quvchi boshqa tashkilotga biriktirilgan")
            other_class = await conn.fetchval(
                """
                SELECT c.name FROM school_class_students cs
                JOIN school_classes c ON c.id=cs.class_id
                WHERE cs.student_id=$1 AND cs.left_at IS NULL AND c.center_id=$2 AND c.id<>$3
                """,
                user["id"], school["id"], class_id
            )
            if other_class:
                raise HTTPException(status_code=409, detail=f"O'quvchi hozir {other_class} sinfida")
            await conn.execute(
                """
                INSERT INTO school_class_students(class_id, student_id, left_at)
                VALUES ($1, $2, NULL)
                ON CONFLICT(class_id, student_id) DO UPDATE SET left_at=NULL, joined_at=NOW()
                """,
                class_id, user["id"]
            )
            await conn.execute("UPDATE users SET center_id=$1 WHERE id=$2", school["id"], user["id"])
    return {"message": "O'quvchi sinfga qo'shildi"}


@router.delete("/classes/{class_id}/students/{student_id}")
async def remove_class_student(class_id: int, student_id: int, current_user: dict = Depends(get_current_head_teacher)):
    db = await get_pool()
    async with db.acquire() as conn:
        school = await _school_or_404(conn, current_user)
        row = await conn.fetchrow(
            """
            UPDATE school_class_students cs SET left_at=NOW()
            FROM school_classes c
            WHERE cs.class_id=c.id AND c.id=$1 AND c.center_id=$2
              AND cs.student_id=$3 AND cs.left_at IS NULL
            RETURNING cs.id
            """,
            class_id, school["id"], student_id
        )
        if not row:
            raise HTTPException(status_code=404, detail="O'quvchi sinfda topilmadi")
    return {"message": "O'quvchi sinfdan chiqarildi"}
