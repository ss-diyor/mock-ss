import csv
import hashlib
import io
import json
import mimetypes
import random
import zipfile
from collections import defaultdict
from datetime import datetime
from typing import Any, Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import Response
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel, Field

from db import get_pool
from notification_center import create_notification, notify_admin
from test_builder_routes import (
    BuilderQuestionIn,
    MEDIA_LIMITS,
    MEDIA_MIMES,
    QUESTION_TYPES,
    SECTIONS,
    _owned_test,
    _validate_question,
    get_current_head_teacher,
)


router = APIRouter(prefix="/api/question-bank", tags=["question-bank"])
DIFFICULTIES = {"easy", "medium", "hard"}
STATUSES = {"draft", "review", "approved", "retired"}
IMPORT_COLUMNS = [
    "set_title", "section", "instructions", "passage", "topic", "difficulty",
    "band_min", "band_max", "source_name", "source_url", "license_note",
    "question_type", "prompt", "options", "correct_answer", "points",
    "explanation", "media_filename", "skill", "tags",
]


class BankSetIn(BaseModel):
    section: str
    title: str
    instructions: str = ""
    passage: str = ""
    topic: str = ""
    difficulty: str = "medium"
    band_min: Optional[float] = None
    band_max: Optional[float] = None
    source_name: str = ""
    source_url: str = ""
    license_note: str = ""
    planned_publish_at: Optional[datetime] = None
    review_due_at: Optional[datetime] = None


class BankQuestionIn(BuilderQuestionIn):
    skill: str = ""
    tags: list[str] = Field(default_factory=list)
    difficulty: str = "medium"
    band_min: Optional[float] = None
    band_max: Optional[float] = None


class ReviewIn(BaseModel):
    note: str = ""


class AutoBuildIn(BaseModel):
    sections: dict[str, int] = Field(default_factory=lambda: {
        "listening": 40, "reading": 40, "writing": 2, "speaking": 3,
    })
    difficulty: Optional[str] = None
    topic: str = ""
    avoid_recent_days: int = 30


def _clean_text(value: Any, limit: int = 10_000) -> str:
    return str(value or "").strip()[:limit]


def _normal(value: Any) -> str:
    return str(value if value is not None else "").strip().casefold()


def _decoded(value: Any, fallback: Any):
    if value is None:
        return fallback
    if isinstance(value, str):
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            return fallback
    return value


def _question_hash(section: str, data: BankQuestionIn, context: str = "") -> str:
    payload = {
        "section": section,
        "type": data.question_type,
        "prompt": " ".join(data.prompt.casefold().split()),
        "options": [_normal(item) for item in data.options],
        "answer": data.correct_answer,
        "context": " ".join(context.casefold().split())[:20_000],
    }
    return hashlib.sha256(json.dumps(payload, ensure_ascii=False, sort_keys=True).encode()).hexdigest()


def _validate_set(data: BankSetIn):
    if data.section not in SECTIONS:
        raise HTTPException(400, "Bo'lim noto'g'ri")
    if not data.title.strip() or len(data.title) > 180:
        raise HTTPException(400, "To'plam nomi noto'g'ri")
    if data.difficulty not in DIFFICULTIES:
        raise HTTPException(400, "Qiyinlik darajasi noto'g'ri")
    for value in (data.band_min, data.band_max):
        if value is not None and not 0 <= value <= 9:
            raise HTTPException(400, "Band 0–9 oralig'ida bo'lishi kerak")
    if data.band_min is not None and data.band_max is not None and data.band_min > data.band_max:
        raise HTTPException(400, "Band oralig'i noto'g'ri")


def _validate_bank_question(section: str, data: BankQuestionIn):
    _validate_question(data)
    if data.difficulty not in DIFFICULTIES:
        raise HTTPException(400, "Qiyinlik darajasi noto'g'ri")
    if data.question_type == "writing_task" and section != "writing":
        raise HTTPException(400, "Writing task faqat Writing bo'limida bo'lishi mumkin")
    if data.question_type == "speaking_prompt" and section != "speaking":
        raise HTTPException(400, "Speaking prompt faqat Speaking bo'limida bo'lishi mumkin")
    if section in {"listening", "reading"} and data.question_type in {"writing_task", "speaking_prompt"}:
        raise HTTPException(400, "Savol turi bo'limga mos emas")
    for value in (data.band_min, data.band_max):
        if value is not None and not 0 <= value <= 9:
            raise HTTPException(400, "Band 0–9 oralig'ida bo'lishi kerak")


async def _set_row(conn, set_id: int, user: dict, editable: bool = False):
    if user.get("is_admin"):
        row = await conn.fetchrow("SELECT * FROM question_bank_sets WHERE id=$1", set_id)
    elif editable:
        row = await conn.fetchrow(
            "SELECT * FROM question_bank_sets WHERE id=$1 AND center_id=$2", set_id, user["center_id"]
        )
    else:
        row = await conn.fetchrow(
            """SELECT * FROM question_bank_sets WHERE id=$1
               AND (center_id=$2 OR (center_id IS NULL AND status='approved'))""",
            set_id, user["center_id"],
        )
    if not row:
        raise HTTPException(404, "Savollar to'plami topilmadi")
    return row


@router.get("/sets")
async def list_bank_sets(
    q: str = "",
    section: str = "",
    status: str = "",
    difficulty: str = "",
    skill: str = "",
    scope: str = "all",
    page: int = Query(1, ge=1),
    page_size: int = Query(24, ge=1, le=100),
    current_user: dict = Depends(get_current_head_teacher),
):
    if section and section not in SECTIONS:
        raise HTTPException(400, "Bo'lim noto'g'ri")
    if status and status not in STATUSES:
        raise HTTPException(400, "Holat noto'g'ri")
    if difficulty and difficulty not in DIFFICULTIES:
        raise HTTPException(400, "Qiyinlik noto'g'ri")
    center_id = None if current_user.get("is_admin") else current_user["center_id"]
    offset = (page - 1) * page_size
    db = await get_pool()
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT s.id,s.center_id,s.section,s.title,s.topic,s.difficulty,s.band_min,s.band_max,
                   s.source_name,s.status,s.planned_publish_at,s.review_due_at,s.version,s.usage_count,s.last_used_at,s.created_at,s.updated_at,
                   COUNT(DISTINCT qn.id) AS question_count,
                   COALESCE(ARRAY_AGG(DISTINCT qn.skill) FILTER(WHERE qn.skill IS NOT NULL AND qn.skill<>''),ARRAY[]::TEXT[]) AS skills
            FROM question_bank_sets s
            LEFT JOIN question_bank_questions qn ON qn.set_id=s.id
            WHERE ($1::integer IS NULL OR s.center_id=$1 OR (s.center_id IS NULL AND s.status='approved'))
              AND ($2='' OR s.section=$2) AND ($3='' OR s.status=$3)
              AND ($4='' OR s.difficulty=$4)
              AND ($5='' OR s.title ILIKE '%'||$5||'%' OR COALESCE(s.topic,'') ILIKE '%'||$5||'%'
                   OR EXISTS(SELECT 1 FROM question_bank_questions sq WHERE sq.set_id=s.id AND sq.prompt ILIKE '%'||$5||'%'))
              AND ($6='' OR EXISTS(SELECT 1 FROM question_bank_questions sq WHERE sq.set_id=s.id AND sq.skill=$6))
              AND ($7='all' OR ($7='public' AND s.center_id IS NULL) OR ($7='organization' AND s.center_id IS NOT NULL))
            GROUP BY s.id
            ORDER BY CASE s.status WHEN 'review' THEN 0 WHEN 'draft' THEN 1 WHEN 'approved' THEN 2 ELSE 3 END,
                     s.updated_at DESC
            LIMIT $8 OFFSET $9
            """, center_id, section, status, difficulty, q.strip(), skill.strip(), scope, page_size, offset,
        )
        total = await conn.fetchval(
            """
            SELECT COUNT(*) FROM question_bank_sets s
            WHERE ($1::integer IS NULL OR s.center_id=$1 OR (s.center_id IS NULL AND s.status='approved'))
              AND ($2='' OR s.section=$2) AND ($3='' OR s.status=$3) AND ($4='' OR s.difficulty=$4)
              AND ($5='' OR s.title ILIKE '%'||$5||'%' OR COALESCE(s.topic,'') ILIKE '%'||$5||'%'
                   OR EXISTS(SELECT 1 FROM question_bank_questions sq WHERE sq.set_id=s.id AND sq.prompt ILIKE '%'||$5||'%'))
              AND ($6='' OR EXISTS(SELECT 1 FROM question_bank_questions sq WHERE sq.set_id=s.id AND sq.skill=$6))
              AND ($7='all' OR ($7='public' AND s.center_id IS NULL) OR ($7='organization' AND s.center_id IS NOT NULL))
            """, center_id, section, status, difficulty, q.strip(), skill.strip(), scope,
        )
    return {
        "items": [dict(row) for row in rows], "total": total, "page": page,
        "page_size": page_size, "pages": max(1, (total + page_size - 1) // page_size),
    }


@router.post("/sets")
async def create_bank_set(data: BankSetIn, current_user: dict = Depends(get_current_head_teacher)):
    _validate_set(data)
    db = await get_pool()
    async with db.acquire() as conn:
        set_id = await conn.fetchval(
            """
            INSERT INTO question_bank_sets(center_id,section,title,instructions,passage,topic,difficulty,
              band_min,band_max,source_name,source_url,license_note,planned_publish_at,review_due_at,created_by)
            VALUES($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15) RETURNING id
            """,
            None if current_user.get("is_admin") else current_user["center_id"], data.section,
            data.title.strip(), data.instructions[:10_000], data.passage[:150_000], data.topic.strip()[:160],
            data.difficulty, data.band_min, data.band_max, data.source_name.strip()[:300],
            data.source_url.strip()[:1000], data.license_note[:2000], data.planned_publish_at,
            data.review_due_at, current_user.get("id"),
        )
    return {"id": set_id, "message": "Savollar to'plami yaratildi"}


@router.get("/sets/{set_id}")
async def bank_set_detail(set_id: int, current_user: dict = Depends(get_current_head_teacher)):
    db = await get_pool()
    async with db.acquire() as conn:
        set_row = await _set_row(conn, set_id, current_user)
        questions = await conn.fetch(
            """SELECT id,question_type,prompt,options,correct_answer,points,explanation,media_id,skill,tags,
                      difficulty,band_min,band_max,status,version,usage_count,last_used_at,review_note,created_at,updated_at
               FROM question_bank_questions WHERE set_id=$1 ORDER BY id""", set_id,
        )
        media = await conn.fetch(
            "SELECT id,kind,original_filename,mime_type,file_size,created_at FROM question_bank_media WHERE set_id=$1 ORDER BY id",
            set_id,
        )
    payload = dict(set_row)
    payload["questions"] = []
    for row in questions:
        item = dict(row)
        item["options"] = _decoded(item.get("options"), [])
        item["correct_answer"] = _decoded(item.get("correct_answer"), None)
        item["points"] = float(item["points"])
        payload["questions"].append(item)
    payload["media"] = [dict(row) for row in media]
    return payload


@router.put("/sets/{set_id}")
async def update_bank_set(set_id: int, data: BankSetIn, current_user: dict = Depends(get_current_head_teacher)):
    _validate_set(data)
    db = await get_pool()
    async with db.acquire() as conn:
        await _set_row(conn, set_id, current_user, editable=True)
        await conn.execute(
            """
            UPDATE question_bank_sets SET section=$2,title=$3,instructions=$4,passage=$5,topic=$6,
              difficulty=$7,band_min=$8,band_max=$9,source_name=$10,source_url=$11,license_note=$12,
              planned_publish_at=$13,review_due_at=$14,status='draft',version=version+1,
              reviewed_by=NULL,review_note=NULL,updated_at=NOW()
            WHERE id=$1
            """, set_id, data.section, data.title.strip(), data.instructions[:10_000], data.passage[:150_000],
            data.topic.strip()[:160], data.difficulty, data.band_min, data.band_max,
            data.source_name.strip()[:300], data.source_url.strip()[:1000], data.license_note[:2000],
            data.planned_publish_at, data.review_due_at,
        )
        for question in await conn.fetch("SELECT * FROM question_bank_questions WHERE set_id=$1", set_id):
            question_data = BankQuestionIn(
                question_type=question["question_type"], prompt=question["prompt"],
                options=_decoded(question["options"], []), correct_answer=_decoded(question["correct_answer"], None),
                points=float(question["points"]), explanation=question["explanation"] or "",
                media_id=question["media_id"], skill=question["skill"] or "", tags=list(question["tags"] or []),
                difficulty=question["difficulty"], band_min=question["band_min"], band_max=question["band_max"],
            )
            new_hash = _question_hash(data.section, question_data, f"{data.title}\n{data.passage}")
            await conn.execute(
                """UPDATE question_bank_questions SET content_hash=$2,
                     status=CASE WHEN status='retired' THEN status ELSE 'draft' END,updated_at=NOW() WHERE id=$1""",
                question["id"], new_hash,
            )
    return {"message": "To'plam yangilandi va draft holatiga qaytdi"}


@router.delete("/sets/{set_id}")
async def retire_bank_set(set_id: int, current_user: dict = Depends(get_current_head_teacher)):
    db = await get_pool()
    async with db.acquire() as conn:
        await _set_row(conn, set_id, current_user, editable=True)
        await conn.execute(
            "UPDATE question_bank_sets SET status='retired',updated_at=NOW() WHERE id=$1", set_id
        )
        await conn.execute(
            "UPDATE question_bank_questions SET status='retired',updated_at=NOW() WHERE set_id=$1", set_id
        )
    return {"message": "To'plam arxivlandi"}


@router.post("/sets/{set_id}/status/{status}")
async def update_bank_status(
    set_id: int, status: str, data: ReviewIn, current_user: dict = Depends(get_current_head_teacher)
):
    if status not in STATUSES:
        raise HTTPException(400, "Holat noto'g'ri")
    db = await get_pool()
    async with db.acquire() as conn:
        row = await _set_row(conn, set_id, current_user, editable=True)
        if status == "approved" and not current_user.get("is_admin") and row["center_id"] is None:
            raise HTTPException(403, "Public savollarni faqat super-admin tasdiqlaydi")
        if status == "approved" and row["center_id"] is None and (
            not (row["source_name"] or "").strip() or not (row["license_note"] or "").strip()
        ):
            raise HTTPException(400, "Public to'plam uchun manba va foydalanish huquqi izohi majburiy")
        if status in {"review", "approved"}:
            count = await conn.fetchval(
                "SELECT COUNT(*) FROM question_bank_questions WHERE set_id=$1 AND status<>'retired'", set_id
            )
            if not count:
                raise HTTPException(400, "Bo'sh to'plamni tasdiqlab bo'lmaydi")
        reviewer = current_user.get("id") if status in {"approved", "retired"} else None
        await conn.execute(
            """UPDATE question_bank_sets SET status=$2,reviewed_by=$3,review_note=$4,updated_at=NOW()
               WHERE id=$1""", set_id, status, reviewer, data.note[:2000],
        )
        await conn.execute(
            """UPDATE question_bank_questions SET status=$2,reviewed_by=$3,review_note=$4,updated_at=NOW()
               WHERE set_id=$1 AND ($2='retired' OR status<>'retired')""", set_id, status, reviewer, data.note[:2000],
        )
        if status == "review" and not current_user.get("is_admin"):
            await notify_admin(
                conn, "Savollar to'plami tekshiruvga yuborildi",
                f"{row['title']} ({row['section']}) to'plamini tekshirish kerak.",
                kind="task", action_url="/test-builder", metadata={"question_bank_set_id": set_id},
            )
        if status == "approved" and row["created_by"] and row["created_by"] != current_user.get("id"):
            await create_notification(
                conn, recipient_user_id=row["created_by"], kind="success",
                title="Savollar to'plami tasdiqlandi", message=f"{row['title']} foydalanishga tayyor.",
                action_url="/test-builder", metadata={"question_bank_set_id": set_id},
            )
    return {"status": status, "message": "Holat yangilandi"}


@router.post("/sets/{set_id}/questions")
async def create_bank_question(
    set_id: int, data: BankQuestionIn, current_user: dict = Depends(get_current_head_teacher)
):
    db = await get_pool()
    async with db.acquire() as conn:
        set_row = await _set_row(conn, set_id, current_user, editable=True)
        _validate_bank_question(set_row["section"], data)
        if data.media_id and not await conn.fetchval(
            "SELECT 1 FROM question_bank_media WHERE id=$1 AND set_id=$2", data.media_id, set_id
        ):
            raise HTTPException(404, "Media topilmadi")
        content_hash = _question_hash(set_row["section"], data, f"{set_row['title']}\n{set_row['passage'] or ''}")
        if await conn.fetchval("SELECT 1 FROM question_bank_questions WHERE content_hash=$1", content_hash):
            raise HTTPException(409, "Bu savol bankda allaqachon mavjud")
        question_id = await conn.fetchval(
            """
            INSERT INTO question_bank_questions(set_id,question_type,prompt,options,correct_answer,points,
              explanation,media_id,skill,tags,difficulty,band_min,band_max,content_hash,created_by)
            VALUES($1,$2,$3,$4::jsonb,$5::jsonb,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15) RETURNING id
            """, set_id, data.question_type, data.prompt.strip(), json.dumps(data.options, ensure_ascii=False),
            json.dumps(data.correct_answer, ensure_ascii=False) if data.correct_answer is not None else None,
            data.points, data.explanation[:5000], data.media_id, data.skill.strip()[:160],
            [tag.strip()[:60] for tag in data.tags if tag.strip()][:20], data.difficulty,
            data.band_min, data.band_max, content_hash, current_user.get("id"),
        )
        await conn.execute("UPDATE question_bank_sets SET status='draft',updated_at=NOW() WHERE id=$1", set_id)
    return {"id": question_id, "message": "Savol bankka qo'shildi"}


@router.put("/questions/{question_id}")
async def update_bank_question(
    question_id: int, data: BankQuestionIn, current_user: dict = Depends(get_current_head_teacher)
):
    db = await get_pool()
    async with db.acquire() as conn:
        old = await conn.fetchrow(
            """SELECT q.*,s.section,s.center_id FROM question_bank_questions q
               JOIN question_bank_sets s ON s.id=q.set_id WHERE q.id=$1""", question_id
        )
        if not old or (not current_user.get("is_admin") and old["center_id"] != current_user["center_id"]):
            raise HTTPException(404, "Savol topilmadi")
        _validate_bank_question(old["section"], data)
        if data.media_id and not await conn.fetchval(
            "SELECT 1 FROM question_bank_media WHERE id=$1 AND set_id=$2", data.media_id, old["set_id"]
        ):
            raise HTTPException(404, "Media topilmadi")
        set_context = await conn.fetchrow("SELECT title,passage FROM question_bank_sets WHERE id=$1", old["set_id"])
        content_hash = _question_hash(old["section"], data, f"{set_context['title']}\n{set_context['passage'] or ''}")
        if await conn.fetchval(
            "SELECT 1 FROM question_bank_questions WHERE content_hash=$1 AND id<>$2", content_hash, question_id
        ):
            raise HTTPException(409, "Bu savol bankda allaqachon mavjud")
        snapshot = {key: (float(old[key]) if key == "points" else old[key]) for key in (
            "question_type", "prompt", "options", "correct_answer", "points", "explanation", "media_id",
            "skill", "tags", "difficulty", "band_min", "band_max", "status", "version",
        )}
        await conn.execute(
            """INSERT INTO question_bank_versions(bank_question_id,version,snapshot,changed_by)
               VALUES($1,$2,$3::jsonb,$4) ON CONFLICT(bank_question_id,version) DO NOTHING""",
            question_id, old["version"], json.dumps(snapshot, ensure_ascii=False, default=str), current_user.get("id"),
        )
        await conn.execute(
            """
            UPDATE question_bank_questions SET question_type=$2,prompt=$3,options=$4::jsonb,
              correct_answer=$5::jsonb,points=$6,explanation=$7,media_id=$8,skill=$9,tags=$10,
              difficulty=$11,band_min=$12,band_max=$13,content_hash=$14,status='draft',
              version=version+1,reviewed_by=NULL,review_note=NULL,updated_at=NOW() WHERE id=$1
            """, question_id, data.question_type, data.prompt.strip(), json.dumps(data.options, ensure_ascii=False),
            json.dumps(data.correct_answer, ensure_ascii=False) if data.correct_answer is not None else None,
            data.points, data.explanation[:5000], data.media_id, data.skill.strip()[:160],
            [tag.strip()[:60] for tag in data.tags if tag.strip()][:20], data.difficulty,
            data.band_min, data.band_max, content_hash,
        )
        await conn.execute(
            "UPDATE question_bank_sets SET status='draft',updated_at=NOW() WHERE id=$1", old["set_id"]
        )
    return {"message": "Savol yangilandi; avvalgi versiya tarixda saqlandi"}


@router.delete("/questions/{question_id}")
async def retire_bank_question(question_id: int, current_user: dict = Depends(get_current_head_teacher)):
    db = await get_pool()
    async with db.acquire() as conn:
        row = await conn.fetchrow(
            """SELECT q.id,q.set_id,s.center_id FROM question_bank_questions q
               JOIN question_bank_sets s ON s.id=q.set_id WHERE q.id=$1""", question_id
        )
        if not row or (not current_user.get("is_admin") and row["center_id"] != current_user["center_id"]):
            raise HTTPException(404, "Savol topilmadi")
        await conn.execute(
            "UPDATE question_bank_questions SET status='retired',updated_at=NOW() WHERE id=$1", question_id
        )
    return {"message": "Savol arxivlandi"}


@router.get("/questions/{question_id}/versions")
async def bank_question_versions(
    question_id: int, current_user: dict = Depends(get_current_head_teacher)
):
    db = await get_pool()
    async with db.acquire() as conn:
        question = await conn.fetchrow(
            """SELECT q.id,q.set_id,q.version FROM question_bank_questions q WHERE q.id=$1""", question_id
        )
        if not question:
            raise HTTPException(404, "Savol topilmadi")
        await _set_row(conn, question["set_id"], current_user)
        rows = await conn.fetch(
            """SELECT version,snapshot,changed_by,created_at FROM question_bank_versions
               WHERE bank_question_id=$1 ORDER BY version DESC""", question_id
        )
    return {"current_version": question["version"], "versions": [dict(row) for row in rows]}


def _valid_media(kind: str, mime: str, content: bytes) -> bool:
    if kind == "image":
        return content.startswith((b"\xff\xd8", b"\x89PNG", b"GIF8")) or (
            content.startswith(b"RIFF") and content[8:12] == b"WEBP"
        )
    return content.startswith((b"OggS", b"ID3", b"\x1aE\xdf\xa3")) or (
        content.startswith(b"RIFF") and content[8:12] == b"WAVE"
    ) or b"ftyp" in content[:32] or content[:2] in {b"\xff\xfb", b"\xff\xf3", b"\xff\xf2"}


@router.post("/sets/{set_id}/media")
async def upload_bank_media(
    set_id: int, kind: str = Form(...), file: UploadFile = File(...),
    current_user: dict = Depends(get_current_head_teacher),
):
    if kind not in MEDIA_LIMITS:
        raise HTTPException(400, "Media turi noto'g'ri")
    mime = (file.content_type or "").lower()
    if mime not in MEDIA_MIMES[kind]:
        raise HTTPException(400, "Media formati qo'llab-quvvatlanmaydi")
    content = await file.read(MEDIA_LIMITS[kind] + 1)
    if not content or len(content) > MEDIA_LIMITS[kind] or not _valid_media(kind, mime, content):
        raise HTTPException(400, "Media fayli noto'g'ri yoki hajmi limitdan oshgan")
    db = await get_pool()
    async with db.acquire() as conn:
        await _set_row(conn, set_id, current_user, editable=True)
        used = await conn.fetchval(
            "SELECT COALESCE(SUM(file_size),0) FROM question_bank_media WHERE set_id=$1", set_id
        )
        if used + len(content) > 60 * 1024 * 1024:
            raise HTTPException(413, "To'plam medialari jami 60 MB dan oshmasligi kerak")
        media_id = await conn.fetchval(
            """INSERT INTO question_bank_media(set_id,kind,original_filename,mime_type,file_data,file_size)
               VALUES($1,$2,$3,$4,$5,$6) RETURNING id""",
            set_id, kind, (file.filename or kind)[:180], mime, content, len(content),
        )
    return {"id": media_id, "message": "Media savollar bankiga yuklandi"}


@router.delete("/media/{media_id}")
async def delete_bank_media(media_id: int, current_user: dict = Depends(get_current_head_teacher)):
    db = await get_pool()
    async with db.acquire() as conn:
        row = await conn.fetchrow(
            """SELECT m.id,m.set_id,s.center_id FROM question_bank_media m
               JOIN question_bank_sets s ON s.id=m.set_id WHERE m.id=$1""", media_id
        )
        if not row or (not current_user.get("is_admin") and row["center_id"] != current_user["center_id"]):
            raise HTTPException(404, "Media topilmadi")
        await conn.execute("DELETE FROM question_bank_media WHERE id=$1", media_id)
    return {"message": "Media o'chirildi"}


def _template_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Savollar"
    ws.append(IMPORT_COLUMNS)
    ws.append([
        "Reading Set 1", "reading", "Matnni o'qing", "Passage matni", "Education", "medium",
        5.0, 7.0, "Muallif", "https://example.com", "Foydalanishga ruxsat berilgan",
        "single_choice", "Asosiy fikr qaysi?", "A|B|C|D", "B", 1,
        "B javobi matnga mos", "passage-1.png", "main idea", "reading|main-idea",
    ])
    header_fill = PatternFill("solid", fgColor="1A56E8")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    ws.freeze_panes = "A2"
    widths = {"A": 24, "B": 12, "C": 26, "D": 44, "M": 44, "N": 28, "O": 22, "Q": 34}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    guide = wb.create_sheet("Qo'llanma")
    guide.append(["Maydon", "Izoh"])
    guide_rows = [
        ("section", "listening, reading, writing yoki speaking"),
        ("difficulty", "easy, medium yoki hard"),
        ("question_type", "single_choice, multiple_choice, short_answer, true_false, yes_no_not_given, matching_headings, matching_information, sentence_completion, summary_completion, note_completion, table_completion, flow_chart_completion, diagram_label, map_label, writing_task, speaking_prompt"),
        ("options", "Variantlarni | belgisi bilan ajrating"),
        ("correct_answer", "Multiple choice uchun javoblarni | bilan ajrating"),
        ("tags", "Teglarni | belgisi bilan ajrating"),
        ("media_filename", "Ixtiyoriy ZIP ichidagi rasm/audio fayl nomi"),
        ("set_title", "Bir xil nom va section qatordagi savollar bitta to'plamga birlashadi"),
    ]
    for row in guide_rows:
        guide.append(row)
    guide.column_dimensions["A"].width = 24
    guide.column_dimensions["B"].width = 85
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


@router.get("/import/template")
async def bank_import_template(_: dict = Depends(get_current_head_teacher)):
    return Response(
        _template_workbook(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="ielts-question-bank-template.xlsx"'},
    )


def _rows_from_upload(filename: str, content: bytes) -> list[dict[str, Any]]:
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if suffix == "csv":
        text = content.decode("utf-8-sig")
        return [dict(row) for row in csv.DictReader(io.StringIO(text))]
    if suffix == "xlsx":
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb["Savollar"] if "Savollar" in wb.sheetnames else wb.active
        values = ws.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(values)]
        return [dict(zip(headers, row)) for row in values if any(value not in (None, "") for value in row)]
    raise HTTPException(400, "Faqat XLSX yoki CSV fayl yuklang")


def _split_values(value: Any) -> list[str]:
    if value in (None, ""):
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    return [part.strip() for part in str(value).replace("\n", "|").split("|") if part.strip()]


def _row_models(row: dict[str, Any]) -> tuple[BankSetIn, BankQuestionIn]:
    section = _clean_text(row.get("section"), 30).lower()
    qtype = _clean_text(row.get("question_type"), 40).lower()
    options = _split_values(row.get("options"))
    answer_parts = _split_values(row.get("correct_answer"))
    answer: Any = answer_parts
    if qtype != "multiple_choice":
        answer = answer_parts[0] if answer_parts else None
    if qtype in {"writing_task", "speaking_prompt"}:
        answer = None
    set_data = BankSetIn(
        section=section, title=_clean_text(row.get("set_title"), 180),
        instructions=_clean_text(row.get("instructions")), passage=_clean_text(row.get("passage"), 150_000),
        topic=_clean_text(row.get("topic"), 160), difficulty=_clean_text(row.get("difficulty"), 20).lower() or "medium",
        band_min=float(row["band_min"]) if row.get("band_min") not in (None, "") else None,
        band_max=float(row["band_max"]) if row.get("band_max") not in (None, "") else None,
        source_name=_clean_text(row.get("source_name"), 300), source_url=_clean_text(row.get("source_url"), 1000),
        license_note=_clean_text(row.get("license_note"), 2000),
        planned_publish_at=None, review_due_at=None,
    )
    q_data = BankQuestionIn(
        question_type=qtype, prompt=_clean_text(row.get("prompt")), options=options,
        correct_answer=answer, points=float(row.get("points") or 1),
        explanation=_clean_text(row.get("explanation"), 5000), skill=_clean_text(row.get("skill"), 160),
        tags=_split_values(row.get("tags")), difficulty=set_data.difficulty,
        band_min=set_data.band_min, band_max=set_data.band_max,
    )
    return set_data, q_data


@router.post("/import")
async def import_question_bank(
    file: UploadFile = File(...), media_archive: Optional[UploadFile] = File(None),
    current_user: dict = Depends(get_current_head_teacher)
):
    content = await file.read(5 * 1024 * 1024 + 1)
    if not content or len(content) > 5 * 1024 * 1024:
        raise HTTPException(413, "Import fayli 5 MB dan oshmasligi kerak")
    filename = file.filename or "import.xlsx"
    try:
        rows = _rows_from_upload(filename, content)
    except (UnicodeDecodeError, ValueError, KeyError, StopIteration) as exc:
        raise HTTPException(400, f"Import faylini o'qib bo'lmadi: {exc}")
    if len(rows) > 5000:
        raise HTTPException(413, "Bir importda ko'pi bilan 5000 qator")
    archive_files: dict[str, tuple[str, str, bytes]] = {}
    if media_archive:
        archive_content = await media_archive.read(25 * 1024 * 1024 + 1)
        if len(archive_content) > 25 * 1024 * 1024:
            raise HTTPException(413, "Media ZIP 25 MB dan oshmasligi kerak")
        try:
            with zipfile.ZipFile(io.BytesIO(archive_content)) as archive:
                total_uncompressed = sum(info.file_size for info in archive.infolist() if not info.is_dir())
                if total_uncompressed > 60 * 1024 * 1024:
                    raise HTTPException(413, "ZIP ochilganda 60 MB dan oshmasligi kerak")
                for info in archive.infolist():
                    if info.is_dir():
                        continue
                    safe_name = info.filename.replace("\\", "/").rsplit("/", 1)[-1]
                    if not safe_name or safe_name in archive_files:
                        continue
                    mime = (mimetypes.guess_type(safe_name)[0] or "").lower()
                    kind = "image" if mime in MEDIA_MIMES["image"] else "audio" if mime in MEDIA_MIMES["audio"] else ""
                    if not kind or info.file_size > MEDIA_LIMITS[kind]:
                        continue
                    media_content = archive.read(info)
                    if _valid_media(kind, mime, media_content):
                        archive_files[safe_name] = (kind, mime, media_content)
        except (zipfile.BadZipFile, RuntimeError, NotImplementedError):
            raise HTTPException(400, "Media arxivi ZIP formatida emas")
    errors = []
    imported = 0
    set_cache: dict[tuple[str, str], int] = {}
    media_cache: dict[tuple[int, str], int] = {}
    db = await get_pool()
    async with db.acquire() as conn:
        async with conn.transaction():
            for index, row in enumerate(rows, 2):
                created_set_key = None
                created_media_key = None
                try:
                    set_data, q_data = _row_models(row)
                    _validate_set(set_data)
                    _validate_bank_question(set_data.section, q_data)
                    content_hash = _question_hash(set_data.section, q_data, f"{set_data.title}\n{set_data.passage}")
                    if await conn.fetchval(
                        "SELECT 1 FROM question_bank_questions WHERE content_hash=$1", content_hash
                    ):
                        raise ValueError("Takroriy savol")
                    media_filename = _clean_text(row.get("media_filename"), 180)
                    archive_item = archive_files.get(media_filename) if media_filename else None
                    if media_filename and not archive_item:
                        raise ValueError(f"ZIP ichida media topilmadi: {media_filename}")
                    key = (set_data.section, set_data.title.casefold())
                    async with conn.transaction():
                        set_id = set_cache.get(key)
                        if not set_id:
                            set_id = await conn.fetchval(
                                """
                                INSERT INTO question_bank_sets(center_id,section,title,instructions,passage,topic,difficulty,
                                  band_min,band_max,source_name,source_url,license_note,created_by)
                                VALUES($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13) RETURNING id
                                """, None if current_user.get("is_admin") else current_user["center_id"], set_data.section,
                                set_data.title, set_data.instructions, set_data.passage, set_data.topic, set_data.difficulty,
                                set_data.band_min, set_data.band_max, set_data.source_name, set_data.source_url,
                                set_data.license_note, current_user.get("id"),
                            )
                            set_cache[key] = set_id
                            created_set_key = key
                        media_id = None
                        if media_filename:
                            media_key = (set_id, media_filename)
                            media_id = media_cache.get(media_key)
                            if not media_id:
                                kind, mime, media_content = archive_item
                                media_id = await conn.fetchval(
                                    """INSERT INTO question_bank_media(set_id,kind,original_filename,mime_type,file_data,file_size)
                                       VALUES($1,$2,$3,$4,$5,$6) RETURNING id""",
                                    set_id, kind, media_filename, mime, media_content, len(media_content),
                                )
                                media_cache[media_key] = media_id
                                created_media_key = media_key
                        await conn.execute(
                            """
                            INSERT INTO question_bank_questions(set_id,question_type,prompt,options,correct_answer,points,
                              explanation,media_id,skill,tags,difficulty,band_min,band_max,content_hash,created_by)
                            VALUES($1,$2,$3,$4::jsonb,$5::jsonb,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15)
                            """, set_id, q_data.question_type, q_data.prompt,
                            json.dumps(q_data.options, ensure_ascii=False),
                            json.dumps(q_data.correct_answer, ensure_ascii=False) if q_data.correct_answer is not None else None,
                            q_data.points, q_data.explanation, media_id, q_data.skill, q_data.tags, q_data.difficulty,
                            q_data.band_min, q_data.band_max, content_hash, current_user.get("id"),
                        )
                    imported += 1
                except Exception as exc:
                    if created_set_key:
                        set_cache.pop(created_set_key, None)
                    if created_media_key:
                        media_cache.pop(created_media_key, None)
                    message = exc.detail if isinstance(exc, HTTPException) else str(exc)
                    errors.append({"row": index, "message": message})
            suffix = filename.lower().rsplit(".", 1)[-1]
            import_id = await conn.fetchval(
                """
                INSERT INTO question_bank_imports(center_id,uploaded_by,original_filename,file_format,status,
                  total_rows,imported_rows,errors)
                VALUES($1,$2,$3,$4,$5,$6,$7,$8::jsonb) RETURNING id
                """, None if current_user.get("is_admin") else current_user["center_id"], current_user.get("id"),
                filename[:180], suffix, "completed_with_errors" if errors else "completed", len(rows), imported,
                json.dumps(errors, ensure_ascii=False),
            )
    return {"id": import_id, "total": len(rows), "imported": imported, "error_count": len(errors), "errors": errors[:200]}


@router.get("/imports")
async def import_history(current_user: dict = Depends(get_current_head_teacher)):
    center_id = None if current_user.get("is_admin") else current_user["center_id"]
    db = await get_pool()
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """SELECT id,original_filename,file_format,status,total_rows,imported_rows,errors,created_at
               FROM question_bank_imports WHERE ($1::integer IS NULL OR center_id=$1)
               ORDER BY created_at DESC LIMIT 100""", center_id,
        )
    return [dict(row) for row in rows]


@router.get("/export")
async def export_question_bank(current_user: dict = Depends(get_current_head_teacher)):
    center_id = None if current_user.get("is_admin") else current_user["center_id"]
    db = await get_pool()
    async with db.acquire() as conn:
        rows = await conn.fetch(
            """
            SELECT s.title AS set_title,s.section,s.instructions,s.passage,s.topic,s.difficulty AS set_difficulty,
                   s.band_min AS set_band_min,s.band_max AS set_band_max,s.source_name,s.source_url,s.license_note,
                   s.status AS set_status,q.question_type,q.prompt,q.options,q.correct_answer,q.points,q.explanation,
                   m.original_filename AS media_filename,q.skill,q.tags,q.difficulty,q.band_min,q.band_max,
                   q.status,q.version,q.usage_count,q.last_used_at
            FROM question_bank_sets s JOIN question_bank_questions q ON q.set_id=s.id
            LEFT JOIN question_bank_media m ON m.id=q.media_id
            WHERE ($1::integer IS NULL OR s.center_id=$1 OR (s.center_id IS NULL AND s.status='approved'))
            ORDER BY s.section,s.title,q.id
            """, center_id,
        )
    wb = Workbook()
    ws = wb.active
    ws.title = "Savollar banki"
    headers = [
        "set_title","section","instructions","passage","topic","set_difficulty","set_band_min","set_band_max",
        "source_name","source_url","license_note","set_status","question_type","prompt","options","correct_answer",
        "points","explanation","media_filename","skill","tags","difficulty","band_min","band_max","status",
        "version","usage_count","last_used_at",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1A56E8")
        cell.font = Font(color="FFFFFF", bold=True)
    for row in rows:
        values = []
        for key in headers:
            value = row[key]
            if key in {"options", "correct_answer"}:
                value = _decoded(value, value)
                value = "|".join(map(str, value)) if isinstance(value, list) else value
            elif key == "tags":
                value = "|".join(value or [])
            elif isinstance(value, datetime):
                value = value.isoformat(sep=" ", timespec="seconds")
            values.append(value)
        ws.append(values)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in ("A","C","D","N","R"):
        ws.column_dimensions[column].width = 34 if column != "D" else 60
    stream = io.BytesIO()
    wb.save(stream)
    return Response(
        stream.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="ielts-question-bank-export.xlsx"'},
    )


async def _copy_set_to_test(conn, set_id: int, test_id: int, section: str) -> tuple[int, int]:
    set_row = await conn.fetchrow("SELECT * FROM question_bank_sets WHERE id=$1", set_id)
    if not set_row or set_row["section"] != section:
        raise HTTPException(400, "To'plam bo'limga mos emas")
    section_id = await conn.fetchval(
        "SELECT id FROM test_builder_sections WHERE test_id=$1 AND section=$2", test_id, section
    )
    if not section_id:
        raise HTTPException(404, "Test bo'limi topilmadi")
    if await conn.fetchval(
        "SELECT 1 FROM test_builder_question_groups WHERE section_id=$1 AND bank_set_id=$2", section_id, set_id
    ):
        raise HTTPException(409, "Bu to'plam testga oldin qo'shilgan")
    media_map = {}
    bank_media = await conn.fetch("SELECT * FROM question_bank_media WHERE set_id=$1 ORDER BY id", set_id)
    for media in bank_media:
        media_map[media["id"]] = await conn.fetchval(
            """INSERT INTO test_builder_media(test_id,kind,original_filename,mime_type,file_data,file_size)
               VALUES($1,$2,$3,$4,$5,$6) RETURNING id""",
            test_id, media["kind"], media["original_filename"], media["mime_type"], media["file_data"], media["file_size"],
        )
    group_media_bank_id = next((media["id"] for media in bank_media if section == "listening" and media["kind"] == "audio"), None)
    group_media = media_map.get(group_media_bank_id)
    group_order = await conn.fetchval(
        "SELECT COALESCE(MAX(sort_order),-1)+1 FROM test_builder_question_groups WHERE section_id=$1", section_id
    )
    group_id = await conn.fetchval(
        """INSERT INTO test_builder_question_groups(section_id,title,instructions,passage,media_id,bank_set_id,sort_order)
           VALUES($1,$2,$3,$4,$5,$6,$7) RETURNING id""",
        section_id, set_row["title"], set_row["instructions"], set_row["passage"], group_media, set_id, group_order,
    )
    questions = await conn.fetch(
        "SELECT * FROM question_bank_questions WHERE set_id=$1 AND status<>'retired' ORDER BY id", set_id
    )
    start_order = await conn.fetchval(
        "SELECT COALESCE(MAX(sort_order),-1)+1 FROM test_builder_questions WHERE section_id=$1", section_id
    )
    for offset, question in enumerate(questions):
        await conn.execute(
            """
            INSERT INTO test_builder_questions(section_id,question_type,prompt,options,correct_answer,points,
              explanation,media_id,sort_order,bank_question_id,bank_question_version,group_id)
            VALUES($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12)
            """, section_id, question["question_type"], question["prompt"], question["options"],
            question["correct_answer"], question["points"], question["explanation"],
            None if question["media_id"] == group_media_bank_id else media_map.get(question["media_id"]),
            start_order + offset, question["id"], question["version"], group_id,
        )
    await conn.execute(
        "UPDATE question_bank_sets SET usage_count=usage_count+1,last_used_at=NOW() WHERE id=$1", set_id
    )
    await conn.execute(
        "UPDATE question_bank_questions SET usage_count=usage_count+1,last_used_at=NOW() WHERE set_id=$1", set_id
    )
    return group_id, len(questions)


@router.post("/tests/{test_id}/sections/{section}/sets/{set_id}")
async def attach_bank_set(
    test_id: int, section: str, set_id: int, current_user: dict = Depends(get_current_head_teacher)
):
    if section not in SECTIONS:
        raise HTTPException(400, "Bo'lim noto'g'ri")
    db = await get_pool()
    async with db.acquire() as conn:
        async with conn.transaction():
            await _owned_test(conn, test_id, current_user, editable=True)
            set_row = await _set_row(conn, set_id, current_user)
            if set_row["status"] not in {"approved", "draft"}:
                raise HTTPException(400, "Faqat tasdiqlangan yoki o'zingizning draft to'plamingizni qo'shing")
            group_id, count = await _copy_set_to_test(conn, set_id, test_id, section)
    return {"group_id": group_id, "question_count": count, "message": "To'plam testga qo'shildi"}


@router.post("/tests/{test_id}/submit-review")
async def submit_builder_test_review(
    test_id: int, current_user: dict = Depends(get_current_head_teacher)
):
    db = await get_pool()
    async with db.acquire() as conn:
        await _owned_test(conn, test_id, current_user, editable=True)
        section_count = await conn.fetchval(
            """SELECT COUNT(*) FROM test_builder_sections s WHERE s.test_id=$1
               AND EXISTS(SELECT 1 FROM test_builder_questions q WHERE q.section_id=s.id)""", test_id
        )
        if not section_count:
            raise HTTPException(400, "Testda kamida bitta savolli bo'lim bo'lishi kerak")
        row = await conn.fetchrow(
            "UPDATE tests SET status='pending',updated_at=NOW() WHERE id=$1 RETURNING title", test_id
        )
        await notify_admin(
            conn, "Konstruktor testi tekshiruvga yuborildi",
            f"{row['title']} testini ko'rib chiqish va publish qilish kerak.",
            kind="task", action_url="/admin", metadata={"test_id": test_id},
        )
    return {"status": "pending", "message": "Test super-adminga tasdiqlash uchun yuborildi"}


@router.post("/tests/{test_id}/save-to-bank")
async def save_builder_test_to_bank(
    test_id: int, current_user: dict = Depends(get_current_head_teacher)
):
    db = await get_pool()
    created_sets = 0
    created_questions = 0
    duplicates = 0
    async with db.acquire() as conn:
        async with conn.transaction():
            test = await _owned_test(conn, test_id, current_user)
            sections = await conn.fetch(
                "SELECT * FROM test_builder_sections WHERE test_id=$1 ORDER BY sort_order", test_id
            )
            for section in sections:
                groups = await conn.fetch(
                    "SELECT * FROM test_builder_question_groups WHERE section_id=$1 ORDER BY sort_order,id",
                    section["id"],
                )
                sources = []
                for group in groups:
                    if group["bank_set_id"]:
                        continue
                    questions = await conn.fetch(
                        "SELECT * FROM test_builder_questions WHERE group_id=$1 ORDER BY sort_order,id", group["id"]
                    )
                    if questions:
                        sources.append((group["title"], group["instructions"], group["passage"], group["media_id"], questions))
                ungrouped = await conn.fetch(
                    """SELECT * FROM test_builder_questions WHERE section_id=$1 AND group_id IS NULL
                       AND bank_question_id IS NULL ORDER BY sort_order,id""", section["id"]
                )
                if ungrouped:
                    sources.append((section["title"], section["instructions"], section["passage"], None, ungrouped))
                for title, instructions, passage, group_media_id, questions in sources:
                    set_id = await conn.fetchval(
                        """
                        INSERT INTO question_bank_sets(center_id,section,title,instructions,passage,topic,difficulty,
                          source_name,license_note,created_by)
                        VALUES($1,$2,$3,$4,$5,$6,'medium',$7,$8,$9) RETURNING id
                        """, None if current_user.get("is_admin") else current_user["center_id"], section["section"],
                        f"{test['title']} · {title}"[:180], instructions, passage, test["title"][:160],
                        "Mavjud test konstruktori", "Ichki test materialidan bankka ko'chirilgan",
                        current_user.get("id"),
                    )
                    created_sets += 1
                    media_ids = {q["media_id"] for q in questions if q["media_id"]}
                    if group_media_id:
                        media_ids.add(group_media_id)
                    media_map = {}
                    if media_ids:
                        for media in await conn.fetch(
                            "SELECT * FROM test_builder_media WHERE id=ANY($1::integer[])", list(media_ids)
                        ):
                            media_map[media["id"]] = await conn.fetchval(
                                """INSERT INTO question_bank_media(set_id,kind,original_filename,mime_type,file_data,file_size)
                                   VALUES($1,$2,$3,$4,$5,$6) RETURNING id""",
                                set_id, media["kind"], media["original_filename"], media["mime_type"],
                                media["file_data"], media["file_size"],
                            )
                    for question in questions:
                        qdata = BankQuestionIn(
                            question_type=question["question_type"], prompt=question["prompt"],
                            options=_decoded(question["options"], []),
                            correct_answer=_decoded(question["correct_answer"], None), points=float(question["points"]),
                            explanation=question["explanation"] or "", media_id=media_map.get(question["media_id"]),
                            difficulty="medium",
                        )
                        content_hash = _question_hash(
                            section["section"], qdata, f"{test['title']} · {title}\n{passage or ''}"
                        )
                        if await conn.fetchval(
                            "SELECT 1 FROM question_bank_questions WHERE content_hash=$1", content_hash
                        ):
                            duplicates += 1
                            continue
                        await conn.execute(
                            """
                            INSERT INTO question_bank_questions(set_id,question_type,prompt,options,correct_answer,points,
                              explanation,media_id,difficulty,content_hash,created_by)
                            VALUES($1,$2,$3,$4,$5,$6,$7,$8,'medium',$9,$10)
                            """, set_id, question["question_type"], question["prompt"], question["options"],
                            question["correct_answer"], question["points"], question["explanation"],
                            media_map.get(question["media_id"]), content_hash, current_user.get("id"),
                        )
                        created_questions += 1
                    if not await conn.fetchval(
                        "SELECT 1 FROM question_bank_questions WHERE set_id=$1", set_id
                    ):
                        await conn.execute("DELETE FROM question_bank_sets WHERE id=$1", set_id)
                        created_sets -= 1
    return {
        "sets": created_sets, "questions": created_questions, "duplicates": duplicates,
        "message": "Mavjud konstruktor materiallari savollar bankiga saqlandi",
    }


@router.post("/tests/{test_id}/auto-build")
async def auto_build_test(
    test_id: int, data: AutoBuildIn, current_user: dict = Depends(get_current_head_teacher)
):
    if data.difficulty and data.difficulty not in DIFFICULTIES:
        raise HTTPException(400, "Qiyinlik noto'g'ri")
    if not 0 <= data.avoid_recent_days <= 3650:
        raise HTTPException(400, "Takrorlanmaslik muddati noto'g'ri")
    for section, target in data.sections.items():
        if section not in SECTIONS or not 0 <= target <= 200:
            raise HTTPException(400, "Blueprint noto'g'ri")
    added = {}
    shortages = {}
    recent_fallback = {}
    center_id = None if current_user.get("is_admin") else current_user["center_id"]
    db = await get_pool()
    async with db.acquire() as conn:
        async with conn.transaction():
            await _owned_test(conn, test_id, current_user, editable=True)
            for section, target in data.sections.items():
                if target <= 0:
                    continue
                existing = await conn.fetchval(
                    """SELECT COUNT(*) FROM test_builder_questions q JOIN test_builder_sections s ON s.id=q.section_id
                       WHERE s.test_id=$1 AND s.section=$2""", test_id, section,
                )
                needed = max(0, target - existing)
                if not needed:
                    added[section] = 0
                    continue
                candidates = await conn.fetch(
                    """
                    SELECT s.id,COUNT(q.id) AS question_count FROM question_bank_sets s
                    JOIN question_bank_questions q ON q.set_id=s.id AND q.status='approved'
                    WHERE s.section=$1 AND s.status='approved'
                      AND ($2::integer IS NULL OR s.center_id IS NULL OR s.center_id=$2)
                      AND ($3::text IS NULL OR s.difficulty=$3)
                      AND ($4='' OR COALESCE(s.topic,'') ILIKE '%'||$4||'%')
                      AND ($5=0 OR s.last_used_at IS NULL OR s.last_used_at < NOW()-($5||' days')::interval)
                      AND NOT EXISTS(
                        SELECT 1 FROM test_builder_question_groups g JOIN test_builder_sections bs ON bs.id=g.section_id
                        WHERE bs.test_id=$6 AND g.bank_set_id=s.id
                      )
                    GROUP BY s.id,s.usage_count,s.last_used_at
                    ORDER BY s.usage_count ASC,s.last_used_at NULLS FIRST,RANDOM()
                    """, section, center_id, data.difficulty, data.topic.strip(), data.avoid_recent_days, test_id,
                )
                section_added = 0
                for candidate in candidates:
                    _, count = await _copy_set_to_test(conn, candidate["id"], test_id, section)
                    section_added += count
                    if section_added >= needed:
                        break
                if section_added < needed and data.avoid_recent_days:
                    fallback_candidates = await conn.fetch(
                        """
                        SELECT s.id,COUNT(q.id) AS question_count FROM question_bank_sets s
                        JOIN question_bank_questions q ON q.set_id=s.id AND q.status='approved'
                        WHERE s.section=$1 AND s.status='approved'
                          AND ($2::integer IS NULL OR s.center_id IS NULL OR s.center_id=$2)
                          AND ($3::text IS NULL OR s.difficulty=$3)
                          AND ($4='' OR COALESCE(s.topic,'') ILIKE '%'||$4||'%')
                          AND NOT EXISTS(
                            SELECT 1 FROM test_builder_question_groups g JOIN test_builder_sections bs ON bs.id=g.section_id
                            WHERE bs.test_id=$5 AND g.bank_set_id=s.id
                          )
                        GROUP BY s.id,s.usage_count,s.last_used_at
                        ORDER BY s.usage_count ASC,s.last_used_at ASC NULLS FIRST,RANDOM()
                        """, section, center_id, data.difficulty, data.topic.strip(), test_id,
                    )
                    fallback_count = 0
                    for candidate in fallback_candidates:
                        _, count = await _copy_set_to_test(conn, candidate["id"], test_id, section)
                        section_added += count
                        fallback_count += count
                        if section_added >= needed:
                            break
                    if fallback_count:
                        recent_fallback[section] = fallback_count
                added[section] = section_added
                if section_added < needed:
                    shortages[section] = needed - section_added
    return {
        "added": added, "shortages": shortages, "recent_fallback": recent_fallback,
        "message": "Test blueprint asosida yig'ildi",
    }


@router.get("/analytics")
async def bank_analytics(current_user: dict = Depends(get_current_head_teacher)):
    center_id = None if current_user.get("is_admin") else current_user["center_id"]
    db = await get_pool()
    async with db.acquire() as conn:
        summary = await conn.fetch(
            """
            SELECT s.section,s.status,COUNT(DISTINCT s.id) AS set_count,COUNT(q.id) AS question_count
            FROM question_bank_sets s LEFT JOIN question_bank_questions q ON q.set_id=s.id
            WHERE ($1::integer IS NULL OR s.center_id=$1 OR (s.center_id IS NULL AND s.status='approved'))
            GROUP BY s.section,s.status ORDER BY s.section,s.status
            """, center_id,
        )
        mappings = await conn.fetch(
            """
            SELECT bq.id AS builder_id,bq.bank_question_id,bs.test_id,bs.section,qb.prompt,qb.correct_answer,
                   qb.question_type,qb.usage_count
            FROM test_builder_questions bq JOIN test_builder_sections bs ON bs.id=bq.section_id
            JOIN question_bank_questions qb ON qb.id=bq.bank_question_id
            JOIN question_bank_sets qs ON qs.id=qb.set_id
            WHERE ($1::integer IS NULL OR qs.center_id IS NULL OR qs.center_id=$1)
            """, center_id,
        )
        results = await conn.fetch(
            """SELECT test_id,section,answers FROM exam_results
               WHERE test_id IS NOT NULL AND answers IS NOT NULL ORDER BY id DESC LIMIT 5000"""
        )
        due_sets = await conn.fetch(
            """
            SELECT id,title,section,status,planned_publish_at,review_due_at
            FROM question_bank_sets
            WHERE ($1::integer IS NULL OR center_id=$1 OR (center_id IS NULL AND status='approved'))
              AND status<>'retired'
              AND ((review_due_at IS NOT NULL AND review_due_at<=NOW()+INTERVAL '14 days')
                OR (planned_publish_at IS NOT NULL AND planned_publish_at<=NOW()+INTERVAL '30 days'))
            ORDER BY LEAST(COALESCE(review_due_at,'infinity'::timestamp),COALESCE(planned_publish_at,'infinity'::timestamp))
            LIMIT 50
            """, center_id,
        )
    map_by_test = defaultdict(list)
    stats = defaultdict(lambda: {"attempts": 0, "correct": 0})
    metadata = {}
    for item in mappings:
        map_by_test[(item["test_id"], item["section"])].append(item)
        metadata[item["bank_question_id"]] = item
    for result in results:
        answers = _decoded(result["answers"], {})
        for item in map_by_test.get((result["test_id"], result["section"]), []):
            key = str(item["builder_id"])
            if key not in answers:
                continue
            expected = _decoded(item["correct_answer"], None)
            submitted = answers[key]
            record = stats[item["bank_question_id"]]
            record["attempts"] += 1
            if isinstance(expected, list):
                left = sorted(_normal(value) for value in (submitted if isinstance(submitted, list) else [submitted]))
                right = sorted(_normal(value) for value in expected)
                correct = left == right
            else:
                correct = _normal(submitted) == _normal(expected)
            record["correct"] += int(correct)
    flags = []
    for question_id, record in stats.items():
        if record["attempts"] < 5:
            continue
        accuracy = record["correct"] / record["attempts"]
        if accuracy <= .2 or accuracy >= .95:
            item = metadata[question_id]
            flags.append({
                "question_id": question_id, "prompt": item["prompt"][:180],
                "attempts": record["attempts"], "accuracy": round(accuracy * 100, 1),
                "flag": "juda_qiyin" if accuracy <= .2 else "juda_oson",
            })
    approved_counts = defaultdict(int)
    for row in summary:
        if row["status"] == "approved":
            approved_counts[row["section"]] += row["question_count"]
    targets = {"listening": 80, "reading": 80, "writing": 10, "speaking": 20}
    low_stock = [
        {"section": section, "approved": approved_counts[section], "target": target,
         "missing": max(0, target - approved_counts[section])}
        for section, target in targets.items() if approved_counts[section] < target
    ]
    return {
        "summary": [dict(row) for row in summary], "low_stock": low_stock,
        "quality_flags": flags[:100], "due_sets": [dict(row) for row in due_sets],
    }
