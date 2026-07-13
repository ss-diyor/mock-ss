"""Teacher va head-teacher natija eksportlari uchun umumiy yordamchilar."""

from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime
from typing import Iterable, Mapping, Any

from fastapi import HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from scoring import get_band_score


EXPORT_COLUMNS = [
    ("result_id", "Result ID"),
    ("group_name", "Guruh"),
    ("full_name", "Student"),
    ("email", "Email"),
    ("section", "Bo'lim"),
    ("score", "To'g'ri javob"),
    ("total", "Jami savol"),
    ("band", "Band"),
    ("task_achievement", "Task Achievement / Response"),
    ("coherence_cohesion", "Coherence & Cohesion"),
    ("lexical_resource", "Lexical Resource"),
    ("grammar_accuracy", "Grammar Range & Accuracy"),
    ("writing_feedback", "Writing feedback"),
    ("graded_by", "Baholagan teacher"),
    ("graded_at", "Baholangan vaqt"),
    ("submitted_at", "Topshirilgan vaqt"),
]

_DANGEROUS_CELL_PREFIXES = ("=", "+", "-", "@")


def _safe_cell(value: Any) -> Any:
    """CSV/XLSX formula injection xavfini kamaytiradi."""
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ") if isinstance(value, datetime) else value.isoformat()
    if isinstance(value, (int, float)):
        return value
    text = str(value)
    if text.startswith(_DANGEROUS_CELL_PREFIXES):
        return "'" + text
    return text


def _band_for(row: Mapping[str, Any]) -> float | None:
    if row["section"] == "writing":
        value = row["writing_band"]
        return float(value) if value is not None else None
    score, total = row["score"], row["total"]
    if score is None or total is None:
        return None
    value = get_band_score(score, total, row["section"])
    return float(value) if value is not None else None


def normalise_export_rows(rows: Iterable[Mapping[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for row in rows:
        result.append(
            {
                "result_id": row["id"],
                "group_name": row["group_name"],
                "full_name": row["full_name"],
                "email": row["email"],
                "section": row["section"],
                "score": row["score"],
                "total": row["total"],
                "band": _band_for(row),
                "task_achievement": row["writing_task_achievement"],
                "coherence_cohesion": row["writing_coherence_cohesion"],
                "lexical_resource": row["writing_lexical_resource"],
                "grammar_accuracy": row["writing_grammar_accuracy"],
                "writing_feedback": row["writing_feedback"],
                "graded_by": row["grader_name"],
                "graded_at": row["writing_graded_at"],
                "submitted_at": row["submitted_at"],
            }
        )
    return result


def _filename(prefix: str, extension: str) -> str:
    safe_prefix = re.sub(r"[^a-zA-Z0-9_-]+", "-", prefix).strip("-") or "results"
    return f"{safe_prefix}-{datetime.utcnow():%Y-%m-%d}.{extension}"


def build_results_export(
    rows: Iterable[Mapping[str, Any]],
    export_format: str,
    filename_prefix: str,
) -> StreamingResponse:
    fmt = export_format.lower().strip()
    if fmt == "excel":
        fmt = "xlsx"
    if fmt not in {"csv", "xlsx"}:
        raise HTTPException(status_code=400, detail="Format faqat csv, excel yoki xlsx bo'lishi mumkin")

    records = normalise_export_rows(rows)
    headers = [title for _, title in EXPORT_COLUMNS]

    if fmt == "csv":
        text_buffer = io.StringIO(newline="")
        writer = csv.writer(text_buffer)
        writer.writerow(headers)
        for record in records:
            writer.writerow([_safe_cell(record[key]) for key, _ in EXPORT_COLUMNS])
        payload = io.BytesIO(text_buffer.getvalue().encode("utf-8-sig"))
        media_type = "text/csv; charset=utf-8"
        filename = _filename(filename_prefix, "csv")
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Natijalar"
        sheet.append(headers)
        header_fill = PatternFill("solid", fgColor="DCE7FF")
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for record in records:
            sheet.append([_safe_cell(record[key]) for key, _ in EXPORT_COLUMNS])
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for index, column in enumerate(EXPORT_COLUMNS, start=1):
            key, title = column
            max_length = len(title)
            for record in records[:1000]:
                max_length = max(max_length, len(str(_safe_cell(record[key]))))
            sheet.column_dimensions[get_column_letter(index)].width = min(max_length + 2, 48)
        payload = io.BytesIO()
        workbook.save(payload)
        payload.seek(0)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = _filename(filename_prefix, "xlsx")

    payload.seek(0)
    return StreamingResponse(
        payload,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
